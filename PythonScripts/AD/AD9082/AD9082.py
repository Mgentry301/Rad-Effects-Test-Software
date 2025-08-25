import threading
import queue
import time
from openpyxl import Workbook
from datetime import datetime
from threading import Lock
import concurrent.futures
from PythonScripts.radbench import Keithley2230 as keithley
from PythonScripts.FieldFox.MaxCapture import capture_spectrum_worker
from PythonScripts.AD.AD9082.API.ad9082_python_eval_app import txfe_main

# Shared variable and lock
total_capture_count = 0
total_reg_count = 0
capture_count_lock = Lock()
latest_supply_data = None
supply_data_lock = threading.Lock()


supply_labels=[    
    'k0 ch1 V', 'k0 ch2 V', 'k0 ch3 V',
    'k0 ch1 I', 'k0 ch2 I', 'k0 ch3 I',
    'k1 ch1 V', 'k1 ch2 V', 'k1 ch3 V',
    'k1 ch1 I', 'k1 ch2 I', 'k1 ch3 I',
    'k2 ch1 V', 'k2 ch2 V', 'k2 ch3 V',
    'k2 ch1 I', 'k2 ch2 I', 'k2 ch3 I',
    'k3 ch1 V', 'k3 ch2 V', 'k3 ch3 V',
    'k3 ch1 I', 'k3 ch2 I', 'k3 ch3 I',          
]

def save_workbook_async(workbook, filename):
    def save():
        workbook.save(filename)
    threading.Thread(target=save).start()

class supplies:
    supply_list = []
    data_labels=[    
        'k0 ch1 V', 'k0 ch2 V', 'k0 ch3 V',
        'k0 ch1 I', 'k0 ch2 I', 'k0 ch3 I',
        'k1 ch1 V', 'k1 ch2 V', 'k1 ch3 V',
        'k1 ch1 I', 'k1 ch2 I', 'k1 ch3 I',
        'k2 ch1 V', 'k2 ch2 V', 'k2 ch3 V',
        'k2 ch1 I', 'k2 ch2 I', 'k2 ch3 I',
        'adc ch1 V', 'adc ch2 V', 
        'adc ch1 I', 'adc ch2 I',                
    ]
    supply_values = []
    def __init__(self):
        k0 = keithley('USB0::0x05E6::0x2230::9201602::INSTR')    
        k1 = keithley('USB0::0x05E6::0x2230::9030930::INSTR')
        k2 = keithley('USB0::0x05E6::0x2230::9201548::INSTR')
        k3 = keithley('USB0::0x05E6::0x2230::9100252::INSTR')
        self.supply_list = [k0, k1, k2, k3]  
        
    def configure_supplies(self):
        """
        Configure each supply and channel to a voltage between 1 and 1.5V
        with the maximum allowed current (6A for Keithley2230).
        """
        k0_v = [2, 0, 1]
        k1_v = [3.3, 1, 1]
        k2_v = [1, 2, 1]
        k3_v = [1, 1, 1]
        currents = [1.5, 1.5, 5]

        voltage_lists = [k0_v, k1_v, k2_v, k3_v]

        for supply, v_list in zip(self.supply_list, voltage_lists):
            for ch in range(1, 4):  # CH1, CH2, CH3
                supply.set_volt_and_curr(f"CH{ch}", v_list[ch-1], currents[ch-1])
            supply.on()
    
    def disable_supplies(self):
        """
        Turn off all supplies.
        """
        for supply in self.supply_list:
            supply.off()
         

    def get_supplies(self):
        data = []

        def read_supply(supply):
            voltages = supply.meas_volt("ALL")
            currents = supply.meas_curr("ALL")
            return voltages + currents

        with concurrent.futures.ThreadPoolExecutor() as executor:
            results = list(executor.map(read_supply, self.supply_list))
        for result in results:
            data.extend(result)
        return data

def supply_reader_thread(supplies, stop_event):
    global latest_supply_data
    all_supplies = supplies
    interval = 0.02  # 20ms for 50Hz
    while not stop_event.is_set():
        start = time.perf_counter()
        data = all_supplies.get_supplies()
        with supply_data_lock:
            latest_supply_data = data
        elapsed = time.perf_counter() - start
        sleep_time = max(0, interval - elapsed)
        time.sleep(sleep_time)

def data_reader(data_queue, stop_event, run_type):
    """
    Reads data continuously and stores it in a queue.
    Tracks the number of register reads and data captures.
    Logs the counts every 10 seconds.
    """
    global total_capture_count
    global total_reg_count
    last_capture_count = 0
    last_reg_count = 0
    start_time = time.time()    
    interval = 0.02  # 20ms for 50Hz

    while not stop_event.is_set():
        loop_start = time.perf_counter()
        with supply_data_lock:
            supply_data = latest_supply_data
        if supply_data is not None:
            data_queue.put(supply_data)

        with capture_count_lock:
            total_capture_count += 1    

        # Log performance every sec_interval seconds
        sec_interval = 10
        if time.time() - start_time >= sec_interval:
            if run_type ==1: captures_last_x_sec = total_capture_count - last_capture_count
            else: captures_last_x_sec = total_reg_count - last_reg_count

            print(f"Avg Reads in last {sec_interval} sec: {captures_last_x_sec/sec_interval}")

            last_capture_count = total_capture_count
            last_reg_count = total_reg_count
            start_time = time.time()

        if data_queue.qsize() > 900:
            print(f"Data queue is nearing capacity! Size is {data_queue.qsize()}")

        elapsed = time.perf_counter() - loop_start
        sleep_time = max(0, interval - elapsed)
        time.sleep(sleep_time)

def data_logger(data_queue, stop_event, run_number, freq_queue=None):
    """
    Logs data from the queue into an Excel file with timestamps.
    """
    # Create an Excel workbook and sheets
    workbook = Workbook()
    capture_sheet = workbook.active
    capture_sheet.title = "Capture Data"
    supply_sheet = workbook.create_sheet(title = "Supply Data")
    for col_idx, value in enumerate(supply_labels, start=2):  # Write data vertically
        supply_sheet.cell(row=1, column=col_idx, value=value)

    buffer = []
    buffer_size = 10  # Adjust buffer size as needed
    supply_row = 2
    cap_row = 2

    # Wait for frequency axis if freq_queue is provided
    freq_labels = None
    if freq_queue is not None:
        try:
            # Wait up to 10 seconds for freq axis
            freq_labels = freq_queue.get(timeout=10)
        except queue.Empty:
            freq_labels = None

    #Write capture sheet header row
    if freq_labels is not None:
        capture_sheet.cell(row=1, column=1, value="Timestamp")
        for col_idx, freq in enumerate(freq_labels, start=2):
            capture_sheet.cell(row=1, column=col_idx, value=f"{freq:.6f} GHz")
        cap_row = 2
    else:
        capture_sheet.cell(row=1, column=1, value="Timestamp")
        capture_sheet.cell(row=1, column=2, value="Amplitude Data")
        cap_row = 2

    while not stop_event.is_set() or not data_queue.empty():
        try:
            # Retrieve data from the queue
            data = data_queue.get(timeout=0.1)
            buffer.append(data)
            
            # Process or log data in batches
            # Save the workbook less frequently
            if len(buffer) >= buffer_size:
                for item in buffer:
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
                    if len(item)==24: #supply data
                        supply_sheet.cell(row=supply_row, column=1, value=timestamp)  # Timestamp in col 1
                        for col_idx, value in enumerate(item, start=2):  # Write data vertically
                            supply_sheet.cell(row=supply_row, column=col_idx, value=value)
                        supply_row += 1
                    else:  # Capture data
                        capture_sheet.cell(row=cap_row, column=1, value=timestamp)  # Timestamp in row 1
                        for col_idx, value in enumerate(item.tolist(), start=2):  # Write data vertically
                            capture_sheet.cell(row=cap_row, column=col_idx, value=value)
                        cap_row += 1
                buffer.clear()

                # Save the workbook every x amount of entries instead of every batch
                if total_capture_count % 5000 ==0:
                    print(f"total_capture_count:{total_capture_count} total_reg_count: {total_reg_count}")
                    print("Saving workbook...")
                    save_workbook_async(workbook, f"Run_{run_number}_data_log.xlsx")
                    print("Workbook saved")
        except queue.Empty:
            continue

    # Final save when the logger stops

    if len(buffer) >= buffer_size:
        for item in buffer:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
            if len(item)==24: #supply data
                supply_sheet.cell(row=supply_row, column=1, value=timestamp)  # Timestamp in col 1
                for col_idx, value in enumerate(item, start=2):  # Write data vertically
                    supply_sheet.cell(row=supply_row, column=col_idx, value=value)
                supply_row += 1
            else:  # Capture data
                capture_sheet.cell(row=cap_row, column=1, value=timestamp)  # Timestamp in row 1
                for col_idx, value in enumerate(item.tolist(), start=2):  # Write data vertically
                    capture_sheet.cell(row=cap_row, column=col_idx, value=value)
                cap_row += 1
    buffer.clear()
    print("Last save completed")
    save_workbook_async(workbook, f"Run_{run_number}_data_log.xlsx")


if __name__ == '__main__':
    run_number = input('\nEnter the run number: ')
    run_type = input("Enter the run type (1 for SEL, anything else for SEU): ")
    print("powering up ad9082 & configring system")

    # Configure supplies before starting threads
    supplies = supplies()
    supplies.configure_supplies()

    # Configure and initialize AD9082
    ads9, ad9082, uc, args = txfe_main.init()

    # Create a queue for buffering data
    data_queue = queue.Queue(maxsize=1000)
    stop_event = threading.Event()
    freq_queue = queue.Queue()  # For frequency axis from FieldFox

    input('press enter to start reccording')

    # Start threads
    supply_thread = threading.Thread(target=supply_reader_thread, args=(supplies, stop_event))
    logger_thread = threading.Thread(target=data_logger, args=(data_queue, stop_event, run_number, freq_queue))
    reader_thread = threading.Thread(target=data_reader, args=(data_queue, stop_event, 1))

    supply_thread.start()
    logger_thread.start()
    reader_thread.start()

    # Only start capture thread if run_type is not "1"
    if run_type != "1":        
        capture_thread = threading.Thread(
            target=capture_spectrum_worker,
            args=(data_queue, stop_event),
            kwargs={'freq_queue': freq_queue}
        )
        capture_thread.start()
    else:
        capture_thread = None

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        x = 0
        while x != -1:
            x = input('Enter 1 to reconfigure the AD9082, 2 to Power Cycle and reconfigure, -1 to power down: ')
        
        supplies.disable_supplies()
        print("supplies disabled")

        print("Stopping threads...")       
        stop_event.set()
        # Join all threads that use the supplies before closing them!
        supply_thread.join()
        logger_thread.join()
        reader_thread.join()
        if capture_thread is not None:
            capture_thread.join()
        print("Threads stopped.")
