"""
Multi-instrument GUI: Keithley 2230 panels + Keysight EL34243A panel.
- Add instruments by typing SN or full VISA resource.
- Choose instrument type on add (Keithley or Keysight).
- Scan VISA to detect connected devices.
- Each instrument gets its own tab and independent controls.

Notes:
- Keysight EL34243A wrapper uses common SCPI patterns for electronic loads
  (INP ON/OFF, FUNC:MODE <CURR|VOLT|RES>, CURR/VOLT/RES <value>, MEAS:CURR?).
  If your model uses different commands, the wrapper will report errors and
  can be adjusted easily.
"""
import sys
import os
import json
import tempfile
import datetime
import time
from typing import List, Tuple
 
from PyQt5 import QtWidgets, QtCore
import pyvisa

from Support_Scrips.power_sequence_builder import PowerSequenceBuilder
from Support_Scrips.Front_Panels.keithley_panel import KeithleyPanel
from Support_Scrips.Front_Panels.keysightEL_panel import KeysightELPanel
from Support_Scrips.Front_Panels.keysight_e36233a_panel import KeysightE36233APanel
from Support_Scrips.Front_Panels.hittite_siggen_panel import HittiteSigGenPanel
from Support_Scrips.Front_Panels.rhodeschwarz_sma_panel import RhodeSchwarzSMAPanel
from Support_Scrips.Front_Panels.fieldfox_sa_panel import FieldFoxSAPanel

class MainWindow(QtWidgets.QMainWindow):
    # Thread-safe log signal
    log_signal = QtCore.pyqtSignal(str)
    # --- Internal helpers to reduce duplication ---
    def _keithley_apply_settings(self, panel: KeithleyPanel, on: bool):
        if not getattr(panel, 'inst', None):
            return
        for ch in (1, 2, 3):
            try:
                if on:
                    V = float(panel.vol_edits[ch].text())
                    I = float(panel.iam_edits[ch].text())
                    panel.inst.set_voltage(ch, V) 
                    panel.inst.set_current(ch, I)
                    panel.inst.set_output(ch, True)
                    panel.output_btns[ch].setChecked(True)
                    panel.output_btns[ch].setText('Output On')
                    panel.output_btns[ch].setStyleSheet('background-color: #4CAF50; color: white;')
                else:
                    panel.inst.set_output(ch, False)
                    panel.output_btns[ch].setChecked(False)
                    panel.output_btns[ch].setText('Output Off')
                    panel.output_btns[ch].setStyleSheet('background-color: #F44336; color: white;')
            except Exception:
                pass
        panel.master_out_btn.setChecked(on)
        panel.master_out_btn.setText('All On' if on else 'All Off')

    def _keysight_el_apply_settings(self, panel: KeysightELPanel, on: bool):
        if not getattr(panel, 'dev', None):
            return
        for ch in (1, 2):
            try:
                mode = (panel.mode_combo_ch1.currentText() if ch == 1 else panel.mode_combo_ch2.currentText())
                if mode == 'Disable':
                    panel.dev.set_input(ch, False)
                    if hasattr(panel, '_set_input_toggle_ui'):
                        panel._set_input_toggle_ui(ch, False)
                    continue
                if on:
                    # Read per-channel UI controls
                    try:
                        value = float(panel.mode_value_ch1.text() if ch == 1 else panel.mode_value_ch2.text())
                    except Exception:
                        value = 0.0
                    panel.dev.set_mode(ch, mode)
                    ramp_enabled = panel.ramp_enable_ch1.isChecked() if ch == 1 else panel.ramp_enable_ch2.isChecked()
                    try:
                        rise_time = float(panel.rise_time_ch1.text() if ch == 1 else panel.rise_time_ch2.text())
                    except Exception:
                        rise_time = 0.0
                    panel.dev.set_input(ch, True)
                    if ramp_enabled and rise_time > 0:
                        import numpy as np
                        import time
                        steps = 20
                        for v in np.linspace(0, value, steps):
                            panel.dev.set_parameter(ch, mode, v)
                            QtWidgets.QApplication.processEvents()
                            time.sleep(rise_time / steps)
                        panel.dev.set_parameter(ch, mode, value)
                    else:
                        panel.dev.set_parameter(ch, mode, value)
                else:
                    panel.dev.set_input(ch, False)
            except Exception:
                pass
        # Update per-channel UI toggles if available
        try:
            if hasattr(panel, '_set_input_toggle_ui'):
                panel._set_input_toggle_ui(1, on)
                panel._set_input_toggle_ui(2, on)
        except Exception:
            pass

    def _generic_output_toggle(self, panel, on: bool):
        dev = getattr(panel, 'dev', None)
        if not dev:
            return
        try:
            dev.set_output(on)
            if hasattr(panel, 'output_btn'):
                panel.output_btn.setChecked(on)
                panel.output_btn.setText('Output On' if on else 'Output Off')
            if hasattr(panel, 'status_label'):
                panel.status_label.setText('Output ON' if on else 'Output OFF')
        except Exception:
            pass
    
    def on_record_clicked(self):
        """Unified record button: start/stop recording for selected metrics."""
        # If not primed yet, behave like legacy flow: ask path and start now
        if self.record_btn.isChecked():
            if getattr(self, '_record_primed', False):
                # Fire start for already-primed threads
                try:
                    if getattr(self, '_record_start_event', None):
                        self._record_start_event.set()
                    # Nudge sync time to now for any loops using _record_sync_start
                    self._record_sync_start = time.time()
                except Exception:
                    pass
                self.record_btn.setText('Stop Recording')
                self.statusBar().showMessage('Recording started', 3000)
                return
            excel_path = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Reads', '', 'Excel Files (*.xlsx)')[0]
            if excel_path and not excel_path.lower().endswith('.xlsx'):
                excel_path += '.xlsx'
            if not excel_path:
                return
            # Remember path for cleanup (e.g., removing any backups on stop)
            try:
                self._last_excel_path = excel_path
            except Exception:
                pass
            # Create a global excel write lock once per session
            try:
                import threading
                if not hasattr(self, '_excel_lock') or self._excel_lock is None:
                    self._excel_lock = threading.Lock()
            except Exception:
                self._excel_lock = None
            # Reset shared workbook state for this recording session
            try:
                self._excel_wb = None
                self._excel_ws_cache = {}
                # Working/final paths + markers
                self._excel_final_path = excel_path
                self._excel_working_path = excel_path + '.working.xlsx'
                self._excel_marker_path = excel_path + '.saving'
                self._excel_wb_path = self._excel_working_path
                self._excel_last_save = 0.0
                self._excel_last_snapshot = 0.0
                self._excel_saving = False
            except Exception:
                pass
            # Create saving marker
            try:
                with open(self._excel_marker_path, 'w') as f:
                    f.write('saving')
            except Exception:
                pass
            # Establish a synchronized start time a short interval in the future so all recorders align
            try:
                self._record_sync_start = time.time() + 1.5  # 1.5s lead time
            except Exception:
                self._record_sync_start = None
            # Write one-time bench summary sheet at recording start
            try:
                self._write_bench_info(excel_path)
            except Exception as e:
                # Non-fatal if we cannot write the bench info
                try:
                    self._log(f'Bench Info write failed: {e}')
                except Exception:
                    pass
            # Start recording selected metrics
            if self.supply_record_toggle.isChecked():
                panels = self.get_all_supply_panels()
                if not panels:
                    self.statusBar().showMessage('No supply panels to record', 4000)
                    # Do not return; continue to other selected metrics (e.g., Spectrum)
                self.global_recorder = self._create_supply_recorder(panels, excel_path)
                self.global_recorder.start()
                self.statusBar().showMessage('Started recording supplies', 4000)
            if self.spectrum_record_toggle.isChecked():
                # Excel lock already created at beginning of recording if needed
                # Continuous Spectrum Recording
                import threading
                fieldfox_panel = None
                for i in range(self.tabs.count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel):
                        fieldfox_panel = w
                        break
                if fieldfox_panel is None:
                    self.statusBar().showMessage('No FieldFox panel found', 4000)
                    # Skip spectrum but continue to other metrics
                    fieldfox_panel = None
                if fieldfox_panel is not None:
                    def update_spectrum_read_speed(sps: float):
                        self.spectrum_read_speed_label.setText(f'Spectrum Sample Rate: {sps:.2f} samples/sec')
                    self._start_spectrum_capture(fieldfox_panel, excel_path, update_spectrum_read_speed)
                    self._spectrum_recording = True
                    self._spectrum_thread_running = True
                    def running_flag():
                        return self._spectrum_recording and self._spectrum_thread_running
                    # Thread was created inside helper
                    self.statusBar().showMessage('Started recording spectrum', 4000)
            if self.register_record_toggle.isChecked():
                # Register Recording using ACE Client; read list from current config if available
                registers = []
                try:
                    if hasattr(self, 'register_read_array') and self.register_read_array:
                        registers = self.register_read_array
                    else:
                        # Fallback: try to read from current selection in configs dropdown
                        cfg_name = self.load_combo.currentText() if hasattr(self, 'load_combo') else ''
                        if cfg_name:
                            cfg_path = os.path.join(self.configs_dir, cfg_name)
                            if os.path.exists(cfg_path):
                                with open(cfg_path, 'r') as f:
                                    data = json.load(f)
                                    registers = data.get('register_read_array', [])
                except Exception:
                    registers = []
                if not registers:
                    QtWidgets.QMessageBox.warning(self, 'No registers configured', 'No register_read_array found in current config. Load a config that defines it.')
                else:
                    def update_register_read_speed(sps: float):
                        try:
                            self.register_read_speed_label.setText(f'Register Sample Rate: {sps:.2f} samples/sec')
                        except Exception:
                            pass
                    self._start_register_recording(excel_path, registers, update_register_read_speed)
                    self.statusBar().showMessage('Started recording registers', 4000)
            self.record_btn.setText('Stop Recording')
            # Start/ensure save status timer
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
        else:
            # Stop all recordings
            self._stop_all_recordings()
            self.record_btn.setText('Record')
            self.statusBar().showMessage('Stopped recording', 4000)
            # Reset primed state and enable toggles
            try:
                self._record_primed = False
                self.record_btn.setEnabled(False)
                for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                    cb.setEnabled(True)
                if hasattr(self, 'prime_btn'):
                    self.prime_btn.setText('Prime Recorders')
                    self.prime_btn.setStyleSheet('')
            except Exception:
                pass

    def _create_supply_recorder(self, panels, excel_path):
        """Factory for a PatchedSupplyRecorder with background writer and 25Hz throttle."""
        from supply_recorder import SupplyRecorder
        self_parent = self

        def update_supply_read_speed(sps: float):
            try:
                self_parent.supply_read_speed_label.setText(f'Supply Sample Rate: {sps:.2f} samples/sec')
            except Exception:
                pass

        class PatchedSupplyRecorder(SupplyRecorder):
            def __init__(self, panels, excel_path, sheet_name='supply reads'):
                super().__init__(panels, excel_path, sheet_name)
                import queue
                self._write_q = queue.Queue(maxsize=100000)
                self._writer_stop = None
                self._writer_thread = None
                self._header = self._build_header(panels)

            def _build_header(self, panels):
                # Build header to match row order: timestamp, all voltages (all panels), then all currents (all panels)
                ts = ['timestamp']
                v_headers = []
                i_headers = []
                for p in panels:
                    try:
                        tab_idx = self_parent.tabs.indexOf(p)
                        alias_name = self_parent.tabs.tabText(tab_idx) if tab_idx >= 0 else getattr(p, 'resource', 'Supply')
                    except Exception:
                        alias_name = getattr(p, 'resource', 'Supply')
                    if p.__class__.__name__.startswith('Keithley'):
                        ch_count = 3
                        try:
                            ch_names = [p.ch_name_edits[i].text() if hasattr(p, 'ch_name_edits') and p.ch_name_edits.get(i) else f'CH{i}' for i in (1, 2, 3)]
                        except Exception:
                            ch_names = [f'CH{i}' for i in (1, 2, 3)]
                    else:
                        ch_count = 2
                        ch_names = [f'CH{i}' for i in (1, 2)]
                    for i in range(1, ch_count + 1):
                        nm = ch_names[i - 1] if i - 1 < len(ch_names) else f'CH{i}'
                        v_headers.append(f'{alias_name}_{nm}_V')
                for p in panels:
                    try:
                        tab_idx = self_parent.tabs.indexOf(p)
                        alias_name = self_parent.tabs.tabText(tab_idx) if tab_idx >= 0 else getattr(p, 'resource', 'Supply')
                    except Exception:
                        alias_name = getattr(p, 'resource', 'Supply')
                    if p.__class__.__name__.startswith('Keithley'):
                        ch_count = 3
                        try:
                            ch_names = [p.ch_name_edits[i].text() if hasattr(p, 'ch_name_edits') and p.ch_name_edits.get(i) else f'CH{i}' for i in (1, 2, 3)]
                        except Exception:
                            ch_names = [f'CH{i}' for i in (1, 2, 3)]
                    else:
                        ch_count = 2
                        ch_names = [f'CH{i}' for i in (1, 2)]
                    for i in range(1, ch_count + 1):
                        nm = ch_names[i - 1] if i - 1 < len(ch_names) else f'CH{i}'
                        i_headers.append(f'{alias_name}_{nm}_I')
                return ts + v_headers + i_headers

            def start(self):
                import threading, time
                self._writer_stop = threading.Event()

                def writer_loop():
                    lock = getattr(self_parent, '_excel_lock', None)
                    acquired = lock.acquire(timeout=5) if lock else True
                    if acquired:
                        try:
                            self_parent._excel_open_locked(self.excel_path)
                            ws = self_parent._excel_get_sheet_locked(self.sheet_name)
                            for ci, val in enumerate(self._header, start=1):
                                ws.cell(row=1, column=ci, value=val)
                            self_parent._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                    pending = []
                    last_save = time.time()
                    flush_rows = 1200
                    while True:
                        try:
                            item = self._write_q.get(timeout=0.3)
                            pending.append(item)
                            self._write_q.task_done()
                        except Exception:
                            pass
                        now = time.time()
                        should_save = pending and (len(pending) >= flush_rows or (now - last_save) >= 8.0)
                        should_stop = self._writer_stop.is_set() and self._write_q.empty()
                        if should_save or should_stop:
                            acquired = lock.acquire(timeout=10) if lock else True
                            if acquired:
                                try:
                                    self_parent._excel_open_locked(self.excel_path)
                                    ws = self_parent._excel_get_sheet_locked(self.sheet_name)
                                    for row in pending:
                                        ws.append(row)
                                    pending.clear()
                                    last_save = now
                                    self_parent._excel_save_locked()
                                finally:
                                    if lock and acquired:
                                        try:
                                            lock.release()
                                        except Exception:
                                            pass
                        if should_stop and not pending:
                            break

                self._writer_thread = threading.Thread(target=writer_loop, daemon=True)
                self._writer_thread.start()
                return super().start()

            def stop(self):
                try:
                    self._stop_event.set()
                    if self._thread:
                        self._thread.join()
                except Exception:
                    pass
                try:
                    if self._writer_stop:
                        self._writer_stop.set()
                    if self._writer_thread:
                        self._writer_thread.join(timeout=10)
                except Exception:
                    pass

            def _run(self):
                import time, datetime
                sample_count = 0
                # Wait for prime event then sync time if configured
                try:
                    evt = getattr(self_parent, '_record_start_event', None)
                    if evt is not None:
                        while not evt.is_set() and not self._stop_event.is_set():
                            time.sleep(0.005)
                except Exception:
                    pass
                if getattr(self_parent, '_record_sync_start', None):
                    while time.time() < self_parent._record_sync_start and not self._stop_event.is_set():
                        time.sleep(0.005)
                start_time = time.time()
                last_report_time = start_time
                target_interval = getattr(self_parent, '_supply_target_interval', 0.04)
                next_tick = time.perf_counter()
                while not self._stop_event.is_set():
                    nowp = time.perf_counter()
                    if nowp < next_tick:
                        time.sleep(min(next_tick - nowp, 0.01))
                        continue
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    all_voltages = []
                    all_currents = []
                    for func in self.get_readings_funcs:
                        voltages, currents = func()
                        all_voltages.extend(voltages)
                        all_currents.extend(currents)
                    row = [now] + all_voltages + all_currents
                    try:
                        self._write_q.put_nowait(row)
                    except Exception:
                        pass
                    sample_count += 1
                    now_ts = time.time()
                    if now_ts - last_report_time >= 2.0:
                        elapsed = max(now_ts - start_time, 1e-6)
                        sps = sample_count / elapsed
                        update_supply_read_speed(sps)
                        last_report_time = now_ts
                    next_tick += target_interval
                    if next_tick < nowp - target_interval:
                        next_tick = nowp + target_interval

        return PatchedSupplyRecorder(panels, excel_path, sheet_name='supply reads')

    def on_prime_clicked(self):
        """Prepare selected recorders: choose file, init Excel, and spin threads waiting for start."""
        try:
            # If already primed and user clicks again, unprime/tear-down
            if getattr(self, '_record_primed', False):
                # Don't allow unprime while an active recording is running
                try:
                    if hasattr(self, 'record_btn') and self.record_btn.isChecked():
                        self.statusBar().showMessage('Recording is running; stop it before unpriming', 4000)
                        return
                except Exception:
                    pass
                self._unprime_recorders()
                return
            excel_path = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Reads', '', 'Excel Files (*.xlsx)')[0]
            if excel_path and not excel_path.lower().endswith('.xlsx'):
                excel_path += '.xlsx'
            if not excel_path:
                return
            import threading
            # Create excel lock if needed and reset shared workbook state
            if not hasattr(self, '_excel_lock') or self._excel_lock is None:
                self._excel_lock = threading.Lock()
            self._excel_wb = None
            self._excel_ws_cache = {}
            self._excel_final_path = excel_path
            self._excel_working_path = excel_path + '.working.xlsx'
            self._excel_marker_path = excel_path + '.saving'
            self._excel_wb_path = self._excel_working_path
            self._excel_last_save = 0.0
            self._excel_last_snapshot = 0.0
            self._excel_saving = False
            self._last_excel_path = excel_path
            try:
                with open(self._excel_marker_path, 'w') as f:
                    f.write('saving')
            except Exception:
                pass
            # Create a start event and sync time slightly in future
            self._record_start_event = threading.Event()
            self._record_sync_start = time.time() + 1.0
            # Bench info immediately
            try:
                self._write_bench_info(excel_path)
            except Exception:
                pass
            # Start selected recorders but they will wait for event/sync
            if self.supply_record_toggle.isChecked():
                panels = self.get_all_supply_panels()
                if panels:
                    self.global_recorder = self._create_supply_recorder(panels, excel_path)
                    self.global_recorder.start()
            if self.spectrum_record_toggle.isChecked():
                fieldfox_panel = None
                for i in range(self.tabs.count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel):
                        fieldfox_panel = w
                        break
                if fieldfox_panel is not None:
                    def update_spectrum_read_speed(sps: float):
                        self.spectrum_read_speed_label.setText(f'Spectrum Sample Rate: {sps:.2f} samples/sec')
                    # Reuse existing helper; it respects _record_sync_start and will wait before loop
                    self._start_spectrum_capture(fieldfox_panel, excel_path, update_spectrum_read_speed)
            if self.register_record_toggle.isChecked():
                registers = []
                try:
                    if hasattr(self, 'register_read_array') and self.register_read_array:
                        registers = self.register_read_array
                    else:
                        cfg_name = self.load_combo.currentText() if hasattr(self, 'load_combo') else ''
                        if cfg_name:
                            cfg_path = os.path.join(self.configs_dir, cfg_name)
                            if os.path.exists(cfg_path):
                                with open(cfg_path, 'r') as f:
                                    data = json.load(f)
                                    registers = data.get('register_read_array', [])
                except Exception:
                    registers = []
                if registers:
                    def update_register_read_speed(sps: float):
                        try:
                            self.register_read_speed_label.setText(f'Register Sample Rate: {sps:.2f} samples/sec')
                        except Exception:
                            pass
                    self._start_register_recording(excel_path, registers, update_register_read_speed)
            # Update UI state
            self._record_primed = True
            self.record_btn.setEnabled(True)
            for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                cb.setEnabled(False)
            if hasattr(self, 'prime_btn'):
                self.prime_btn.setText('Recorders Primed (Click to Unprime)')
                self.prime_btn.setStyleSheet('background-color: #4CAF50; color: white;')
                try:
                    self.prime_btn.setToolTip('Click to unprime and tear down prepared recorders')
                except Exception:
                    pass
            self.statusBar().showMessage('Recorders primed. Press Record to start.', 4000)
            # Start/ensure save status timer
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Prime failed', str(e))

    def _unprime_recorders(self):
        """Tear down any primed recorders and reset UI back to pre-primed state."""
        try:
            # Safety: if recording is actually running, do nothing here
            if hasattr(self, 'record_btn') and self.record_btn.isChecked():
                return
        except Exception:
            pass
        # Stop any background primed threads and clean up workbook markers/files
        try:
            # Discard primed artifacts (no final save) when unpriming
            self._stop_all_recordings(finalize=False)
        except Exception:
            pass
        # Reset primed state and UI
        try:
            self._record_primed = False
            # Clear start sync/event so future primes create fresh ones
            try:
                self._record_start_event = None
                self._record_sync_start = None
            except Exception:
                pass
            if hasattr(self, 'record_btn'):
                try:
                    self.record_btn.setChecked(False)
                except Exception:
                    pass
                self.record_btn.setText('Record')
                self.record_btn.setEnabled(False)
            for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                try:
                    cb.setEnabled(True)
                except Exception:
                    pass
            if hasattr(self, 'prime_btn'):
                try:
                    self.prime_btn.setText('Prime Recorders')
                    self.prime_btn.setStyleSheet('')
                    self.prime_btn.setToolTip('Prepare selected recorders and Excel file; enables Start')
                except Exception:
                    pass
            self.statusBar().showMessage('Recorders unprimed', 3000)
        except Exception:
            pass

    def _stop_all_recordings(self, finalize: bool = True):
        """Stop supply, spectrum, and register recordings and reset labels.
        finalize=True: flush/save workbook and snapshot final file
        finalize=False: discard primed working file/marker without saving
        """
        # Supply
        try:
            if hasattr(self, 'global_recorder') and self.global_recorder:
                try:
                    self.global_recorder.stop()
                except Exception:
                    pass
                # Final guarded flush (only when finalizing)
                if finalize:
                    try:
                        lock = getattr(self, '_excel_lock', None)
                        if lock and lock.acquire(timeout=5):
                            try:
                                if getattr(self.global_recorder, 'buffer', None):
                                    self.global_recorder._flush_buffer()
                            finally:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                    except Exception:
                        pass
                self.global_recorder = None
            if hasattr(self, 'supply_read_speed_label'):
                self.supply_read_speed_label.setText('Supply Sample Rate: --')
        except Exception:
            pass
        # Spectrum
        try:
            if hasattr(self, '_spectrum_thread') and self._spectrum_thread:
                self._spectrum_recording = False
                self._spectrum_thread_running = False
                try:
                    self._spectrum_thread.join(timeout=2)
                except Exception:
                    pass
                self._spectrum_thread = None
            # Resume panel streaming if available
            try:
                for i in range(getattr(self, 'tabs', QtWidgets.QTabWidget()).count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel) and hasattr(w, 'toggle_streaming'):
                        if not w.streaming_enabled:
                            # Re-enable streaming and start capture thread
                            w.stream_btn.setChecked(True)
                            w.toggle_streaming(True)
            except Exception:
                pass
        except Exception:
            pass
        # Registers
        try:
            if hasattr(self, '_register_thread') and getattr(self, '_register_thread', None):
                try:
                    self._register_recording = False
                except Exception:
                    pass
                try:
                    self._register_thread_running = False
                except Exception:
                    pass
                try:
                    self._register_thread.join(timeout=2)
                except Exception:
                    pass
                self._register_thread = None
            if hasattr(self, 'register_read_speed_label'):
                self.register_read_speed_label.setText('Register Sample Rate: --')
        except Exception:
            pass
        # Finalize or discard workbook artifacts
        try:
            # For finalize stops, give writer queues a short window to drain before the last save
            if finalize:
                try:
                    import time as _t
                    t0 = _t.time()
                    while (_t.time() - t0) < 2.0:
                        try:
                            pending = self._get_total_pending_writes()
                        except Exception:
                            pending = 0
                        if pending <= 0:
                            break
                        _t.sleep(0.1)
                except Exception:
                    pass
            if finalize:
                lock = getattr(self, '_excel_lock', None)
                if lock and lock.acquire(timeout=5):
                    try:
                        # Force save and snapshot
                        self._excel_save_locked(force=True)
                    finally:
                        try:
                            lock.release()
                        except Exception:
                            pass
            # Try to remove marker and working file with small retries; keep status updates active until done
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
            ok = self._cleanup_excel_artifacts(finalize=finalize, timeout_s=3.0)
            # Stop save status timer after cleanup
            try:
                if hasattr(self, '_save_status_timer'):
                    self._save_status_timer.stop()
            except Exception:
                pass
            # If discard, also clear more paths so a stale state cannot persist
            if not finalize:
                try:
                    self._excel_final_path = None
                    self._excel_working_path = None
                    self._excel_marker_path = None
                except Exception:
                    pass
        except Exception:
            pass

    def _start_register_recording(self, excel_path: str, registers: list, rate_cb):
        """Start register recording in background with a non-blocking Excel writer."""
        import threading, time, datetime, os, queue
        from openpyxl import Workbook, load_workbook

        # Normalize register list (accept decimal, hex str, int)
        reg_list: list[int] = []
        for r in registers:
            try:
                if isinstance(r, str):
                    reg_list.append(int(r, 0))  # auto-detect base (0x, decimal)
                else:
                    reg_list.append(int(r))
            except Exception:
                continue
        if not reg_list:
            QtWidgets.QMessageBox.warning(self, 'No registers', 'No registers specified to record.')
            return

        sheet_name = 'register reads'
        start_time = time.time()
        last_report_time = start_time
        sample_count = 0

        # Prepare workbook / header immediately using cached helpers
        try:
            header = ['timestamp'] + [f'{addr:#x}' for addr in reg_list]
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    self._excel_open_locked(excel_path)
                    ws = self._excel_get_sheet_locked(sheet_name)
                    for ci, val in enumerate(header, start=1):
                        ws.cell(row=1, column=ci, value=val)
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception:
            pass

        # Background writer using a queue; avoids blocking acquisition during wb.save
        write_q: "queue.Queue[list]" = queue.Queue(maxsize=50000)
        # expose for UI save status
        try:
            self._register_write_q = write_q
        except Exception:
            pass
        writer_stop = threading.Event()
        flush_interval = 150
        header = ['timestamp'] + [f'{addr:#x}' for addr in reg_list]

        def writer_loop():
            pending: list[list] = []
            last_save = time.time()
            while True:
                try:
                    item = write_q.get(timeout=0.3)
                    pending.append(item)
                    write_q.task_done()
                except queue.Empty:
                    pass
                now = time.time()
                should_save = pending and (len(pending) >= flush_interval or (now - last_save) >= 8.0)
                should_stop = writer_stop.is_set() and write_q.empty()
                if should_save or should_stop:
                    lock = getattr(self, '_excel_lock', None)
                    acquired = lock.acquire(timeout=10) if lock else True
                    if acquired:
                        try:
                            # Use shared helpers to write to working file and snapshot
                            self._excel_open_locked(excel_path)
                            ws = self._excel_get_sheet_locked(sheet_name)
                            try:
                                for ci, val in enumerate(header, start=1):
                                    ws.cell(row=1, column=ci, value=val)
                            except Exception:
                                pass
                            for row in pending:
                                ws.append(row)
                            pending.clear()
                            last_save = now
                            self._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                if should_stop and not pending:
                    break

        writer_thread = threading.Thread(target=writer_loop, daemon=True)
        writer_thread.start()

        def running_flag():
            return getattr(self, '_register_recording', False) and getattr(self, '_register_thread_running', False)

        def worker():
            nonlocal sample_count, last_report_time, start_time
            # Wait for start event if primed, then synchronized start time if provided
            try:
                evt = getattr(self, '_record_start_event', None)
                if evt is not None:
                    while not evt.is_set() and running_flag():
                        time.sleep(0.01)
                if getattr(self, '_record_sync_start', None):
                    while time.time() < self._record_sync_start and running_flag():
                        time.sleep(0.01)
            except Exception:
                pass
            # Connect to ACE server
            try:
                import clr  # type: ignore
                clr.AddReference('AnalogDevices.Csa.Remoting.Clients')
                clr.AddReference('AnalogDevices.Csa.Remoting.Contracts')
                from AnalogDevices.Csa.Remoting.Clients import ClientManager  # type: ignore
                manager = ClientManager.Create()
                client = manager.CreateRequestClient('localhost:2357')
            except Exception as e:
                try:
                    self._log(f'ACE connection failed: {e}')
                except Exception:
                    pass
                return
            while running_flag():
                try:
                    nowts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    values = []
                    for addr in reg_list:
                        try:
                            val = client.ReadRegister(str(int(addr)))
                            if hasattr(val, 'strip'):
                                val = val.strip('\r\n')
                        except Exception as err:
                            val = f'ERR:{err}'
                        values.append(val)
                    try:
                        write_q.put_nowait([nowts] + values)
                    except queue.Full:
                        pass
                    sample_count += 1
                    now_time = time.time()
                    if now_time - last_report_time >= 2.0:
                        elapsed = max(now_time - start_time, 1e-6)
                        try:
                            rate_cb(sample_count / elapsed)
                        except Exception:
                            pass
                        last_report_time = now_time
                    time.sleep(0.05)
                except Exception as e:
                    try:
                        self._log(f'Register read loop error: {e}')
                    except Exception:
                        pass
                    time.sleep(0.1)
            # Signal writer to finish and wait for it
            try:
                writer_stop.set()
            except Exception:
                pass
            try:
                writer_thread.join(timeout=10)
            except Exception:
                pass
            try:
                self._register_thread_running = False
            except Exception:
                pass

        # Start register reader thread
        self._register_recording = True
        self._register_thread_running = True
        self._register_thread = threading.Thread(target=worker, daemon=True)
        self._register_thread.start()

    def _excel_open_locked(self, excel_path: str):
        """Open or return a cached workbook for this session. Caller must hold _excel_lock."""
        try:
            import os
            from openpyxl import Workbook, load_workbook
            # Always map to working path if configured
            try:
                target_path = getattr(self, '_excel_working_path', None) or excel_path
            except Exception:
                target_path = excel_path
            if getattr(self, '_excel_wb', None) is None or getattr(self, '_excel_wb_path', None) != target_path:
                try:
                    if os.path.exists(target_path):
                        self._excel_wb = load_workbook(target_path)
                    else:
                        self._excel_wb = Workbook()
                except Exception:
                    self._excel_wb = Workbook()
                self._excel_ws_cache = {}
                self._excel_wb_path = target_path
            return self._excel_wb
        except Exception:
            return None

    def _excel_get_sheet_locked(self, sheet_name: str):
        """Get or create a worksheet; cached by name. Caller must hold _excel_lock."""
        try:
            if sheet_name in self._excel_ws_cache:
                return self._excel_ws_cache[sheet_name]
            wb = getattr(self, '_excel_wb', None)
            if wb is None:
                return None
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                if getattr(wb, 'active', None) is not None and len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1 and not wb.active['A1'].value:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(sheet_name)
            self._excel_ws_cache[sheet_name] = ws
            return ws
        except Exception:
            return None

    def _excel_save_locked(self, force: bool = False):
        """Save cached workbook to working file, and snapshot atomically to final.
        Caller holds _excel_lock.
        Policy: write working file frequently; snapshot final at least every 2.5 minutes
        to keep the final XLSX current and mitigate oversized working files.
        """
        try:
            import time, os, shutil
            wb = getattr(self, '_excel_wb', None)
            wpath = getattr(self, '_excel_wb_path', None)
            if not wb or not wpath:
                return
            now = time.time()
            last = getattr(self, '_excel_last_save', 0.0)
            # Save to working file (default every ~10s; force overrides)
            save_interval = 10.0
            if force or (now - last) >= save_interval:
                self._excel_saving = True
                wb.save(wpath)
                self._excel_last_save = now
            # Periodically snapshot to final using atomic replace
            fpath = getattr(self, '_excel_final_path', None)
            if fpath:
                last_snap = getattr(self, '_excel_last_snapshot', 0.0)
                # Snapshot at least every 150s (2.5 minutes), or immediately if forced.
                snapshot_interval = 150.0
                if force or ((now - last_snap) >= snapshot_interval):
                    tmp_final = fpath + '.tmp'
                    # Write a copy of working to tmp, then replace final
                    try:
                        shutil.copyfile(wpath, tmp_final)
                        os.replace(tmp_final, fpath)
                        self._excel_last_snapshot = now
                    except Exception:
                        try:
                            if os.path.exists(tmp_final):
                                os.remove(tmp_final)
                        except Exception:
                            pass
            self._excel_saving = False
        except Exception:
            try:
                self._excel_saving = False
            except Exception:
                pass
            pass

    def _get_total_pending_writes(self) -> int:
        """Return approximate total pending items across known writer queues."""
        total = 0
        try:
            if hasattr(self, 'global_recorder') and getattr(self.global_recorder, '_write_q', None):
                total += self.global_recorder._write_q.qsize()
        except Exception:
            pass
        try:
            if hasattr(self, '_spectrum_write_q') and getattr(self, '_spectrum_write_q', None):
                total += self._spectrum_write_q.qsize()
        except Exception:
            pass
        try:
            if hasattr(self, '_register_write_q') and getattr(self, '_register_write_q', None):
                total += self._register_write_q.qsize()
        except Exception:
            pass
        return total

    def _start_spectrum_capture(self, panel: FieldFoxSAPanel, excel_path: str, rate_cb):
        """Start spectrum capture in background with a non-blocking Excel writer."""
        import threading, time, datetime, queue
        from openpyxl import Workbook, load_workbook
        import os

        # If spectrum toggle not actually selected anymore, abort silently
        try:
            if not getattr(self, 'spectrum_record_toggle', None) or not self.spectrum_record_toggle.isChecked():
                return
        except Exception:
            return

        # Ensure flags are set before the worker starts
        self._spectrum_recording = True
        self._spectrum_thread_running = True

        # Try to ensure the FieldFox is connected (non-fatal if not yet)
        try:
            if getattr(panel.sa, 'inst', None) is None:
                panel.on_connect()
        except Exception:
            pass

        # Read initial frequency axis (may fail; we'll refresh later)
        try:
            freq = panel.sa.get_freq_axis(panel.unit_combo.currentText())
        except Exception:
            freq = []

        # Create/ensure workbook sheet and header up front using shared workbook
        try:
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    self._excel_open_locked(excel_path)
                    ws = self._excel_get_sheet_locked('capture data')
                    try:
                        n = len(freq)
                    except Exception:
                        n = 0
                    header = ['timestamp'] + ([f"{f:.6f}" for f in freq] if n > 0 else [])
                    for ci, val in enumerate(header, start=1):
                        ws.cell(row=1, column=ci, value=val)
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception:
            pass

        # Pause panel's live streaming to avoid interleaved SCPI during capture
        try:
            if hasattr(panel, 'stop_capture_thread'):
                panel.stop_capture_thread()
        except Exception:
            pass
        # Clear status safely (avoids raw clear race)
        try:
            if hasattr(panel.sa, 'sync'):
                panel.sa.sync()
        except Exception:
            pass

        # Background writer for spectrum rows
        try:
            n0 = len(freq)
        except Exception:
            n0 = 0
        desired_header = ['timestamp'] + ([f"{f:.6f}" for f in freq] if n0 > 0 else [])
        write_q: "queue.Queue[list]" = queue.Queue(maxsize=20000)
        try:
            self._spectrum_write_q = write_q
        except Exception:
            pass
        writer_stop = threading.Event()
        flush_interval = 200

        def spec_writer_loop():
            pending: list[list] = []
            last_save = time.time()
            while True:
                try:
                    item = write_q.get(timeout=0.3)
                    pending.append(item)
                    write_q.task_done()
                except queue.Empty:
                    pass
                now = time.time()
                should_save = pending and (len(pending) >= flush_interval or (now - last_save) >= 6.0)
                should_stop = writer_stop.is_set() and write_q.empty()
                if should_save or should_stop:
                    lock = getattr(self, '_excel_lock', None)
                    acquired = lock.acquire(timeout=10) if lock else True
                    if acquired:
                        try:
                            self._excel_open_locked(excel_path)
                            ws = self._excel_get_sheet_locked('capture data')
                            try:
                                for ci, val in enumerate(desired_header, start=1):
                                    ws.cell(row=1, column=ci, value=val)
                            except Exception:
                                pass
                            for row in pending:
                                ws.append(row)
                            pending.clear()
                            last_save = now
                            self._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                if should_stop and not pending:
                    break

        spec_writer_thread = threading.Thread(target=spec_writer_loop, daemon=True)
        spec_writer_thread.start()

        start_time = time.time()
        last_report_time = start_time
        sample_count = 0
        def running_flag():
            return getattr(self, '_spectrum_recording', False) and getattr(self, '_spectrum_thread_running', False)
        def worker():
            nonlocal sample_count, last_report_time, freq, desired_header
            import pyvisa
            # Wait for start event if primed, then synchronized start if defined
            try:
                evt = getattr(self, '_record_start_event', None)
                if evt is not None:
                    while not evt.is_set() and running_flag():
                        time.sleep(0.01)
                if getattr(self, '_record_sync_start', None):
                    while time.time() < self._record_sync_start and running_flag():
                        time.sleep(0.01)
            except Exception:
                pass
            while running_flag():
                try:
                    # Reconnect if session not open
                    if getattr(panel.sa, 'inst', None) is None:
                        try:
                            panel.on_connect()
                        except Exception:
                            pass
                        time.sleep(0.3)
                        continue
                    amplitudes = panel.sa.capture_spectrum()
                    # Refresh axis if needed
                    try:
                        n_freq = len(freq)
                    except Exception:
                        n_freq = 0
                    if not n_freq or (hasattr(amplitudes, '__len__') and len(amplitudes) != n_freq):
                        try:
                            freq = panel.sa.get_freq_axis(panel.unit_combo.currentText())
                            # Update header for writer to reflect new axis
                            try:
                                n1 = len(freq)
                            except Exception:
                                n1 = 0
                            desired_header = ['timestamp'] + ([f"{f:.6f}" for f in freq] if n1 > 0 else [])
                        except Exception:
                            pass
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    row = [now] + [float(a) for a in amplitudes]
                    try:
                        write_q.put_nowait(row)
                    except queue.Full:
                        pass
                    sample_count += 1
                except pyvisa.errors.InvalidSession:
                    break
                except Exception as e:
                    # Treat common transient VISA errors as recoverable
                    msg = ''
                    try:
                        msg = str(e.args[0]) if getattr(e, 'args', None) else str(e)
                    except Exception:
                        msg = str(e)
                    if any(tok in msg for tok in ['-410', 'Query Interrupted', '-420', 'Query UNTERMINATED', 'Query Unterminated']):
                        # Use safe sync instead of raw VISA clear to avoid -110 header errors
                        try:
                            if hasattr(panel.sa, 'sync'):
                                panel.sa.sync()
                        except Exception:
                            pass
                        time.sleep(0.05)
                        continue
                now_time = time.time()
                if now_time - last_report_time >= 2.0:
                    elapsed = now_time - start_time
                    rate_cb(sample_count / elapsed if elapsed > 0 else 0.0)
                    last_report_time = now_time
                time.sleep(0.1)
            # Finish writer
            try:
                writer_stop.set()
            except Exception:
                pass
            try:
                spec_writer_thread.join(timeout=10)
            except Exception:
                pass

        self._spectrum_thread = threading.Thread(target=worker, daemon=True)
        self._spectrum_thread.start()

    def _update_save_status(self):
        """Update status bar hinting whether Excel is still being saved; discourage opening file prematurely."""
        try:
            saving = bool(getattr(self, '_excel_saving', False))
            has_marker = False
            try:
                marker = getattr(self, '_excel_marker_path', '')
                has_marker = bool(marker) and os.path.exists(marker)
            except Exception:
                has_marker = False
            has_working = False
            try:
                wpath = getattr(self, '_excel_working_path', '')
                has_working = bool(wpath) and os.path.exists(wpath)
            except Exception:
                has_working = False
            pending = 0
            try:
                # Sum known queues
                if hasattr(self, 'global_recorder') and getattr(self.global_recorder, '_write_q', None):
                    pending += self.global_recorder._write_q.qsize()
                if hasattr(self, '_spectrum_write_q') and getattr(self, '_spectrum_write_q', None):
                    pending += self._spectrum_write_q.qsize()
                if hasattr(self, '_register_write_q') and getattr(self, '_register_write_q', None):
                    pending += self._register_write_q.qsize()
            except Exception:
                pass
            if saving or has_marker or has_working or pending > 0:
                self.statusBar().showMessage('Saving Excel data… please do not open the file yet', 2000)
            else:
                # show that file is ready briefly
                self.statusBar().showMessage('Excel file ready', 1500)
        except Exception:
            pass

    def _cleanup_excel_artifacts(self, finalize: bool, timeout_s: float = 3.0) -> bool:
        """Try to remove .saving marker and .working.xlsx file with small retries.
        Returns True if both (if present) are gone.
        If finalize is False, also clear in-memory workbook state.
        """
        import time
        end = time.time() + max(0.5, float(timeout_s))
        ok = False
        while time.time() < end:
            ok = True
            # Try remove marker
            try:
                marker = getattr(self, '_excel_marker_path', '')
                if marker and os.path.exists(marker):
                    try:
                        os.remove(marker)
                    except Exception:
                        ok = False
            except Exception:
                pass
            # Try remove working file
            try:
                wpath = getattr(self, '_excel_working_path', '')
                if wpath and os.path.exists(wpath):
                    try:
                        os.remove(wpath)
                    except Exception:
                        ok = False
            except Exception:
                pass
            if ok:
                break
            time.sleep(0.2)
        # Optionally clear in-memory workbook references to release any latent handles
        if not finalize:
            try:
                self._excel_wb = None
                self._excel_ws_cache = {}
                self._excel_wb_path = None
                # Intentionally keep final path so next prime can reuse, or clear if desired
            except Exception:
                pass
        return ok

    def _write_bench_info(self, excel_path: str):
        """Create or update a 'Bench Info' sheet with a summary of the bench equipment and settings."""
        from openpyxl import Workbook, load_workbook
        # Gather info from all tabs
        rows = []
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        def parse_serial_from_idn(idn: str) -> str:
            try:
                parts = [p.strip() for p in idn.replace(";", ",").split(',') if p.strip()]
                if len(parts) >= 3:
                    return parts[2]
                # fallback: look for SN or serial tokens
                for tok in parts:
                    if tok.upper().startswith('SN'):
                        return tok.split(':')[-1].strip()
            except Exception:
                pass
            return ''
        for i in range(self.tabs.count()):
            panel = self.tabs.widget(i)
            tab_name = self.tabs.tabText(i)
            # Keithley 2230
            if isinstance(panel, KeithleyPanel):
                inst = getattr(panel, 'inst', None)
                connected = bool(inst)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if inst:
                        idn = inst.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                # Setpoints
                settings = {
                    'ch1': {'V': getattr(panel.vol_edits.get(1), 'text', lambda: '0')()},
                    'ch2': {'V': getattr(panel.vol_edits.get(2), 'text', lambda: '0')()},
                    'ch3': {'V': getattr(panel.vol_edits.get(3), 'text', lambda: '0')()},
                }
                try:
                    settings['ch1']['I'] = panel.iam_edits[1].text()
                    settings['ch2']['I'] = panel.iam_edits[2].text()
                    settings['ch3']['I'] = panel.iam_edits[3].text()
                except Exception:
                    pass
                # Output state
                state_parts = []
                for ch in (1, 2, 3):
                    ch_on = None
                    if connected:
                        try:
                            ch_on = bool(inst.get_output_state(ch))
                        except Exception:
                            ch_on = None
                    state_parts.append(f'ch{ch}:{"on" if ch_on else ("off" if ch_on is not None else "?")}')
                state = ', '.join(state_parts)
                rows.append([ts, 'Keithley 2230', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            # Keysight E36233A
            elif isinstance(panel, KeysightE36233APanel):
                supply = getattr(panel, 'supply', None)
                connected = bool(supply)
                # Try to deduce resource from name field if nothing else is available
                resource = ''
                try:
                    resource = panel.name_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if supply and getattr(supply, 'inst', None):
                        idn = supply.inst.query("*IDN?")
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                settings = {'ch1': {}, 'ch2': {}}
                state_parts = []
                for ch in (1, 2):
                    try:
                        settings[f'ch{ch}']['V'] = str(supply.get_voltage_setpoint(ch)) if connected else ''
                    except Exception:
                        pass
                    try:
                        settings[f'ch{ch}']['I'] = str(supply.get_current_setpoint(ch)) if connected else ''
                    except Exception:
                        pass
                    on_val = None
                    try:
                        on_val = bool(supply.get_output_state(ch)) if connected else None
                    except Exception:
                        on_val = None
                    state_parts.append(f'ch{ch}:{"on" if on_val else ("off" if on_val is not None else "?")}')
                state = ', '.join(state_parts)
                rows.append([ts, 'Keysight E36233A', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            # Keysight EL34243A
            elif isinstance(panel, KeysightELPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                settings = {
                    'ch1': {
                        'mode': getattr(panel.mode_combo_ch1, 'currentText', lambda: '')(),
                        'value': getattr(panel.mode_value_ch1, 'text', lambda: '')(),
                    },
                    'ch2': {
                        'mode': getattr(panel.mode_combo_ch2, 'currentText', lambda: '')(),
                        'value': getattr(panel.mode_value_ch2, 'text', lambda: '')(),
                    },
                }
                state = f"ch1:{'on' if panel.input_toggle_ch1.isChecked() else 'off'}, ch2:{'on' if panel.input_toggle_ch2.isChecked() else 'off'}"
                rows.append([ts, 'Keysight EL34243A', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            # Hittite Sig Gen
            elif isinstance(panel, HittiteSigGenPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.freq_unit_combo, 'currentText', lambda: '')()
                settings = {
                    'frequency': getattr(panel.freq_edit, 'text', lambda: '')(),
                    'freq_unit': unit,
                    'power_dB': getattr(panel.pow_edit, 'text', lambda: '')(),
                }
                state = 'on' if getattr(panel.output_btn, 'isChecked', lambda: False)() else 'off'
                rows.append([ts, 'Hittite Sig Gen', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            # RhodeSchwarz SMA
            elif isinstance(panel, RhodeSchwarzSMAPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.freq_unit_combo, 'currentText', lambda: '')()
                settings = {
                    'frequency': getattr(panel.freq_edit, 'text', lambda: '')(),
                    'freq_unit': unit,
                    'power_dBm': getattr(panel.pow_edit, 'text', lambda: '')(),
                }
                state = 'on' if getattr(panel.output_btn, 'isChecked', lambda: False)() else 'off'
                rows.append([ts, 'RhodeSchwarz SMA', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            # Keysight FieldFox
            elif isinstance(panel, FieldFoxSAPanel):
                sa = getattr(panel, 'sa', None)
                connected = bool(getattr(sa, 'inst', None))
                resource = ''
                try:
                    resource = getattr(sa, 'visa_address', '')
                except Exception:
                    pass
                serial = ''
                try:
                    if connected:
                        try:
                            # prefer robust wrapper query if available
                            idn = sa._safe_query("*IDN?") if hasattr(sa, '_safe_query') else sa.inst.query("*IDN?")
                        except Exception:
                            idn = ''
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.unit_combo, 'currentText', lambda: '')()
                settings = {
                    'center': getattr(panel.center_edit, 'text', lambda: '')(),
                    'span': getattr(panel.span_edit, 'text', lambda: '')(),
                    'start': getattr(panel.start_edit, 'text', lambda: '')(),
                    'stop': getattr(panel.stop_edit, 'text', lambda: '')(),
                    'unit': unit,
                }
                # Capture thread as state
                running = bool(getattr(panel, '_capture_thread', None))
                state = 'capture:running' if running else 'capture:stopped'
                rows.append([ts, 'Keysight FieldFox', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
        # Write into Excel via cached helpers (working path) with retry strategy
        try:
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    wb = self._excel_open_locked(excel_path)
                    if wb is None:
                        return
                    if 'Bench Info' in wb.sheetnames:
                        ws = wb['Bench Info']
                        # Clear existing content
                        try:
                            for _ in range(ws.max_row, 0, -1):
                                ws.delete_rows(1)
                        except Exception:
                            pass
                    else:
                        ws = wb.create_sheet('Bench Info')
                    header = ['timestamp', 'instrument', 'name', 'resource', 'serial', 'connected', 'state', 'settings']
                    ws.append(header)
                    if rows:
                        for r in rows:
                            ws.append(r)
                    else:
                        ws.append([ts, '(none)', '(no instruments)', '', '', '', '', '{}'])
                        try:
                            QtCore.QTimer.singleShot(1000, lambda: self._write_bench_info(excel_path))
                        except Exception:
                            pass
                    # Remove default empty sheet if any
                    try:
                        if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1 and wb['Sheet'].max_column == 1 and not wb['Sheet']['A1'].value:
                            del wb['Sheet']
                    except Exception:
                        pass
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception as e:
            try:
                self._log(f'Bench Info write error: {e}')
            except Exception:
                pass

    def add_instrument_panel(self, resource, label, inst_type):
        """Add a new instrument panel to the tabs based on resource, label, and instrument type."""
        panel = None
        if inst_type == 'Keithley 2230':
            panel = KeithleyPanel(resource)
        elif inst_type == 'Keysight EL34243A':
            panel = KeysightELPanel(resource)
        elif inst_type == 'Keysight E36233A':
            panel = KeysightE36233APanel(resource)
        elif inst_type == 'Hittite Sig Gen':
            panel = HittiteSigGenPanel(resource)
        elif inst_type == 'RhodeSchwarz SMA':
            panel = RhodeSchwarzSMAPanel(resource)
        elif inst_type == 'Keysight FieldFox':
            panel = FieldFoxSAPanel(resource)
        else:
            panel = KeithleyPanel(resource)
        # Ensure a canonical 'resource' attribute exists for all panels so it can be saved
        try:
            if not hasattr(panel, 'resource') or not getattr(panel, 'resource'):
                # FieldFox panel stores address inside panel.sa.visa_address
                if isinstance(panel, FieldFoxSAPanel) and hasattr(panel, 'sa'):
                    panel.resource = panel.sa.visa_address
                else:
                    panel.resource = resource
        except Exception:
            pass
        self.tabs.addTab(panel, label)
        if hasattr(panel, 'resource_edit'):
            panel.resource_edit.setText(resource)
        if hasattr(panel, 'name_edit'):
            panel.name_edit.setText(label)
        if hasattr(panel, 'on_connect'):
            panel.on_connect()
        self._auto_turn_off_panel(panel)

    def get_all_supply_panels(self):
        panels = []
        for i in range(self.tabs.count()):
            w = self.tabs.widget(i)
            if isinstance(w, (KeithleyPanel, KeysightE36233APanel)):
                panels.append(w)
        return panels

    # Removed global recording buttons and logic; unified record button handles all recording
    # helper to format timestamps for logs
    def _ts(self):
        return datetime.datetime.now().strftime('%H:%M:%S')

    def _log(self, msg: str):
        # Emit through Qt signal to ensure GUI-thread-safe appending
        try:
            self.log_signal.emit(f'[{self._ts()}] {msg}')
        except Exception:
            # Fallback to status bar if no signal yet
            self.statusBar().showMessage(msg, 4000)

    @QtCore.pyqtSlot(str)
    def _append_log(self, txt: str):
        # Append to both logs if present and keep them scrolled
        try:
            if hasattr(self, 'test_log') and self.test_log is not None:
                self.test_log.appendPlainText(txt)
                try:
                    self.test_log.verticalScrollBar().setValue(self.test_log.verticalScrollBar().maximum())
                except Exception:
                    pass
            if hasattr(self, 'program_log') and self.program_log is not None:
                self.program_log.appendPlainText(txt)
                try:
                    self.program_log.verticalScrollBar().setValue(self.program_log.verticalScrollBar().maximum())
                except Exception:
                    pass
        except Exception:
            try:
                self.statusBar().showMessage(txt, 4000)
            except Exception:
                pass

    def on_abort_clicked(self):
        self._sequence_abort_flag = True
        self._log('Abort requested')

    def _auto_turn_off_panel(self, panel):
        """Safely turn off outputs/inputs for a newly added panel."""
        if isinstance(panel, KeithleyPanel) and getattr(panel, 'inst', None):
            for ch in (1, 2, 3):
                try:
                    panel.inst.set_output(ch, False)
                except Exception:
                    pass
            panel.master_out_btn.setChecked(False)
            panel.master_out_btn.setText('All Off')
        elif isinstance(panel, HittiteSigGenPanel) and getattr(panel, 'dev', None):
            try:
                panel.dev.set_output(False)
            except Exception:
                pass
            panel.output_btn.setChecked(False)
            panel.output_btn.setText('Output Off')
        elif isinstance(panel, KeysightELPanel) and getattr(panel, 'dev', None):
            for ch in (1, 2):
                if hasattr(panel, 'ch_enabled') and not panel.ch_enabled[ch].isChecked():
                    continue
                try:
                    panel.dev.set_input(ch, False)
                except Exception:
                    pass
            # Reflect in UI if available
            try:
                if hasattr(panel, '_set_input_toggle_ui'):
                    panel._set_input_toggle_ui(1, False)
                    panel._set_input_toggle_ui(2, False)
            except Exception:
                pass
        elif isinstance(panel, RhodeSchwarzSMAPanel) and getattr(panel, 'dev', None):
            try:
                panel.dev.set_output(False)
            except Exception:
                pass
            panel.output_btn.setChecked(False)
            panel.output_btn.setText('Output Off')
        elif isinstance(panel, FieldFoxSAPanel) and getattr(panel, 'dev', None):
            try:
                panel.dev.set_output(False)
            except Exception:
                pass
            panel.output_btn.setChecked(False)
            panel.output_btn.setText('Output Off')

    def _find_replacement_candidates(self, saved_resource: str, inst_type_hint: str, rm: pyvisa.ResourceManager = None) -> List[Tuple[str, str, str]]:
        """Return list of (label, resource, inst_type) candidates matching inst_type_hint."""
        candidates: List[Tuple[str, str, str]] = []
        if rm is None:
            try:
                rm = pyvisa.ResourceManager()
            except Exception:
                return candidates
        try:
            resources = rm.list_resources()
        except Exception:
            return candidates
        for res in resources:
            try:
                dev = rm.open_resource(res, timeout=1000)
                try:
                    idn = dev.query('*IDN?')
                except Exception:
                    idn = ''
                finally:
                    try:
                        dev.close()
                    except Exception:
                        pass
                lidn = idn.upper()
                inst_type = 'Unknown'
                if 'KEITHLEY' in lidn or '2230' in lidn:
                    inst_type = 'Keithley 2230'
                elif 'EL34243' in lidn or ('KEYSIGHT' in lidn and 'ELECTRONIC LOAD' in lidn):
                    inst_type = 'Keysight EL34243A'
                elif 'E36233A' in lidn:
                    inst_type = 'Keysight E36233A'
                elif 'HITTITE' in lidn or 'SIG GEN' in lidn:
                    inst_type = 'Hittite Sig Gen'
                elif 'ROHDE' in lidn or 'SCHWARZ' in lidn or 'SMA' in lidn:
                    inst_type = 'RhodeSchwarz SMA'
                elif 'FIELDFOX' in lidn or 'FIELD FOX' in lidn:
                    inst_type = 'Keysight FieldFox'
                if inst_type_hint and inst_type_hint.split()[0] == inst_type.split()[0]:
                    candidates.append((res.split('::')[3] if '::' in res else res, res, inst_type))
                elif inst_type == inst_type_hint:
                    candidates.append((res.split('::')[3] if '::' in res else res, res, inst_type))
            except Exception:
                continue
        return candidates
    
    def sequence_power_on(self):
        order = self.seq_order_edit.text().strip()
        if not order:
            self.statusBar().showMessage('No sequence order specified', 4000)
            return
        names = [n.strip().lower() for n in order.split(',') if n.strip()]
        for name in names:
            for i in range(self.tabs.count()):
                widget = self.tabs.widget(i)
                tab_name = self.tabs.tabText(i).lower()
                if name in tab_name:
                    if isinstance(widget, KeithleyPanel):
                        if widget.inst:
                            for ch in (1, 2, 3):
                                try:
                                    V = float(widget.vol_edits[ch].text())
                                    I = float(widget.iam_edits[ch].text())
                                    widget.inst.set_voltage(ch, V)
                                    widget.inst.set_current(ch, I)
                                    widget.inst.set_output(ch, True)
                                except Exception:
                                    pass
                            widget.master_out_btn.setChecked(True)
                            widget.master_out_btn.setText('All On')
                    elif isinstance(widget, HittiteSigGenPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(True)
                                widget.output_btn.setChecked(True)
                                widget.output_btn.setText('Output On')
                                widget.status_label.setText('Output ON')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
                    elif isinstance(widget, KeysightELPanel):
                        if widget.dev:
                            for ch in (1, 2):
                                if hasattr(widget, 'ch_enabled') and not widget.ch_enabled[ch].isChecked():
                                    continue
                                try:
                                    mode = (widget.mode_combo_ch1.currentText() if ch == 1 else widget.mode_combo_ch2.currentText())
                                    try:
                                        value = float(widget.mode_value_ch1.text() if ch == 1 else widget.mode_value_ch2.text())
                                    except Exception:
                                        value = 0.0
                                    widget.dev.set_mode(ch, mode)
                                    widget.dev.set_parameter(ch, mode, value)
                                    widget.dev.set_input(ch, True)
                                    if hasattr(widget, '_set_input_toggle_ui'):
                                        widget._set_input_toggle_ui(ch, True)
                                except Exception:
                                    pass
                    elif isinstance(widget, RhodeSchwarzSMAPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(True)
                                widget.output_btn.setChecked(True)
                                widget.output_btn.setText('Output On')
                                widget.status_label.setText('Output ON')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
                    elif isinstance(widget, FieldFoxSAPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(True)
                                widget.output_btn.setChecked(True)
                                widget.output_btn.setText('Output On')
                                widget.status_label.setText('Output ON')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
        self.statusBar().showMessage('Sequence power on complete', 4000)

    def sequence_power_off(self):
        order = self.seq_order_edit.text().strip()
        if not order:
            self.statusBar().showMessage('No sequence order specified', 4000)
            return
        names = [n.strip().lower() for n in order.split(',') if n.strip()]
        for name in names:
            for i in range(self.tabs.count()):
                widget = self.tabs.widget(i)
                tab_name = self.tabs.tabText(i).lower()
                if name in tab_name:
                    if isinstance(widget, KeithleyPanel):
                        if widget.inst:
                            for ch in (1, 2, 3):
                                try:
                                    widget.inst.set_output(ch, False)
                                except Exception:
                                    pass
                            widget.master_out_btn.setChecked(False)
                            widget.master_out_btn.setText('All Off')
                    elif isinstance(widget, HittiteSigGenPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(False)
                                widget.output_btn.setChecked(False)
                                widget.output_btn.setText('Output Off')
                                widget.status_label.setText('Output OFF')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
                    elif isinstance(widget, KeysightELPanel):
                        if widget.dev:
                            for ch in (1, 2):
                                if hasattr(widget, 'ch_enabled') and not widget.ch_enabled[ch].isChecked():
                                    continue
                                try:
                                    widget.dev.set_input(ch, False)
                                    if hasattr(widget, '_set_input_toggle_ui'):
                                        widget._set_input_toggle_ui(ch, False)
                                except Exception:
                                    pass
                    elif isinstance(widget, RhodeSchwarzSMAPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(False)
                                widget.output_btn.setChecked(False)
                                widget.output_btn.setText('Output Off')
                                widget.status_label.setText('Output OFF')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
                    elif isinstance(widget, FieldFoxSAPanel):
                        if widget.dev:
                            try:
                                widget.dev.set_output(False)
                                widget.output_btn.setChecked(False)
                                widget.output_btn.setText('Output Off')
                                widget.status_label.setText('Output OFF')
                            except Exception as e:
                                widget.status_label.setText(f'Failed to set output: {e}')
        self.statusBar().showMessage('Sequence power off complete', 4000)

    def reset_part(self):
        try:
            delay = float(self.reset_delay_edit.text())
        except Exception:
            delay = 2.0
        # clear any previous abort request and power off
        try:
            self._sequence_abort_flag = False
        except Exception:
            pass
        self._log(f'Reset requested: powering off, will power on in {delay} seconds')
        self.power_off_all()

        def _do_power_on():
            # If a power-up sequence is configured and enabled, use it; otherwise power all
            use_seq = False
            seq = []
            try:
                if hasattr(self, 'power_seq_builder'):
                    use_seq = self.power_seq_builder.use_sequence()
                    seq = self.power_seq_builder.get_sequence()
            except Exception:
                pass
            if use_seq and seq:
                self._log('Using configured power-up sequence for hard reset')
                if hasattr(self, 'test_power_toggle_btn'):
                    self.test_power_toggle_btn.setChecked(True)
                    self._update_test_power_toggle_btn(True)
                # run sequence then Soft Reset when complete
                def after_seq():
                    try:
                        self.soft_reset()
                    except Exception:
                        pass
                self._run_power_sequence(on_complete=after_seq)
            else:
                self._log('No sequence configured; powering all instruments ON then Soft Reset')
                self.power_on_all()
                try:
                    self.soft_reset()
                except Exception:
                    pass
        QtCore.QTimer.singleShot(int(delay * 1000), _do_power_on)
        self.statusBar().showMessage(f'Reset part: powered off, will power on in {delay} seconds', 4000)
    
    def __init__(self):
        super().__init__()
        # Apple liquid glass design stylesheet
        self.setStyleSheet('''
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e3eafc, stop:1 #cfd9df);
            }
            QWidget {
                background: rgba(255,255,255,0.7);
                border-radius: 16px;
                font-family: "Segoe UI", "San Francisco", Arial, sans-serif;
                font-size: 15px;
                color: #222;
            }
            QTabWidget::pane {
                border: none;
                background: transparent;
            }
            QTabBar::tab {
                background: rgba(255,255,255,0.6);
                border-radius: 12px;
                padding: 8px 24px;
                margin: 4px;
                font-weight: 500;
                color: #222;
                min-width: 140px; /* ensure long labels like 'Programming' are fully visible */
                border: 1px solid rgba(0,0,0,0.15); /* subtle outline for rounded tab */
            }
            QTabBar::tab:selected {
                background: rgba(255,255,255,0.85);
                color: #007aff;
                border: 1px solid #b6c2ce; 
            }
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #fafdff, stop:1 #e3eafc);
                border-radius: 12px;
                border: 1px solid #d1d5db;
                padding: 8px 20px;
                font-size: 15px;
                color: #222;
                font-weight: 500;
            }
            QPushButton:hover {
                background: #e3eafc;
                color: #007aff;
            }
            QPushButton:pressed {
                background: #cfd9df;
                color: #007aff;
            }
            QCheckBox {
                font-size: 15px;
                padding: 4px 12px;
                border-radius: 8px;
                background: rgba(255,255,255,0.5);
            }
            QLabel {
                font-size: 15px;
                color: #222;
                background: transparent;
            }
            QLineEdit {
                background: rgba(255,255,255,0.8);
                border-radius: 8px;
                border: 1px solid #d1d5db;
                padding: 6px 12px;
                font-size: 15px;
            }
            QPlainTextEdit {
                background: rgba(255,255,255,0.8);
                border-radius: 12px;
                border: 1px solid #d1d5db;
                font-size: 15px;
                padding: 8px;
            }
            QComboBox {
                background: rgba(255,255,255,0.8);
                border-radius: 8px;
                border: 1px solid #d1d5db;
                font-size: 15px;
                padding: 6px 12px;
            }
            QStatusBar {
                background: rgba(255,255,255,0.6);
                border-top: 1px solid #d1d5db;
                font-size: 14px;
                color: #222;
            }
    ''')
        # runtime control state for sequencing/program runs
        self._sequence_abort_flag = False
        self.program_process = None
        self._program_tempfile = None
        self._logic_thread = None
        self._logic_poll_timer = None
        self._logic_abort = False
        # aliasing for portable configs
        self.alias_dir = os.path.join(os.path.dirname(__file__), 'bench alias')
        os.makedirs(self.alias_dir, exist_ok=True)
        self.alias_profile = ''
        self.alias_map = {}
        # programming logic directory (external logic files for Configure Part)
        # Force-resolve to <repo>/PythonScripts/programming logics to avoid accidental Support_Scrips pathing
        try:
            # repo_root -> two levels up from Setup_GUI
            repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
            python_scripts_dir = os.path.join(repo_root, 'PythonScripts')
            self.program_logic_dir = os.path.join(python_scripts_dir, 'programming logics')
            os.makedirs(self.program_logic_dir, exist_ok=True)
            # Fallbacks if expected structure is not present
            if not os.path.isdir(self.program_logic_dir):
                self.program_logic_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'programming logics')
                os.makedirs(self.program_logic_dir, exist_ok=True)
        except Exception:
            # Final fallback: current module directory
            self.program_logic_dir = os.path.dirname(__file__)
        self.setWindowTitle("Instrument Controller - Multiple Instruments")
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        main_layout = QtWidgets.QVBoxLayout(central)

        # Top-level tabs: Setup and Test
        self.top_tabs = QtWidgets.QTabWidget()
        main_layout.addWidget(self.top_tabs)

        # --- Setup Tab ---
        setup_widget = QtWidgets.QWidget()
        setup_layout = QtWidgets.QVBoxLayout(setup_widget)

        # Configs folder
        self.configs_dir = os.path.join(os.path.dirname(__file__), 'configs')
        os.makedirs(self.configs_dir, exist_ok=True)

        # Top: Instrument type selector, Scan VISA, detected devices, Save/Load Config
        top_row = QtWidgets.QHBoxLayout()
        self.scan_btn = QtWidgets.QPushButton('Scan VISA')
        self.scan_btn.clicked.connect(self.on_scan_instruments)
        self.detected_combo = QtWidgets.QComboBox()
        self.detected_combo.setMinimumWidth(350)
        self.add_selected_btn = QtWidgets.QPushButton('Add Selected')
        self.add_selected_btn.clicked.connect(self.on_add_selected_instrument)
        self.save_btn = QtWidgets.QPushButton('Save Config')
        self.save_btn.clicked.connect(self.save_config_dialog)
        self.load_combo = QtWidgets.QComboBox()
        self.load_combo.setMinimumWidth(200)
        self.load_btn = QtWidgets.QPushButton('Load Config')
        self.load_btn.clicked.connect(self.load_config_dialog)
        self.refresh_configs_list()

        top_row.addWidget(self.scan_btn)
        top_row.addWidget(self.detected_combo)
        top_row.addWidget(self.add_selected_btn)
        top_row.addWidget(self.save_btn)
        top_row.addWidget(self.load_combo)
        top_row.addWidget(self.load_btn)
        setup_layout.addLayout(top_row)

        # Tabs for instruments
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.setTabsClosable(True)
        self.tabs.tabCloseRequested.connect(self.close_tab)
        # Auto-pause FieldFox streaming when tab not active
        try:
            self.tabs.currentChanged.connect(self._on_tab_changed)
        except Exception:
            pass
        setup_layout.addWidget(self.tabs)

    # Leave status blank until the initial VISA scan completes

        self.top_tabs.addTab(setup_widget, "Setup")

        # Connect logger signal to UI appender
        try:
            self.log_signal.connect(self._append_log)
        except Exception:
            pass

        # --- Test Tab ---
        test_widget = QtWidgets.QWidget()
        test_layout = QtWidgets.QVBoxLayout(test_widget)



    # (Removed duplicate Test tab UI block. All UI code is inside __init__ with correct indentation.)

        # Power sequence builder (wrapper)
        self.power_seq_builder = PowerSequenceBuilder(
            parent=self,
            get_instruments_callback=lambda: [self.tabs.tabText(i) for i in range(self.tabs.count())]
        )
        test_layout.addWidget(self.power_seq_builder)

        # Abort / Cancel button and run log
        abort_row = QtWidgets.QHBoxLayout()
        self.abort_btn = QtWidgets.QPushButton('Abort')
        self.abort_btn.setStyleSheet('background-color: #FF5722; color: white;')
        self.abort_btn.clicked.connect(self.on_abort_clicked)
        abort_row.addWidget(self.abort_btn)
        test_layout.addLayout(abort_row)

        # Run log is presented on the Test tab (live updates via self._log)

        self.top_tabs.addTab(test_widget, "Sequencing")

        # Ensure sequence builder combo is populated with any existing tabs
        try:
            if hasattr(self, 'power_seq_builder'):
                self.power_seq_builder.refresh_instr_combo()
        except Exception:
            pass

        # --- Programming Tab ---
        prog_widget = QtWidgets.QWidget()
        prog_layout = QtWidgets.QVBoxLayout(prog_widget)

        # Configure Part row only (select logic file and run)
        cfg_row = QtWidgets.QHBoxLayout()
        self.logic_combo = QtWidgets.QComboBox()
        self.logic_combo.setMinimumWidth(420)
        self._refresh_logic_combo()
        cfg_row.addWidget(QtWidgets.QLabel('Logic file:'))
        cfg_row.addWidget(self.logic_combo, 1)
        self.logic_browse_btn = QtWidgets.QPushButton('Browse…')
        self.logic_browse_btn.clicked.connect(self._browse_logic_file)
        cfg_row.addWidget(self.logic_browse_btn)
        self.configure_part_btn = QtWidgets.QPushButton('Configure Part')
        self.configure_part_btn.clicked.connect(self.on_configure_part_clicked)
        cfg_row.addWidget(self.configure_part_btn)
        prog_layout.addLayout(cfg_row)

        # Search for new programmings row
        search_row = QtWidgets.QHBoxLayout()
        self.logic_rescan_btn = QtWidgets.QPushButton('Search for New Programmings')
        def _on_rescan_logic():
            try:
                # Refresh list
                self._refresh_logic_combo()
                # Pick the most recently modified .py in the directory
                try:
                    files = [f for f in os.listdir(self.program_logic_dir) if f.lower().endswith('.py')]
                except Exception:
                    files = []
                newest_path = ''
                newest_mtime = -1
                for f in files:
                    full = os.path.join(self.program_logic_dir, f)
                    try:
                        mt = os.path.getmtime(full)
                        if mt > newest_mtime:
                            newest_mtime = mt
                            newest_path = full
                    except Exception:
                        pass
                if newest_path:
                    self._ensure_logic_in_combo(newest_path)
                    self.statusBar().showMessage(f'Selected newest logic: {os.path.basename(newest_path)}', 4000)
                else:
                    self.statusBar().showMessage('No logic files found', 3000)
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, 'Rescan failed', str(e))
        self.logic_rescan_btn.clicked.connect(_on_rescan_logic)
        search_row.addWidget(self.logic_rescan_btn)
        prog_layout.addLayout(search_row)

        # Programming tab run log (live)
        prog_layout.addWidget(QtWidgets.QLabel('Run Log:'))
        self.program_log = QtWidgets.QPlainTextEdit()
        self.program_log.setReadOnly(True)
        self.program_log.setMaximumHeight(220)
        prog_layout.addWidget(self.program_log)

        self.top_tabs.addTab(prog_widget, 'Programming')

        # --- Test Tab ---
        test2_widget = QtWidgets.QWidget()
        test2_layout = QtWidgets.QVBoxLayout(test2_widget)

        # Power On/Off toggle (syncs with Sequencing tab power button)
        tpower_row = QtWidgets.QHBoxLayout()
        self.test_tab_power_btn = QtWidgets.QPushButton('Power Off All')
        self.test_tab_power_btn.setCheckable(True)
        self.test_tab_power_btn.setChecked(False)
        self.test_tab_power_btn.clicked.connect(self.on_test_power_toggle)
        if hasattr(self, '_update_test_power_toggle_btn'):
            self._update_test_power_toggle_btn(False)
        tpower_row.addWidget(self.test_tab_power_btn)
        test2_layout.addLayout(tpower_row)

        # --- Unified Record Controls ---
        record_row = QtWidgets.QHBoxLayout()
        # Prime button prepares threads and enables Start
        self.prime_btn = QtWidgets.QPushButton('Prime Recorders')
        self.prime_btn.setToolTip('Prepare selected recorders and Excel file; enables Start')
        self.prime_btn.clicked.connect(self.on_prime_clicked)
        record_row.addWidget(self.prime_btn)
        self.record_btn = QtWidgets.QPushButton('Record')
        self.record_btn.setCheckable(True)
        self.record_btn.setChecked(False)
        self.record_btn.clicked.connect(self.on_record_clicked)
        # Start disabled until primed
        self.record_btn.setEnabled(False)
        record_row.addWidget(self.record_btn)
        self.supply_record_toggle = QtWidgets.QCheckBox('Supply')
        self.supply_record_toggle.setChecked(True)
        record_row.addWidget(self.supply_record_toggle)
        self.spectrum_record_toggle = QtWidgets.QCheckBox('Spectrum')
        self.spectrum_record_toggle.setChecked(False)
        record_row.addWidget(self.spectrum_record_toggle)
        # Register recording toggle
        self.register_record_toggle = QtWidgets.QCheckBox('Register')
        self.register_record_toggle.setChecked(False)
        record_row.addWidget(self.register_record_toggle)
        self.supply_read_speed_label = QtWidgets.QLabel('Supply Sample Rate: --')
        self.spectrum_read_speed_label = QtWidgets.QLabel('Spectrum Sample Rate: --')
        self.register_read_speed_label = QtWidgets.QLabel('Register Sample Rate: --')
        record_row.addWidget(self.supply_read_speed_label)
        record_row.addWidget(self.spectrum_read_speed_label)
        record_row.addWidget(self.register_read_speed_label)
        # Add more toggles here for future metrics
        test2_layout.addLayout(record_row)
        # Priming state
        self._record_primed = False

        # Soft Reset button (runs Configure Part logic)
        prog_row = QtWidgets.QHBoxLayout()
        self.program_now_btn = QtWidgets.QPushButton('Program Device / Soft Reset')
        self.program_now_btn.clicked.connect(self.soft_reset)
        prog_row.addWidget(self.program_now_btn)
        test2_layout.addLayout(prog_row)

        # Reset controls
        reset_row = QtWidgets.QHBoxLayout()
        reset_row.addWidget(QtWidgets.QLabel('Reset delay (s):'))
        self.reset_delay_edit = QtWidgets.QLineEdit('2.0')
        self.reset_delay_edit.setMaximumWidth(80)
        reset_row.addWidget(self.reset_delay_edit)
        self.reset_btn = QtWidgets.QPushButton('Hard Reset')
        self.reset_btn.clicked.connect(self.reset_part)
        reset_row.addWidget(self.reset_btn)
        test2_layout.addLayout(reset_row)

        # Live run log for Configure Part
        test2_layout.addWidget(QtWidgets.QLabel('Run Log:'))
        self.test_log = QtWidgets.QPlainTextEdit()
        self.test_log.setReadOnly(True)
        self.test_log.setMaximumHeight(220)
        test2_layout.addWidget(self.test_log)

        self.top_tabs.addTab(test2_widget, 'Test')

        # --- Settings Tab (Aliases) ---
        self._init_settings_tab()

        # Prompt for aliasing mode BEFORE any VISA scan; the prompt will trigger the first scan.
        try:
            QtCore.QTimer.singleShot(0, self._prompt_alias_startup)
        except Exception:
            pass

    def _on_tab_changed(self, index: int):
        try:
            # Determine which FieldFox panels exist and whether one is active
            active_panel = None
            try:
                w = self.tabs.widget(index)
                if isinstance(w, FieldFoxSAPanel):
                    active_panel = w
            except Exception:
                active_panel = None
            for i in range(self.tabs.count()):
                panel = self.tabs.widget(i)
                if isinstance(panel, FieldFoxSAPanel):
                    if panel is active_panel:
                        # Ensure streaming is on
                        if hasattr(panel, 'stream_btn') and hasattr(panel, 'toggle_streaming'):
                            if not panel.streaming_enabled:
                                panel.stream_btn.setChecked(True)
                                panel.toggle_streaming(True)
                    else:
                        # Pause streaming on non-active FieldFox tabs
                        if hasattr(panel, 'stream_btn') and hasattr(panel, 'toggle_streaming'):
                            if panel.streaming_enabled:
                                panel.stream_btn.setChecked(False)
                                panel.toggle_streaming(False)
        except Exception:
            pass

    def get_instrument_names(self):
        names = []
        for i in range(self.tabs.count()):
            names.append(self.tabs.tabText(i))
        return names

    def power_off_all(self):
        """Turn off outputs/inputs for all instruments in tabs."""
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            if isinstance(widget, KeithleyPanel):
                self._keithley_apply_settings(widget, False)
            elif isinstance(widget, HittiteSigGenPanel):
                self._generic_output_toggle(widget, False)
            elif isinstance(widget, KeysightELPanel):
                self._keysight_el_apply_settings(widget, False)
            elif isinstance(widget, RhodeSchwarzSMAPanel):
                self._generic_output_toggle(widget, False)
            elif isinstance(widget, FieldFoxSAPanel):
                self._generic_output_toggle(widget, False)
        if hasattr(self, 'test_power_toggle_btn'):
            self.test_power_toggle_btn.setChecked(False)
            self._update_test_power_toggle_btn(False)
        self._update_global_power_btns(False)
        self.statusBar().showMessage('All instrument outputs/inputs turned OFF', 4000)

    def power_on_all(self):
        """Turn on outputs/inputs for all instruments in tabs."""
        # First, power on all instruments except Keysight EL
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            if isinstance(widget, KeithleyPanel):
                self._keithley_apply_settings(widget, True)
            elif isinstance(widget, HittiteSigGenPanel):
                self._generic_output_toggle(widget, True)
            elif isinstance(widget, RhodeSchwarzSMAPanel):
                self._generic_output_toggle(widget, True)
            elif isinstance(widget, FieldFoxSAPanel):
                self._generic_output_toggle(widget, True)
        # Then, ramp and power on Keysight EL
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            if isinstance(widget, KeysightELPanel):
                self._keysight_el_apply_settings(widget, True)
        if hasattr(self, 'test_power_toggle_btn'):
            self.test_power_toggle_btn.setChecked(True)
            self._update_test_power_toggle_btn(True)
        self._update_global_power_btns(True)
        self.statusBar().showMessage('All instrument outputs/inputs turned ON', 4000)

    def refresh_configs_list(self):
        self.load_combo.clear()
        # Add placeholder entry first
        placeholder = '---Config List--'
        self.load_combo.addItem(placeholder, None)
        # Populate available configs (sorted for stability)
        files = sorted([f for f in os.listdir(self.configs_dir) if f.lower().endswith('.json')])
        for f in files:
            full = os.path.join(self.configs_dir, f)
            self.load_combo.addItem(f, full)
        # Default to placeholder regardless of files present
        self.load_combo.setCurrentIndex(0)
        # Also refresh alias profiles list in Settings tab
        try:
            if hasattr(self, 'alias_profile_combo'):
                self._refresh_alias_profiles()
        except Exception:
            pass

    # ---------- Alias profiles (Settings tab) ----------
    def _alias_profile_path(self, name: str) -> str:
        return os.path.join(self.alias_dir, f"{name}.json")

    def _list_alias_profiles(self):
        try:
            return [os.path.splitext(f)[0] for f in os.listdir(self.alias_dir) if f.lower().endswith('.json')]
        except Exception:
            return []

    def _refresh_alias_profiles(self):
        try:
            names = sorted(self._list_alias_profiles())
            self.alias_profile_combo.clear()
            for n in names:
                self.alias_profile_combo.addItem(n)
            # reflect current
            if self.alias_profile:
                idx = self.alias_profile_combo.findText(self.alias_profile)
                if idx != -1:
                    self.alias_profile_combo.setCurrentIndex(idx)
        except Exception:
            pass

    def _apply_alias_profile_to_ui(self):
        try:
            # Populate dual lists: aliases on left, resources on right
            aliases = sorted(self.alias_map.keys())
            self.alias_list.clear()
            if hasattr(self, 'alias_res_list'):
                self.alias_res_list.clear()
            for k in aliases:
                self.alias_list.addItem(k)
                if hasattr(self, 'alias_res_list'):
                    self.alias_res_list.addItem(self.alias_map.get(k, ''))
            self.alias_profile_name.setText(self.alias_profile or '')
        except Exception:
            pass

    def _sync_alias_map_from_lists(self):
        """Rebuild alias_map from the current lists (index-aligned)."""
        try:
            if not hasattr(self, 'alias_res_list'):
                return
            new_map = {}
            n = min(self.alias_list.count(), self.alias_res_list.count())
            for i in range(n):
                alias_item = self.alias_list.item(i)
                res_item = self.alias_res_list.item(i)
                if not alias_item:
                    continue
                alias = alias_item.text()
                res = res_item.text() if res_item else ''
                new_map[alias] = res
            self.alias_map = new_map
            self.statusBar().showMessage('Updated alias mapping from drag-and-drop', 3000)
        except Exception:
            pass
        # Also update the resource lists in the settings tab if visible
        try:
            if hasattr(self, 'alias_profile_combo'):
                self._refresh_alias_profiles()
        except Exception:
            pass

    def _load_alias_profile(self, name: str):
        try:
            path = self._alias_profile_path(name)
            if not os.path.exists(path):
                return
            with open(path, 'r') as f:
                data = json.load(f)
            self.alias_profile = data.get('name', name)
            self.alias_map = data.get('aliases', {}) or {}
            self._apply_alias_profile_to_ui()
            self.statusBar().showMessage(f"Loaded alias profile: {self.alias_profile}", 4000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Alias load failed', str(e))

    def _save_alias_profile(self, name: str = None):
        try:
            name = name or self.alias_profile_name.text().strip()
            if not name:
                QtWidgets.QMessageBox.information(self, 'Missing name', 'Enter a profile name.')
                return
            payload = {'name': name, 'aliases': self.alias_map}
            path = self._alias_profile_path(name)
            with open(path, 'w') as f:
                json.dump(payload, f, indent=2)
            self.alias_profile = name
            self._refresh_alias_profiles()
            self._apply_alias_profile_to_ui()
            self.statusBar().showMessage(f"Saved alias profile: {name}", 4000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Alias save failed', str(e))

    def _delete_alias_profile(self):
        try:
            name = self.alias_profile_combo.currentText().strip()
            if not name:
                return
            resp = QtWidgets.QMessageBox.question(self, 'Delete profile', f'Delete alias profile "{name}"?',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
            if resp != QtWidgets.QMessageBox.Yes:
                return
            path = self._alias_profile_path(name)
            if os.path.exists(path):
                os.remove(path)
            if self.alias_profile == name:
                self.alias_profile = ''
                self.alias_map = {}
            self._refresh_alias_profiles()
            self._apply_alias_profile_to_ui()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Alias delete failed', str(e))

    def _build_aliases_from_resources(self, resources: list):
        """Classify and assign type+index aliases for current VISA resources."""
        # Simple classification via *IDN?
        try:
            rm = pyvisa.ResourceManager()
        except Exception:
            rm = None
        cats = {}
        for res in resources:
            typ = 'Unknown'
            idn = ''
            try:
                if rm:
                    inst = rm.open_resource(res, timeout=1500)
                    try:
                        try:
                            inst.clear()
                        except Exception:
                            pass
                        idn = inst.query('*IDN?').upper()
                    finally:
                        try:
                            inst.close()
                        except Exception:
                            pass
            except Exception:
                idn = ''
            if 'KEITHLEY' in idn or '2230' in idn:
                typ = 'Keithley'
            elif 'EL34243' in idn or ('KEYSIGHT' in idn and 'ELECTRONIC LOAD' in idn):
                typ = 'Keysight EL'
            elif 'E36233A' in idn:
                typ = 'E36233A'
            elif 'FIELDFOX' in idn or 'N99' in idn or 'HANDHELD SPECTRUM' in idn:
                typ = 'FieldFox'
            elif 'HITTITE' in idn or 'SIG GEN' in idn:
                typ = 'Hittite'
            elif 'ROHDE' in idn or 'SCHWARZ' in idn or 'SMA' in idn:
                typ = 'RohdeSchwarz SMA'
            cats.setdefault(typ, []).append(res)
        # Assign aliases: "Keithley 0..", "FieldFox 0..", generic type+index otherwise
        new_map = {}
        for typ, lst in cats.items():
            for i, res in enumerate(sorted(lst)):
                if typ == 'Keithley':
                    alias = f'Keithley {i}'
                elif typ == 'FieldFox':
                    alias = f'FieldFox {i}'
                else:
                    alias = f'{typ} {i}'
                new_map[alias] = res
        self.alias_map = new_map
        self._apply_alias_profile_to_ui()

    def _refresh_unmapped_instruments(self):
        """Scan VISA and list instruments not present in current alias_map."""
        try:
            self.unmapped_list.clear()
        except Exception:
            pass
        try:
            rm = pyvisa.ResourceManager()
            resources = list(rm.list_resources())
        except Exception:
            resources = []
        # Build set of already-mapped resources
        mapped = set((self.alias_map or {}).values())
        # For display, include a simple type guess via *IDN?
        def classify(res: str) -> str:
            try:
                inst = rm.open_resource(res, timeout=1200)
                try:
                    try:
                        inst.clear()
                    except Exception:
                        pass
                    idn = inst.query('*IDN?')
                finally:
                    try:
                        inst.close()
                    except Exception:
                        pass
                lidn = (idn or '').upper()
                if 'KEITHLEY' in lidn or '2230' in lidn:
                    return 'Keithley'
                if 'EL34243' in lidn or ('KEYSIGHT' in lidn and 'ELECTRONIC LOAD' in lidn):
                    return 'Keysight EL'
                if 'E36233A' in lidn:
                    return 'E36233A'
                if 'FIELDFOX' in lidn or 'N99' in lidn or 'HANDHELD SPECTRUM' in lidn:
                    return 'FieldFox'
                if 'HITTITE' in lidn or 'SIG GEN' in lidn:
                    return 'Hittite'
                if 'ROHDE' in lidn or 'SCHWARZ' in lidn or 'SMA' in lidn:
                    return 'RohdeSchwarz SMA'
                return 'Unknown'
            except Exception:
                return 'Unknown'
        try:
            for res in resources:
                if res in mapped:
                    continue
                typ = classify(res)
                # Show short label (last token) alongside full resource and type
                short = res.split('::')[3] if '::' in res and len(res.split('::')) > 3 else res
                item = QtWidgets.QListWidgetItem(f'{short}    ({typ})\n{res}')
                # Store original resource and type in item data
                item.setData(QtCore.Qt.UserRole, (res, typ))
                self.unmapped_list.addItem(item)
            self.statusBar().showMessage('Unmapped instruments list refreshed', 3000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Scan failed', str(e))

    def _add_selected_unmapped_to_profile(self, add_all: bool = False):
        """Add selected (or all) unmapped instruments into the current alias profile."""
        # Ensure we have a profile name to save into
        try:
            if not self.alias_profile:
                # If no named profile, default to 'default' in-memory without saving yet
                self.alias_profile = self.alias_profile or 'default'
        except Exception:
            pass
        try:
            items = []
            if add_all:
                for i in range(self.unmapped_list.count()):
                    items.append(self.unmapped_list.item(i))
            else:
                items = self.unmapped_list.selectedItems() or []
            if not items:
                QtWidgets.QMessageBox.information(self, 'No selection', 'No instruments selected to add.')
                return
            # Build counters per type to generate sensible alias names
            counters = {}
            # Seed counters based on existing aliases to avoid collisions
            try:
                for alias in (self.alias_map or {}).keys():
                    base = ''.join(ch for ch in alias if not ch.isdigit()).strip()
                    num = ''.join(ch for ch in alias if ch.isdigit())
                    if base:
                        counters[base] = max(counters.get(base, -1), int(num) if num.isdigit() else -1)
            except Exception:
                pass
            for it in items:
                res, typ = it.data(QtCore.Qt.UserRole)
                # Map type to a base alias token
                base = {
                    'Keithley': 'Keithley',
                    'Keysight EL': 'Keysight EL',
                    'E36233A': 'E36233A',
                    'FieldFox': 'FieldFox',
                    'Hittite': 'Hittite',
                    'RohdeSchwarz SMA': 'RohdeSchwarz SMA',
                }.get(typ, 'Unknown')
                idx = counters.get(base, -1) + 1
                counters[base] = idx
                alias = f'{base} {idx}' if base != 'Unknown' else f'Unknown {idx}'
                # Add to alias map
                self.alias_map[alias] = res
            # Refresh UI
            self._apply_alias_profile_to_ui()
            # Persist profile if a name is present
            try:
                if self.alias_profile:
                    self._save_alias_profile(self.alias_profile)
            except Exception:
                pass
            # Remove added items from the unmapped list
            try:
                for it in items:
                    row = self.unmapped_list.row(it)
                    self.unmapped_list.takeItem(row)
            except Exception:
                pass
            self.statusBar().showMessage('Added instrument(s) to alias profile', 3000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Add failed', str(e))

    def _suggest_aliases_from_scan(self):
        """Scan VISA and populate alias_map with suggested aliases for all detected instruments.
        Existing aliases can be kept and only missing ones added, or replaced based on a prompt.
        """
        try:
            rm = pyvisa.ResourceManager()
            resources = list(rm.list_resources())
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Scan failed', str(e))
            return
        # Build a fresh suggestion map by classifying and indexing
        suggestion = {}
        try:
            cats = {}
            for res in resources:
                typ = 'Unknown'
                idn = ''
                try:
                    inst = rm.open_resource(res, timeout=1200)
                    try:
                        try:
                            inst.clear()
                        except Exception:
                            pass
                        idn = inst.query('*IDN?').upper()
                    finally:
                        try:
                            inst.close()
                        except Exception:
                            pass
                except Exception:
                    idn = ''
                if 'KEITHLEY' in idn or '2230' in idn:
                    typ = 'Keithley'
                elif 'EL34243' in idn or ('KEYSIGHT' in idn and 'ELECTRONIC LOAD' in idn):
                    typ = 'Keysight EL'
                elif 'E36233A' in idn:
                    typ = 'E36233A'
                elif 'FIELDFOX' in idn or 'N99' in idn or 'HANDHELD SPECTRUM' in idn:
                    typ = 'FieldFox'
                elif 'HITTITE' in idn or 'SIG GEN' in idn:
                    typ = 'Hittite'
                elif 'ROHDE' in idn or 'SCHWARZ' in idn or 'SMA' in idn:
                    typ = 'RohdeSchwarz SMA'
                cats.setdefault(typ, []).append(res)
            for typ, lst in cats.items():
                for i, res in enumerate(sorted(lst)):
                    alias = f'{typ} {i}' if typ != 'Unknown' else f'Unknown {i}'
                    suggestion[alias] = res
        except Exception:
            suggestion = {f'Resource {i}': res for i, res in enumerate(resources)}
        # Ask how to apply
        try:
            resp = QtWidgets.QMessageBox.question(
                self, 'Apply suggested aliases',
                'Apply suggested aliases to current profile?\nYes: merge (keep existing, add new)\nNo: replace current profile',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Yes)
        except Exception:
            resp = QtWidgets.QMessageBox.Yes
        if resp == QtWidgets.QMessageBox.Cancel:
            return
        try:
            if resp == QtWidgets.QMessageBox.No:
                # Replace
                self.alias_map = suggestion
            else:
                # Merge: keep existing; add missing resources
                existing_res = set(self.alias_map.values()) if self.alias_map else set()
                for alias, res in suggestion.items():
                    if res in existing_res:
                        continue
                    new_alias = alias
                    k = 1
                    while new_alias in self.alias_map:
                        new_alias = f'{alias} {k}'
                        k += 1
                    self.alias_map[new_alias] = res
            self._apply_alias_profile_to_ui()
            if self.alias_profile:
                self._save_alias_profile(self.alias_profile)
            self.statusBar().showMessage('Aliases updated from scan', 3000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Apply failed', str(e))

    def _auto_select_alias_profile(self, resources: list):
        """Auto-pick the best matching alias profile based on current resources."""
        try:
            profiles = self._list_alias_profiles()
            best = None
            best_score = -1
            res_set = set(resources)
            for name in profiles:
                try:
                    with open(self._alias_profile_path(name), 'r') as f:
                        data = json.load(f)
                    amap = data.get('aliases', {}) or {}
                    vals = set(amap.values())
                    # score: exact match gets a bonus
                    score = len(vals & res_set)
                    if vals and vals <= res_set and len(vals) == len(res_set):
                        score += 1000
                    if score > best_score:
                        best_score = score
                        best = (name, amap)
                except Exception:
                    continue
            if best and (self.alias_profile != best[0]):
                self.alias_profile = best[0]
                self.alias_map = best[1]
                self._refresh_alias_profiles()
                self._apply_alias_profile_to_ui()
                self.statusBar().showMessage(f"Auto-selected alias profile: {self.alias_profile}", 5000)
        except Exception:
            pass

    def _init_settings_tab(self):
        settings_widget = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(settings_widget)
        # Profile row
       
        prow = QtWidgets.QHBoxLayout()
        self.alias_profile_combo = QtWidgets.QComboBox()
        self.alias_profile_combo.setMinimumWidth(220)
        self.alias_profile_name = QtWidgets.QLineEdit()
        self.alias_profile_name.setPlaceholderText('Profile name')
        self._refresh_alias_profiles()
        load_btn = QtWidgets.QPushButton('Load Profile')
        load_btn.clicked.connect(lambda: self._load_alias_profile(self.alias_profile_combo.currentText()))
        save_btn = QtWidgets.QPushButton('Save Profile')
        save_btn.clicked.connect(lambda: self._save_alias_profile(self.alias_profile_name.text()))
        del_btn = QtWidgets.QPushButton('Delete Profile')
        del_btn.clicked.connect(self._delete_alias_profile)
        prow.addWidget(QtWidgets.QLabel('Profile:'))
        prow.addWidget(self.alias_profile_combo)
        prow.addWidget(QtWidgets.QLabel('Name:'))
        prow.addWidget(self.alias_profile_name)
        prow.addWidget(load_btn)
        prow.addWidget(save_btn)
        prow.addWidget(del_btn)
        layout.addLayout(prow)
        # Alias mapping area with drag-and-drop reorder support (right list reorders resources)
        layout.addWidget(QtWidgets.QLabel('Drag instruments (right) to align with alias names (left):'))
        lists_row = QtWidgets.QHBoxLayout()
        # Left: alias names (static)
        self.alias_list = QtWidgets.QListWidget()
        self.alias_list.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.alias_list.setDragEnabled(False)
        self.alias_list.setAcceptDrops(False)
        # Right: resources (draggable to reorder)
        self.alias_res_list = QtWidgets.QListWidget()
        self.alias_res_list.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.alias_res_list.setDragEnabled(True)
        self.alias_res_list.setAcceptDrops(True)
        self.alias_res_list.setDefaultDropAction(QtCore.Qt.MoveAction)
        self.alias_res_list.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        try:
            self.alias_res_list.model().rowsMoved.connect(lambda *args: self._sync_alias_map_from_lists())
        except Exception:
            pass
        lists_row.addWidget(self.alias_list, 1)
        lists_row.addWidget(self.alias_res_list, 2)
        layout.addLayout(lists_row)
        # Unmapped instruments area
        layout.addWidget(QtWidgets.QLabel('Unmapped instruments (present on VISA but not in this alias profile):'))
        self.unmapped_list = QtWidgets.QListWidget()
        self.unmapped_list.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        layout.addWidget(self.unmapped_list)

        # Controls for refreshing and adding to profile
        controls_row = QtWidgets.QHBoxLayout()
        refresh_btn = QtWidgets.QPushButton('Refresh Instruments')
        refresh_btn.setToolTip('Scan VISA and list devices not currently mapped in the alias profile')
        refresh_btn.clicked.connect(self._refresh_unmapped_instruments)
        add_sel_btn = QtWidgets.QPushButton('Add Selected to Profile')
        add_sel_btn.setToolTip('Add the selected unmapped instruments to this alias profile')
        add_sel_btn.clicked.connect(self._add_selected_unmapped_to_profile)
        add_all_btn = QtWidgets.QPushButton('Add All Missing')
        add_all_btn.setToolTip('Add all unmapped instruments to this alias profile')
        add_all_btn.clicked.connect(lambda: self._add_selected_unmapped_to_profile(add_all=True))
        controls_row.addWidget(refresh_btn)
        controls_row.addWidget(add_sel_btn)
        controls_row.addWidget(add_all_btn)
        controls_row.addStretch(1)
        layout.addLayout(controls_row)

        # Utilities row: Suggest Aliases (auto-build from current VISA scan)
        util_row = QtWidgets.QHBoxLayout()
        suggest_btn = QtWidgets.QPushButton('Suggest Aliases')
        suggest_btn.setToolTip('Automatically generate alias names for all detected instruments')
        suggest_btn.clicked.connect(self._suggest_aliases_from_scan)
        util_row.addWidget(suggest_btn)
        util_row.addStretch(1)
        layout.addLayout(util_row)

        self.top_tabs.addTab(settings_widget, 'Settings')

    def disconnect_all_instruments(self):
        """Cleanly stop recordings, disconnect/clear all panels and instruments, and remove tabs.
        This prevents VISA/SCPI contention when loading a new config (e.g., FieldFox double-connect).
        """
        try:
            self.statusBar().showMessage('Disconnecting all instruments…', 3000)
        except Exception:
            pass
        # Stop any ongoing recordings first
        try:
            # If Record toggle is on, untoggle it for UI consistency
            if hasattr(self, 'record_btn') and self.record_btn.isChecked():
                try:
                    self.record_btn.setChecked(False)
                except Exception:
                    pass
            self._stop_all_recordings()
        except Exception:
            pass
        # Iterate tabs backwards: close panels and remove tabs
        try:
            for i in range(self.tabs.count() - 1, -1, -1):
                w = self.tabs.widget(i)
                # Try to turn off outputs/inputs first (best-effort)
                try:
                    if isinstance(w, KeithleyPanel):
                        self._keithley_apply_settings(w, False)
                    elif isinstance(w, KeysightELPanel):
                        self._keysight_el_apply_settings(w, False)
                    elif isinstance(w, (HittiteSigGenPanel, RhodeSchwarzSMAPanel, FieldFoxSAPanel)):
                        self._generic_output_toggle(w, False)
                except Exception:
                    pass
                # Close panel (panels should release VISA sessions/threads in close())
                try:
                    if hasattr(w, 'close'):
                        w.close()
                except Exception:
                    pass
                # As a fallback, try to close raw instrument handles
                try:
                    for attr in ('inst', 'dev', 'sa'):
                        obj = getattr(w, attr, None)
                        if obj is None:
                            continue
                        # FieldFox wrapper: has close()
                        if hasattr(obj, 'close'):
                            try:
                                obj.close()
                            except Exception:
                                pass
                        # VISA instrument directly
                        if hasattr(obj, 'inst') and hasattr(obj.inst, 'close'):
                            try:
                                obj.inst.close()
                            except Exception:
                                pass
                except Exception:
                    pass
                # Remove tab
                try:
                    self.tabs.removeTab(i)
                except Exception:
                    pass
        except Exception:
            pass
        # Small yield to let timers/threads settle
        try:
            QtWidgets.QApplication.processEvents()
        except Exception:
            pass

    def save_config_dialog(self):
        # Ask for filename to save config
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Config', self.configs_dir, 'JSON Files (*.json)')
        if not fname:
            return
        if not fname.lower().endswith('.json'):
            fname += '.json'
        self.save_config(fname)
        self.refresh_configs_list()

    def save_config(self, path):
        instruments = []
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            # Determine panel type
            if isinstance(widget, KeithleyPanel):
                tab_type = 'Keithley 2230'
            elif isinstance(widget, KeysightELPanel):
                tab_type = 'Keysight EL34243A'
            elif isinstance(widget, KeysightE36233APanel):
                tab_type = 'Keysight E36233A'
            elif isinstance(widget, HittiteSigGenPanel):
                tab_type = 'Hittite Sig Gen'
            elif isinstance(widget, RhodeSchwarzSMAPanel):
                tab_type = 'RhodeSchwarz SMA'
            elif isinstance(widget, FieldFoxSAPanel):
                tab_type = 'Keysight FieldFox'
            else:
                tab_type = 'Unknown'
            # determine a saved name: prefer the editable name if present, otherwise the current tab text
            try:
                if hasattr(widget, 'name_edit') and widget.name_edit.text().strip():
                    saved_name = widget.name_edit.text().strip()
                else:
                    saved_name = self.tabs.tabText(i)
            except Exception:
                try:
                    saved_name = self.tabs.tabText(i)
                except Exception:
                    saved_name = ''
            # Robust resource determination (FieldFox panel may not expose 'resource' attribute)
            resource_val = getattr(widget, 'resource', '')
            if (not resource_val) and isinstance(widget, FieldFoxSAPanel) and hasattr(widget, 'sa'):
                try:
                    resource_val = getattr(widget.sa, 'visa_address', '')
                except Exception:
                    resource_val = ''
            # If an alias profile is active, reverse-map resource to alias key for portability
            try:
                if isinstance(resource_val, str) and self.alias_map:
                    for alias_key, res in self.alias_map.items():
                        if res == resource_val:
                            resource_val = f"alias:{alias_key}"
                            break
            except Exception:
                pass
            entry = {'type': tab_type, 'resource': resource_val, 'name': saved_name}
            # Save per-panel attributes
            if isinstance(widget, KeithleyPanel):
                entry['channels'] = {}
                for ch in (1, 2, 3):
                    entry['channels'][ch] = {
                        'voltage': widget.vol_edits[ch].text(),
                        'current': widget.iam_edits[ch].text(),
                        'output': widget.master_out_btn.isChecked(),
                        'name': (widget.ch_name_edits[ch].text() if hasattr(widget, 'ch_name_edits') and widget.ch_name_edits.get(ch) else f'CH{ch}')
                    }
            elif isinstance(widget, KeysightELPanel):
                entry['channels'] = {}
                for ch in (1, 2):
                    # Read per-channel UI
                    mode = (widget.mode_combo_ch1.currentText() if ch == 1 else widget.mode_combo_ch2.currentText())
                    value = (widget.mode_value_ch1.text() if ch == 1 else widget.mode_value_ch2.text())
                    # Determine input state: prefer device state if connected
                    inp = False
                    try:
                        if getattr(widget, 'dev', None):
                            inp = bool(widget.dev.get_input_state(ch))
                        elif hasattr(widget, '_get_input_toggle_ui'):
                            inp = bool(widget._get_input_toggle_ui(ch))
                    except Exception:
                        pass
                    # Save ramp enable and rise time in addition to mode/value/input
                    ramp_enable = (widget.ramp_enable_ch1.isChecked() if ch == 1 else widget.ramp_enable_ch2.isChecked())
                    rise_time = (widget.rise_time_ch1.text() if ch == 1 else widget.rise_time_ch2.text())
                    entry['channels'][ch] = {
                        'mode': mode,
                        'value': value,
                        'input': inp,
                        'ramp_enabled': ramp_enable,
                        'rise_time': rise_time
                    }
            elif isinstance(widget, KeysightE36233APanel):
                entry['channels'] = {}
                for ch in (1, 2):
                    entry['channels'][ch] = {
                        'voltage': widget.voltage_edit.text(),
                        'current': widget.current_edit.text(),
                        'output': widget.onoff_btn.isChecked(),
                        'name': f'CH{ch}'
                    }
            elif isinstance(widget, HittiteSigGenPanel):
                # Save frequency (value + unit if separated) and power + output state if available
                try:
                    entry['settings'] = {
                        'frequency': getattr(widget, 'freq_edit', None).text() if hasattr(widget, 'freq_edit') else '',
                        'freq_unit': getattr(widget, 'freq_unit_combo', None).currentText() if hasattr(widget, 'freq_unit_combo') else 'MHz',
                        'power': getattr(widget, 'pow_edit', None).text() if hasattr(widget, 'pow_edit') else '',
                        'output': getattr(widget, 'output_btn', None).isChecked() if hasattr(widget, 'output_btn') else False
                    }
                except Exception:
                    pass
            # FieldFox SA panel settings
            if isinstance(widget, FieldFoxSAPanel):
                entry['settings'] = {
                    'center': widget.center_edit.text(),
                    'span': widget.span_edit.text(),
                    'start': widget.start_edit.text(),
                    'stop': widget.stop_edit.text(),
                    'unit': widget.unit_combo.currentText()
                }
            # RhodeSchwarz SMA panel settings
            if isinstance(widget, RhodeSchwarzSMAPanel):
                entry['settings'] = {
                    'frequency': widget.freq_edit.text(),
                    'freq_unit': widget.freq_unit_combo.currentText(),
                    'power': widget.pow_edit.text(),
                    'output': widget.output_btn.isChecked()
                }
            instruments.append(entry)

        # Save sequencing information if available
        seq = []
        use_seq = False
        try:
            if hasattr(self, 'power_seq_builder'):
                seq = self.power_seq_builder.get_sequence()
                use_seq = self.power_seq_builder.use_sequence()
        except Exception:
            seq = []
            use_seq = False

        payload = {
            'instruments': instruments,
            'sequence': seq,
            'use_sequence': use_seq,
            'program_logic_path': self._get_selected_logic_path(),
            'register_read_array': getattr(self, 'register_read_array', [])
        }
        try:
            with open(path, 'w') as f:
                json.dump(payload, f, indent=2)
            self.statusBar().showMessage(f'Saved config to {path}', 4000)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Save failed', str(e))

    def load_config_dialog(self):
        # Load config from selected file in dropdown
        data = self.load_combo.currentData()
        text = self.load_combo.currentText()
        # Guard against placeholder or empty selection
        if (data is None) or (not text) or text.startswith('---'):
            QtWidgets.QMessageBox.information(self, 'No config', 'No config file selected.')
            return
        path = str(data)
        self.load_config(path)

    def load_config(self, path):
        if not os.path.exists(path):
            QtWidgets.QMessageBox.information(self, 'No config', f'Config file not found: {path}')
            return
        # Proactively disconnect and clear any existing instruments/panels
        try:
            self.disconnect_all_instruments()
        except Exception:
            # Fallback: remove tabs if disconnect failed
            while self.tabs.count():
                try:
                    self.tabs.removeTab(0)
                except Exception:
                    break

        # Read config file and extract instruments, sequence, use_sequence
        content = None
        if not os.path.exists(path):
            QtWidgets.QMessageBox.information(self, 'No config', f'Config file not found: {path}')
            return
        try:
            with open(path, 'r') as f:
                content = json.load(f)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Load failed', str(e))
            content = None

        # Build current VISA resource list to support substitutions
        try:
            rm_for_load = pyvisa.ResourceManager()
            current_resources = list(rm_for_load.list_resources())
        except Exception:
            rm_for_load = None
            current_resources = []

        # Backwards compatibility: older configs were lists of instruments
        if content is None:
            instruments = []
            sequence = []
            use_sequence = False
        elif isinstance(content, list):
            instruments = content
            sequence = []
            use_sequence = False
        else:
            instruments = content.get('instruments', [])
            sequence = content.get('sequence', [])
            use_sequence = content.get('use_sequence', False)
            # restore selected program logic file if present
            try:
                logic_path = content.get('program_logic_path', '')
                if logic_path and os.path.exists(logic_path):
                    # Use the path saved in the config if it still exists
                    self._ensure_logic_in_combo(logic_path)
                else:
                    # No path provided or file missing: try to auto-select a matching logic
                    try:
                        auto_logic = self._auto_select_logic_for_config(path, instruments, content)
                    except Exception:
                        auto_logic = ''
                    if auto_logic:
                        self._ensure_logic_in_combo(auto_logic)
                        try:
                            self.statusBar().showMessage(f"Auto-selected logic: {os.path.basename(auto_logic)}", 4000)
                        except Exception:
                            pass
            except Exception:
                pass
            # load register list if present
            try:
                self.register_read_array = content.get('register_read_array', [])
            except Exception:
                self.register_read_array = []

        # Restore sequence builder state from config
        if hasattr(self, 'power_seq_builder'):
            try:
                self.power_seq_builder.set_sequence(sequence)
                self.power_seq_builder.set_use_sequence(use_sequence)
            except Exception:
                pass
        used_resources = set()
        for entry in instruments:
            inst_type = entry.get('type', 'Keithley 2230')
            # Heuristic reclassification for legacy configs where Hittite was saved as 'Unknown'
            if inst_type == 'Unknown':
                name_l = entry.get('name', '').lower()
                # Prefer explicit name cues
                if any(k in name_l for k in ['hittite', 'sig', 'signal']):
                    inst_type = 'Hittite Sig Gen'
                elif any(k in name_l for k in ['rohde', 'schwarz', 'sma', 'rs ']):
                    inst_type = 'RhodeSchwarz SMA'
            resource = entry.get('resource', '')
            # Resolve alias to actual VISA resource if needed
            try:
                if isinstance(resource, str) and resource.startswith('alias:'):
                    alias_key = resource.split(':', 1)[1]
                    # Ensure we have an alias map; if empty, try auto select based on current resources
                    if not self.alias_map:
                        try:
                            rm_tmp = pyvisa.ResourceManager()
                            self._auto_select_alias_profile(list(rm_tmp.list_resources()))
                        except Exception:
                            pass
                    resource = self.alias_map.get(alias_key, '')
            except Exception:
                pass
            # If resource is missing (e.g., alias not mapped), try sensible auto-substitution for known types
            if (not resource) and inst_type in ('Hittite Sig Gen', 'RhodeSchwarz SMA', 'Keysight E36233A', 'Keysight EL34243A', 'Keithley 2230'):
                try:
                    candidates = self._find_replacement_candidates(resource, inst_type, rm_for_load)
                except Exception:
                    candidates = []
                if candidates:
                    resource = candidates[0][1]
                    self.statusBar().showMessage(f'{inst_type} resource missing; auto-connected to: {resource}', 5000)
                else:
                    # leave empty, downstream code may prompt or skip depending on type
                    pass
            # For FieldFox: if resource not available or not detected, auto-connect to another available FieldFox
            if inst_type == 'Keysight FieldFox' and (not resource or resource not in current_resources):
                # Attempt automatic substitution without prompting the user.
                try:
                    candidates = self._find_replacement_candidates(resource, inst_type, rm_for_load)
                except Exception:
                    candidates = []
                if not candidates:
                    # Fresh scan (in case initial resource list was stale)
                    try:
                        fresh_rm = pyvisa.ResourceManager()
                        for res in fresh_rm.list_resources():
                            try:
                                dev = fresh_rm.open_resource(res, timeout=200)
                                try:
                                    idn = dev.query('*IDN?').upper()
                                except Exception:
                                    idn = ''
                                finally:
                                    try:
                                        dev.close()
                                    except Exception:
                                        pass
                                if any(k in idn for k in ['FIELDFOX','N99','HANDHELD SPECTRUM','N991','N992','N993','N994']):
                                    candidates.append((res.split('::')[3] if '::' in res else res, res, 'Keysight FieldFox'))
                                    # don't break; gather all then pick first to allow future heuristics
                            except Exception:
                                continue
                    except Exception:
                        pass
                if candidates:
                    resource = candidates[0][1]
                    self.statusBar().showMessage(f'FieldFox resource missing; auto-connected to: {resource}', 5000)
                else:
                    # No FieldFox found; skip silently with status message.
                    self.statusBar().showMessage('FieldFox from config not found; no FieldFox detected to substitute (skipped).', 5000)
                    continue
            elif resource and resource not in current_resources:
                try:
                    candidates = self._find_replacement_candidates(resource, inst_type, rm_for_load)
                except Exception:
                    candidates = []
                if candidates:
                    items = [f'{c[2]}    ({c[1]})' for c in candidates]
                    item, ok = QtWidgets.QInputDialog.getItem(
                        self, 'Substitute instrument',
                        f'Original resource not found: {resource}\nSelect substitute or Cancel to abort load:', items, 0, False)
                    if not ok:
                        QtWidgets.QMessageBox.information(self, 'Load cancelled', 'Loading cancelled by user')
                        return
                    sel_idx = items.index(item)
                    resource = candidates[sel_idx][1]
                else:
                    resp = QtWidgets.QMessageBox.question(
                        self, 'Instrument missing',
                        f'Instrument {resource} of type {inst_type} not found and no replacements available.\nSkip this instrument? (No = cancel load)',
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.Yes)
                    if resp == QtWidgets.QMessageBox.No:
                        return
                    else:
                        continue
            # Always add FieldFox panel if type matches, even if resource is missing or not detected
            panel = None
            if resource and resource in used_resources:
                for i in range(self.tabs.count()):
                    w = self.tabs.widget(i)
                    if getattr(w, 'resource', None) == resource:
                        panel = w
                        break
            if panel is None:
                if inst_type == 'Keithley 2230':
                    panel = KeithleyPanel(resource)
                elif inst_type == 'Keysight EL34243A':
                    panel = KeysightELPanel(resource)
                elif inst_type == 'Keysight E36233A':
                    panel = KeysightE36233APanel(resource)
                elif inst_type == 'Hittite Sig Gen':
                    panel = HittiteSigGenPanel(resource)
                elif inst_type == 'RhodeSchwarz SMA':
                    panel = RhodeSchwarzSMAPanel(resource)
                elif inst_type == 'Keysight FieldFox':
                    # Use resource or fallback to name if missing
                    visa_addr = resource if resource else entry.get('name', 'FieldFox')
                    panel = FieldFoxSAPanel(visa_addr)
                else:
                    panel = KeithleyPanel(resource)
                tab_label = entry.get('name', resource)
                self.tabs.addTab(panel, tab_label)
                used_resources.add(resource)
                if hasattr(panel, 'resource_edit'):
                    panel.resource_edit.setText(resource)
                if hasattr(panel, 'name_edit'):
                    panel.name_edit.setText(tab_label)
                name_val = entry.get('name') if isinstance(entry, dict) else None
                if name_val and hasattr(panel, 'name_edit'):
                    panel.name_edit.setText(name_val)
                # Auto-connect panel after loading, but defer for FieldFox and Hittite so their
                # specialized blocks can manage connect + apply timing reliably.
                if hasattr(panel, 'on_connect'):
                    try:
                        if inst_type not in ('Keysight FieldFox', 'Hittite Sig Gen'):
                            panel.on_connect()
                    except Exception:
                        pass
                # Set panel values from config
                if inst_type == 'Keithley 2230':
                    for ch in (1, 2, 3):
                        ch_cfg = entry.get('channels', {}).get(str(ch)) or entry.get('channels', {}).get(ch)
                        if ch_cfg:
                            try:
                                # Restore per-channel name if UI supports it
                                if hasattr(panel, 'ch_name_edits') and panel.ch_name_edits.get(ch):
                                    nm = ch_cfg.get('name') if isinstance(ch_cfg, dict) else None
                                    if nm is not None:
                                        panel.ch_name_edits[ch].setText(str(nm))
                                panel.vol_edits[ch].setText(str(ch_cfg.get('voltage', '0')))
                                panel.iam_edits[ch].setText(str(ch_cfg.get('current', '0.03')))
                                panel.master_out_btn.setChecked(bool(ch_cfg.get('output', False)))
                                panel.master_out_btn.setText('All On' if ch_cfg.get('output', False) else 'All Off')
                                # Set instrument output state if connected
                                if getattr(panel, 'inst', None):
                                    panel.inst.set_voltage(ch, float(ch_cfg.get('voltage', '0')))
                                    panel.inst.set_current(ch, float(ch_cfg.get('current', '0.03')))
                                    panel.inst.set_output(ch, bool(ch_cfg.get('output', False)))
                            except Exception:
                                pass
                elif inst_type == 'Keysight EL34243A':
                    # Apply saved per-channel settings directly to the existing panel
                    tab_label = entry.get('name', resource)
                    if hasattr(panel, 'resource_edit'):
                        panel.resource_edit.setText(resource)
                    if hasattr(panel, 'name_edit'):
                        panel.name_edit.setText(tab_label)
                    try:
                        name_val = entry.get('name') if isinstance(entry, dict) else None
                        if name_val and hasattr(panel, 'name_edit'):
                            panel.name_edit.setText(name_val)
                    except Exception:
                        pass
                    # Apply saved channel settings (mode/value/input) for ch 1 and 2
                    try:
                        ch_cfgs = entry.get('channels', {})
                        for ch in (1, 2):
                            ch_cfg = ch_cfgs.get(str(ch)) or ch_cfgs.get(ch)
                            if not ch_cfg:
                                continue
                            # Update per-channel UI fields
                            try:
                                if hasattr(panel, 'mode_combo_ch1') and hasattr(panel, 'mode_combo_ch2'):
                                    combo = panel.mode_combo_ch1 if ch == 1 else panel.mode_combo_ch2
                                    if ch_cfg.get('mode'):
                                        idx = combo.findText(str(ch_cfg.get('mode')))
                                        if idx != -1:
                                            combo.setCurrentIndex(idx)
                                if hasattr(panel, 'mode_value_ch1') and hasattr(panel, 'mode_value_ch2') and (ch_cfg.get('value') is not None):
                                    (panel.mode_value_ch1 if ch == 1 else panel.mode_value_ch2).setText(str(ch_cfg.get('value')))
                                # Restore ramp enable and rise time
                                if hasattr(panel, 'ramp_enable_ch1') and hasattr(panel, 'ramp_enable_ch2'):
                                    ramp_enable = ch_cfg.get('ramp_enabled', False)
                                    (panel.ramp_enable_ch1 if ch == 1 else panel.ramp_enable_ch2).setChecked(bool(ramp_enable))
                                if hasattr(panel, 'rise_time_ch1') and hasattr(panel, 'rise_time_ch2'):
                                    rise_time = ch_cfg.get('rise_time', '')
                                    (panel.rise_time_ch1 if ch == 1 else panel.rise_time_ch2).setText(str(rise_time))
                            except Exception:
                                pass
                            # Apply to hardware if connected
                            if getattr(panel, 'dev', None):
                                try:
                                    mode = str(ch_cfg.get('mode', 'CC'))
                                    try:
                                        value = float(ch_cfg.get('value', '0'))
                                    except Exception:
                                        value = 0.0
                                    panel.dev.set_mode(ch, mode)
                                    panel.dev.set_parameter(ch, mode, value)
                                except Exception:
                                    pass
                                # Input state
                                try:
                                    inp = bool(ch_cfg.get('input', False))
                                    panel.dev.set_input(ch, inp)
                                    if hasattr(panel, '_set_input_toggle_ui'):
                                        panel._set_input_toggle_ui(ch, inp)
                                except Exception:
                                    pass
                    except Exception:
                        pass
                elif inst_type == 'RhodeSchwarz SMA':
                    # Restore Rhode & Schwarz SMA generator settings
                    try:
                        settings = entry.get('settings', {})
                        freq = settings.get('frequency')
                        freq_unit = settings.get('freq_unit')
                        power = settings.get('power')
                        output_state = settings.get('output', False)
                        # Apply GUI state
                        if freq is not None and hasattr(panel, 'freq_edit'):
                            panel.freq_edit.setText(str(freq))
                        if freq_unit and hasattr(panel, 'freq_unit_combo'):
                            idx = panel.freq_unit_combo.findText(str(freq_unit))
                            if idx != -1:
                                panel.freq_unit_combo.setCurrentIndex(idx)
                        if power is not None and hasattr(panel, 'pow_edit'):
                            panel.pow_edit.setText(str(power))
                        if hasattr(panel, 'output_btn'):
                            panel.output_btn.setChecked(bool(output_state))
                            panel.output_btn.setText('Output On' if output_state else 'Output Off')
                        # Defer hardware apply to allow connect handshake
                        def apply_sma_hw():
                            if getattr(panel, 'dev', None):
                                try:
                                    # frequency value is entered along with unit; convert
                                    try:
                                        freq_val = float(panel.freq_edit.text())
                                        unit = panel.freq_unit_combo.currentText()
                                        mult = {'GHz':1e9,'MHz':1e6,'KHz':1e3,'Hz':1}.get(unit,1)
                                        panel.dev.set_frequency(freq_val * mult)
                                    except Exception:
                                        pass
                                    if power is not None:
                                        try:
                                            panel.dev.set_power(float(power))
                                        except Exception:
                                            pass
                                    panel.dev.set_output(bool(output_state))
                                except Exception:
                                    pass
                        QtCore.QTimer.singleShot(1200, apply_sma_hw)
                    except Exception:
                        pass
                elif inst_type == 'Hittite Sig Gen':
                    # Restore Hittite signal generator settings
                    try:
                        settings = entry.get('settings', {})
                        freq = settings.get('frequency')
                        freq_unit = settings.get('freq_unit')
                        power = settings.get('power')
                        output_state = settings.get('output', False)
                        # Ensure panel auto-applies UI settings after a connect
                        try:
                            if hasattr(panel, 'auto_apply_on_connect'):
                                panel.auto_apply_on_connect = False  # we'll explicitly connect and then run our apply routine
                        except Exception:
                            pass
                        if freq is not None and hasattr(panel, 'freq_edit'):
                            panel.freq_edit.setText(str(freq))
                        if freq_unit and hasattr(panel, 'freq_unit_combo'):
                            idx = panel.freq_unit_combo.findText(str(freq_unit))
                            if idx != -1:
                                panel.freq_unit_combo.setCurrentIndex(idx)
                        if power is not None and hasattr(panel, 'pow_edit'):
                            panel.pow_edit.setText(str(power))
                        if hasattr(panel, 'output_btn'):
                            panel.output_btn.setChecked(bool(output_state))
                            panel.output_btn.setText('Output On' if output_state else 'Output Off')
                        # Connect immediately so the user doesn't have to click; apply will be scheduled below
                        try:
                            if hasattr(panel, 'on_connect'):
                                panel.on_connect()
                        except Exception:
                            pass
                        # Ensure connection and apply settings; retry a few times if needed
                        def apply_hittite_hw(attempt_idx: int = 0):
                            try:
                                # Ensure connected
                                if getattr(panel, 'dev', None) is None:
                                    try:
                                        if hasattr(panel, 'on_connect'):
                                            panel.on_connect()
                                    except Exception:
                                        pass
                                # Post-connect settle delay on first configure
                                if attempt_idx == 0:
                                    QtCore.QTimer.singleShot(500, lambda: apply_hittite_hw(attempt_idx + 1))
                                    return
                                # Convert frequency with unit
                                try:
                                    freq_val = float(panel.freq_edit.text())
                                except Exception:
                                    freq_val = None
                                try:
                                    unit = panel.freq_unit_combo.currentText()
                                except Exception:
                                    unit = 'GHz'
                                mult = {'GHz': 1e9, 'MHz': 1e6, 'KHz': 1e3, 'Hz': 1}.get(unit, 1)
                                # Apply frequency
                                if freq_val is not None:
                                    try:
                                        panel.dev.set_frequency(freq_val * mult)
                                    except Exception:
                                        try:
                                            # Secondary try via helper
                                            if hasattr(panel, 'on_set_frequency'):
                                                panel.on_set_frequency()
                                        except Exception:
                                            pass
                                # Apply power
                                if power is not None:
                                    try:
                                        panel.dev.set_power(float(power))
                                    except Exception:
                                        try:
                                            if hasattr(panel, 'on_set_power'):
                                                panel.on_set_power()
                                        except Exception:
                                            pass
                                # Apply output last
                                try:
                                    panel.dev.set_output(bool(output_state))
                                except Exception:
                                    pass
                                # Status hint
                                try:
                                    panel.status_label.setText('Applied JSON freq/power/output')
                                except Exception:
                                    pass
                            except Exception:
                                if attempt_idx < 8:
                                    QtCore.QTimer.singleShot(600, lambda: apply_hittite_hw(attempt_idx + 1))
                        QtCore.QTimer.singleShot(400, apply_hittite_hw)
                        # If already connected for any reason, also schedule a direct UI->HW push as a fallback
                        try:
                            if getattr(panel, 'dev', None) is not None and hasattr(panel, '_apply_ui_to_hw'):
                                QtCore.QTimer.singleShot(800, panel._apply_ui_to_hw)
                        except Exception:
                            pass
                    except Exception:
                        pass
                elif inst_type == 'Keysight FieldFox':
                    try:
                        settings = entry.get('settings', {})
                        center = settings.get('center')
                        span = settings.get('span')
                        start = settings.get('start')
                        stop = settings.get('stop')
                        unit = settings.get('unit')
                        if center is not None:
                            panel.center_edit.setText(str(center))
                        if span is not None:
                            panel.span_edit.setText(str(span))
                        if start is not None:
                            panel.start_edit.setText(str(start))
                        if stop is not None:
                            panel.stop_edit.setText(str(stop))
                        if unit is not None:
                            idx = panel.unit_combo.findText(str(unit))
                            if idx != -1:
                                panel.unit_combo.setCurrentIndex(idx)
                        # Ensure connection and capture thread start after short delay
                        def attempt(idx=0):
                            try:
                                # If already connected, don't re-connect
                                if getattr(panel.sa, 'inst', None) is None and hasattr(panel, 'on_connect'):
                                    panel.on_connect()
                                if getattr(panel.sa, 'inst', None) is None:
                                    raise RuntimeError('Not connected yet')
                                # Clear pending I/O and sync once before applying settings
                                try:
                                    inst = getattr(panel.sa, 'inst', None)
                                    if inst is not None and hasattr(inst, 'clear'):
                                        inst.clear()
                                except Exception:
                                    pass
                                try:
                                    if hasattr(panel.sa, 'sync'):
                                        panel.sa.sync()
                                except Exception:
                                    pass
                                try:
                                    panel.apply_settings()
                                except Exception:
                                    pass
                                # Start capture thread if not already running (idempotent)
                                try:
                                    panel.start_capture_thread()
                                except Exception:
                                    pass
                                return  # success
                            except Exception:
                                pass
                            # retry up to 5 times
                            if idx < 5:
                                QtCore.QTimer.singleShot(500, lambda: attempt(idx+1))
                        QtCore.QTimer.singleShot(400, attempt)
                    except Exception:
                        pass
            # Tab name update logic (no label reference)
            def update_tab_name(panel=panel, resource=resource):
                idx = self.tabs.indexOf(panel)
                if idx != -1:
                    self.tabs.setTabText(idx, getattr(panel, 'name_edit', None).text() if hasattr(panel, 'name_edit') else resource)
            if hasattr(panel, 'name_edit'):
                panel.name_edit.textChanged.connect(lambda: update_tab_name(panel, resource))
                if hasattr(self, 'power_seq_builder'):
                    try:
                        self.power_seq_builder.refresh_instr_combo()
                    except Exception:
                        pass
            update_tab_name(panel, resource)
            self._auto_turn_off_panel(panel)

        # After loading tabs, refresh logic list from disk
        try:
            self._refresh_logic_combo()
        except Exception:
            pass
        # remember last loaded config path for use during recording
        try:
            self._last_loaded_config_path = path
        except Exception:
            pass
    
    def _refresh_seq_instr_combo(self):
        self.seq_instr_combo.clear()
        for i in range(self.tabs.count()):
            name = self.tabs.tabText(i)
            self.seq_instr_combo.addItem(name)

    # Removed manual Add instrument by SN/resource; use VISA scan and Add Selected only

    def on_scan_instruments(self):
        """Scan VISA resources and populate detected_combo with resources. Attempt to classify type by *IDN?"""
        # Indicate scanning in the status bar while we work
        try:
            self.statusBar().showMessage('Scanning instruments please wait')
        except Exception:
            pass
        self.detected_combo.clear()
        import pyvisa
        import concurrent.futures
        rm = pyvisa.ResourceManager()
        try:
            resources = list(rm.list_resources())
        except Exception:
            resources = []

        categories = {
            'Keithley 2230': [],
            'Keysight EL34243A': [],
            'Keysight E36233A': [],
            'Hittite Sig Gen': [],
            'RhodeSchwarz SMA': [],
            'Keysight FieldFox': [],
            'Unknown': []
        }

        def scan_resource(res):
            try:
                inst = rm.open_resource(res, timeout=3000)
                try:
                    # Clear any pending I/O and status before IDN query
                    try:
                        inst.clear()
                    except Exception:
                        pass
                    try:
                        inst.write("*CLS")
                        _ = inst.query("*OPC?")
                    except Exception:
                        pass
                    idn = inst.query("*IDN?").strip()
                finally:
                    try:
                        inst.close()
                    except Exception:
                        pass
                if 'Keithley' in idn and '2230' in idn:
                    return ('Keithley 2230', res, 'Keithley 2230')
                elif 'Keysight' in idn and ('EL34243A' in idn or 'Electronic Load' in idn):
                    return ('Keysight EL34243A', res, 'Keysight EL34243A')
                elif 'Keysight' in idn and ('E36233A' in idn or 'Power Supply' in idn):
                    return ('Keysight E36233A', res, 'Keysight E36233A')
                elif 'Keysight' in idn and ('FieldFox' in idn or 'N99' in idn or 'Handheld Spectrum' in idn):
                    return ('Keysight FieldFox', res, 'Keysight FieldFox')
                elif 'Hittite' in idn or 'Sig Gen' in idn:
                    return ('Hittite Sig Gen', res, 'Hittite Sig Gen')
                elif 'Rhode' in idn or 'Schwarz' in idn or 'SMA' in idn:
                    return ('RhodeSchwarz SMA', res, 'RhodeSchwarz SMA')
                else:
                    return ('Unknown', res, 'Unknown')
            except Exception:
                return ('Unknown', res, 'Unknown')

        results = []
        if resources:
            with concurrent.futures.ThreadPoolExecutor(max_workers=min(8, len(resources))) as executor:
                future_to_res = {executor.submit(scan_resource, res): res for res in resources}
                for future in concurrent.futures.as_completed(future_to_res):
                    cat, res, inst_type = future.result()
                    categories[cat].append((res, inst_type))

        total_found = sum(len(v) for v in categories.values())
        if total_found == 0:
            self.detected_combo.addItem('No devices found', None)

        # Build reverse alias map for quick lookup: resource -> alias
        try:
            rev_alias = {v: k for k, v in (self.alias_map or {}).items()}
        except Exception:
            rev_alias = {}

        # Helper to map instrument type to alias base used in profiles
        def _alias_base_for_type(_t: str) -> str:
            t = (_t or '').strip()
            if t == 'Keithley 2230':
                return 'Keithley'
            if t == 'Keysight FieldFox':
                return 'FieldFox'
            if t == 'Keysight EL34243A':
                return 'Keysight EL'
            if t == 'Keysight E36233A':
                return 'E36233A'
            if t == 'Hittite Sig Gen':
                return 'Hittite'
            if t == 'RhodeSchwarz SMA':
                return 'RohdeSchwarz SMA'
            return 'Unknown'

        # To keep fallback aliases unique per base
        alias_counters = {}

        for cat, items in categories.items():
            for res, inst_type in items:
                # Prefer current profile alias if available; otherwise generate a sensible alias
                alias_name = rev_alias.get(res)
                if not alias_name:
                    base = _alias_base_for_type(inst_type)
                    idx = alias_counters.get(base, 0)
                    alias_name = f"{base} {idx}"
                    alias_counters[base] = idx + 1
                # Show alias in the dropdown; store resource, type, and alias
                self.detected_combo.addItem(alias_name, (res, inst_type, alias_name))

        # Apply aliasing choice from startup prompt
        try:
            mode = getattr(self, '_alias_startup_choice', '')
            if mode == 'preconfigured' and getattr(self, 'alias_profile', ''):
                # Keep the loaded profile; do not auto-select/override
                pass
            elif mode == 'generic':
                # Build a generic alias map from detected resources
                self._build_aliases_from_resources(resources)
            else:
                # Fallback behavior: auto-select the best matching profile
                self._auto_select_alias_profile(resources)
        except Exception:
            pass

        try:
            if hasattr(self, 'power_seq_builder'):
                self.power_seq_builder.refresh_instr_combo()
        except Exception:
            pass

        # Show found count briefly, then return to Ready
        try:
            self.statusBar().showMessage(f'Found {total_found} device(s)', 3000)
            QtCore.QTimer.singleShot(3000, lambda: self.statusBar().showMessage('Ready'))
        except Exception:
            try:
                self.statusBar().showMessage('Ready')
            except Exception:
                pass

    def _prompt_alias_startup(self):
        """Prompt user at startup to choose aliasing approach before any VISA scan."""
        # Only prompt once per app run
        if getattr(self, '_alias_prompt_done', False):
            return
        self._alias_prompt_done = True
        # Simple dialog with two explicit choices
        box = QtWidgets.QMessageBox(self)
        box.setWindowTitle('Aliasing Profile')
        box.setText('Choose an aliasing profile for this session:')
        generic_btn = box.addButton('Generic (auto-suggest)', QtWidgets.QMessageBox.AcceptRole)
        preconf_btn = box.addButton('Load preconfigured…', QtWidgets.QMessageBox.ActionRole)
        box.setIcon(QtWidgets.QMessageBox.Question)
        box.exec_()

        clicked = box.clickedButton()
        if clicked is preconf_btn:
            # Let user pick from existing profiles
            try:
                profiles = self._list_alias_profiles() or []
            except Exception:
                profiles = []
            if not profiles:
                QtWidgets.QMessageBox.information(self, 'No profiles', 'No preconfigured profiles found. Using generic aliasing.')
                self._alias_startup_choice = 'generic'
            else:
                name, ok = QtWidgets.QInputDialog.getItem(self, 'Load Profile', 'Select a profile:', profiles, 0, False)
                if ok and name:
                    try:
                        self._load_alias_profile(name)
                        self._alias_startup_choice = 'preconfigured'
                    except Exception as e:
                        QtWidgets.QMessageBox.warning(self, 'Load failed', f'Failed to load profile: {e}\nUsing generic instead.')
                        self._alias_startup_choice = 'generic'
                else:
                    # Fallback to generic if user cancels selection
                    self._alias_startup_choice = 'generic'
        else:
            # Default to generic
            self._alias_startup_choice = 'generic'

        # After choosing, kick off the first scan
        try:
            self.on_scan_instruments()
        except Exception:
            pass

    def on_add_selected_instrument(self):
        data = self.detected_combo.currentData()
        if not data:
            QtWidgets.QMessageBox.information(self, 'No selection', 'Select a detected device first (Scan VISA).')
            return
        # Unpack resource, instrument type, and the alias label we displayed
        try:
            resource, inst_type, alias_label = data
        except Exception:
            # Backward compatibility if old tuple shape
            resource, inst_type = data
            alias_label = resource
        # Use the alias label for the tab name
        label = alias_label
        self.add_instrument_panel(resource, label, inst_type)
        # refresh sequence builder after adding a tab
        try:
            if hasattr(self, 'power_seq_builder'):
                self.power_seq_builder.refresh_instr_combo()
        except Exception:
            pass
        self.statusBar().showMessage(f'Added {label}', 3000)

    def close_tab(self, index):
        widget = self.tabs.widget(index)
        try:
            widget.close()
        except Exception:
            pass
        self.tabs.removeTab(index)
        self.power_seq_builder.refresh_instr_combo()

    def closeEvent(self, event):
        # Ensure all background recordings are stopped and workbook is saved before exit
        try:
            self._stop_all_recordings()
        except Exception:
            pass
        # close all panels to ensure instruments are closed
        for i in range(self.tabs.count()):
            widget = self.tabs.widget(i)
            try:
                widget.close()
            except Exception:
                pass
        event.accept()

    def _update_global_power_btns(self, on):
        if hasattr(self, 'test_power_toggle_btn'):
            if on:
                self.test_power_toggle_btn.setStyleSheet('background-color: #4CAF50; color: white;')
            else:
                self.test_power_toggle_btn.setStyleSheet('background-color: #F44336; color: white;')

    def _update_test_power_toggle_btn(self, on):
        # Update both Sequencing tab primary button and the Test tab button (if present)
        if hasattr(self, 'test_power_toggle_btn'):
            if on:
                self.test_power_toggle_btn.setText('Power On All')
                self.test_power_toggle_btn.setStyleSheet('background-color: #4CAF50; color: white;')
            else:
                self.test_power_toggle_btn.setText('Power Off All')
                self.test_power_toggle_btn.setStyleSheet('background-color: #F44336; color: white;')
            try:
                self.test_power_toggle_btn.setChecked(on)
            except Exception:
                pass
        if hasattr(self, 'test_tab_power_btn'):
            try:
                self.test_tab_power_btn.setChecked(on)
                if on:
                    self.test_tab_power_btn.setText('Power On All')
                    self.test_tab_power_btn.setStyleSheet('background-color: #4CAF50; color: white;')
                else:
                    self.test_tab_power_btn.setText('Power Off All')
                    self.test_tab_power_btn.setStyleSheet('background-color: #F44336; color: white;')
            except Exception:
                pass

    def on_test_power_toggle(self):
        # Use the clicked button's checked state if possible so the button behaves as a toggle.
        on = None
        try:
            sender = self.sender()
            if sender is not None:
                try:
                    on = bool(sender.isChecked())
                except Exception:
                    on = None
        except Exception:
            on = None

        # Fallback: if we couldn't get the sender, derive from either button
        if on is None:
            try:
                on = False
                if hasattr(self, 'test_power_toggle_btn') and self.test_power_toggle_btn.isChecked():
                    on = True
                if hasattr(self, 'test_tab_power_btn') and self.test_tab_power_btn.isChecked():
                    on = True
            except Exception:
                on = False

        # Synchronize both UI buttons to the chosen state
        try:
            if hasattr(self, 'test_power_toggle_btn'):
                self.test_power_toggle_btn.setChecked(on)
            if hasattr(self, 'test_tab_power_btn'):
                self.test_tab_power_btn.setChecked(on)
        except Exception:
            pass

        # Perform the action for the chosen state
        if on:
            # Power ON (use sequence if configured)
            try:
                self._update_test_power_toggle_btn(True)
            except Exception:
                pass
            try:
                self._run_power_sequence()
            except Exception:
                pass
        else:
            # Power OFF
            try:
                self._update_test_power_toggle_btn(False)
            except Exception:
                pass
            try:
                self.power_off_all()
            except Exception:
                pass

    def _run_power_sequence(self, on_complete=None):
        """Run the configured power-up sequence. If on_complete is provided, it will be
        called (safely via QTimer) when the sequence finishes.
        """
        sequence = self.power_seq_builder.get_sequence()
        use_seq = self.power_seq_builder.use_sequence()
        if use_seq and sequence:
            def run_step(idx):
                # check global abort flag
                if getattr(self, '_sequence_abort_flag', False):
                    self._log('Sequence aborted by user')
                    # reset flag for next run
                    try:
                        self._sequence_abort_flag = False
                    except Exception:
                        pass
                    if on_complete:
                        try:
                            QtCore.QTimer.singleShot(100, on_complete)
                        except Exception:
                            try:
                                on_complete()
                            except Exception:
                                pass
                    return
                if idx >= len(sequence):
                    self.statusBar().showMessage('Sequence power on complete', 4000)
                    if on_complete:
                        # call back shortly after finishing to keep UI responsive
                        try:
                            QtCore.QTimer.singleShot(100, on_complete)
                        except Exception:
                            try:
                                on_complete()
                            except Exception:
                                pass
                    return
                item = sequence[idx]
                self._log(f'Processing step: {item}')
                if item.startswith('KeithleyChannel: '):
                    # Sequence step for individual Keithley channel
                    name = item[len('KeithleyChannel: '):]
                    # name format: 'Keithley ... Channel N'
                    if 'Channel' in name:
                        parts = name.split('Channel')
                        tab_name = parts[0].strip()
                        try:
                            ch = int(parts[1].strip())
                        except Exception:
                            ch = None
                        for i in range(self.tabs.count()):
                            if self.tabs.tabText(i) == tab_name:
                                widget = self.tabs.widget(i)
                                if isinstance(widget, KeithleyPanel) and widget.inst and ch:
                                    try:
                                        V = float(widget.vol_edits[ch].text())
                                        I = float(widget.iam_edits[ch].text())
                                    except Exception:
                                        V, I = 0.0, 0.03
                                    # Set voltage/current for requested channel
                                    widget.inst.set_voltage(ch, V)
                                    widget.inst.set_current(ch, I)
                                    # Set voltage/current to zero for other channels, but do not toggle their output off
                                    for other_ch in (1, 2, 3):
                                        if other_ch != ch and not widget.output_btns[other_ch].isChecked():
                                            widget.inst.set_voltage(other_ch, 0.0)
                                            widget.inst.set_current(other_ch, 0.0)
                                    # Only enable output if not already on
                                    if not widget.output_btns[ch].isChecked():
                                        widget.inst.set_output(ch, True)
                                    widget.output_btns[ch].setChecked(True)
                                    widget.output_btns[ch].setText('Output On')
                                break
                    QtCore.QTimer.singleShot(100, lambda: run_step(idx + 1))
                elif item.startswith('Instrument: '):
                    name = item[len('Instrument: '):]
                    for i in range(self.tabs.count()):
                        tab_name = self.tabs.tabText(i)
                        if tab_name == name:
                            widget = self.tabs.widget(i)
                            if isinstance(widget, KeithleyPanel):
                                if widget.inst:
                                    for ch in (1, 2, 3):
                                        try:
                                            V = float(widget.vol_edits[ch].text())
                                            I = float(widget.iam_edits[ch].text())
                                            widget.inst.set_voltage(ch, V)
                                            widget.inst.set_current(ch, I)
                                            widget.inst.set_output(ch, True)
                                            widget.output_btns[ch].setChecked(True)
                                            widget.output_btns[ch].setText('Output On')
                                        except Exception:
                                            pass
                                    if hasattr(widget, 'master_out_btn'):
                                        widget.master_out_btn.setChecked(True)
                                        widget.master_out_btn.setText('All On')
                            elif isinstance(widget, HittiteSigGenPanel):
                                if widget.dev:
                                    try:
                                        widget.dev.set_output(True)
                                        widget.output_btn.setChecked(True)
                                        widget.output_btn.setText('Output On')
                                        widget.status_label.setText('Output ON')
                                    except Exception as e:
                                        widget.status_label.setText(f'Failed to set output: {e}')
                            elif isinstance(widget, KeysightELPanel):
                                if widget.dev:
                                    for ch in (1, 2):
                                        if hasattr(widget, 'ch_enabled') and not widget.ch_enabled[ch].isChecked():
                                            continue
                                        try:
                                            # Use per-channel controls and ramp logic
                                            mode_combo = getattr(widget, f'mode_combo_ch{ch}', None)
                                            mode_value_edit = getattr(widget, f'mode_value_ch{ch}', None)
                                            ramp_enable = getattr(widget, f'ramp_enable_ch{ch}', None)
                                            rise_time_edit = getattr(widget, f'rise_time_ch{ch}', None)
                                            input_toggle = getattr(widget, f'input_toggle_ch{ch}', None)
                                            if not mode_combo or not mode_value_edit:
                                                continue
                                            mode = mode_combo.currentText()
                                            try:
                                                value = float(mode_value_edit.text())
                                            except Exception:
                                                value = 0.0
                                            ramp_enabled = ramp_enable.isChecked() if ramp_enable else False
                                            try:
                                                rise_time = float(rise_time_edit.text()) if rise_time_edit else 0.0
                                            except Exception:
                                                rise_time = 0.0
                                            if mode == 'Disable':
                                                widget.dev.set_input(ch, False)
                                                if input_toggle:
                                                    input_toggle.setChecked(False)
                                                    input_toggle.setText('Input Off')
                                                continue
                                            widget.dev.set_mode(ch, mode)
                                            widget.dev.set_input(ch, True)
                                            if ramp_enabled and rise_time > 0:
                                                import numpy as np
                                                import time
                                                steps = 20
                                                for v in np.linspace(0, value, steps):
                                                    widget.dev.set_parameter(ch, mode, v)
                                                    QtWidgets.QApplication.processEvents()
                                                    time.sleep(rise_time / steps)
                                                widget.dev.set_parameter(ch, mode, value)
                                            else:
                                                widget.dev.set_parameter(ch, mode, value)
                                            if input_toggle:
                                                input_toggle.setChecked(True)
                                                input_toggle.setText('Input On')
                                        except Exception:
                                            pass
                            elif isinstance(widget, RhodeSchwarzSMAPanel):
                                if widget.dev:
                                    try:
                                        widget.dev.set_output(True)
                                        widget.output_btn.setChecked(True)
                                        widget.output_btn.setText('Output On')
                                        widget.status_label.setText('Output ON')
                                    except Exception as e:
                                        widget.status_label.setText(f'Failed to set output: {e}')
                            elif isinstance(widget, FieldFoxSAPanel):
                                if widget.dev:
                                    try:
                                        widget.dev.set_output(True)
                                        widget.output_btn.setChecked(True)
                                        widget.output_btn.setText('Output On')
                                        widget.status_label.setText('Output ON')
                                    except Exception as e:
                                        widget.status_label.setText(f'Failed to set output: {e}')
                    QtCore.QTimer.singleShot(100, lambda: run_step(idx + 1))
                elif item.startswith('Delay: '):
                    delay_val = float(item[len('Delay: '):-3])
                    self._log(f'Waiting {delay_val} s')
                    QtCore.QTimer.singleShot(int(delay_val * 1000), lambda: run_step(idx + 1))
                else:
                    run_step(idx + 1)
            run_step(0)
        else:
            # No sequence configured or sequence disabled: just power on all now
            self.power_on_all()
            # If caller provided a completion callback, call it shortly after powering on
            if on_complete:
                try:
                    QtCore.QTimer.singleShot(100, on_complete)
                except Exception:
                    try:
                        on_complete()
                    except Exception:
                        pass

    def soft_reset(self):
        """Soft Reset: run the same Configure Part logic as the Programming page."""
        logic_path = self._get_selected_logic_path()
        if not logic_path:
            QtWidgets.QMessageBox.information(self, 'No logic selected', 'Please select a logic .py file on the Programming tab.')
            return
        if not os.path.exists(logic_path):
            QtWidgets.QMessageBox.warning(self, 'Missing file', f'Logic file not found:\n{logic_path}')
            return
        self.run_configure_part(logic_path)

    # --- Configure Part (external logic) ---
    def _refresh_logic_combo(self):
        try:
            if not hasattr(self, 'logic_combo'):
                return
            # Preserve current selection
            try:
                prev_path = self.logic_combo.currentData()
                prev_path = prev_path if isinstance(prev_path, str) else ''
            except Exception:
                prev_path = ''

            self.logic_combo.clear()
            self.logic_combo.addItem('(Select a logic file)', '')
            try:
                files = [f for f in os.listdir(self.program_logic_dir) if f.lower().endswith('.py')]
            except Exception:
                files = []
            for f in sorted(files):
                full = os.path.join(self.program_logic_dir, f)
                self.logic_combo.addItem(f, full)
            # If previously selected path is outside the logic dir, keep it if it still exists
            if prev_path and os.path.exists(prev_path):
                # Check if it's already present
                found_idx = -1
                for i in range(self.logic_combo.count()):
                    if self.logic_combo.itemData(i) == prev_path:
                        found_idx = i
                        break
                if found_idx == -1:
                    # Add as custom entry
                    self.logic_combo.addItem(os.path.basename(prev_path), prev_path)
                    found_idx = self.logic_combo.count() - 1
                # Restore selection
                if found_idx != -1:
                    self.logic_combo.setCurrentIndex(found_idx)
        except Exception:
            pass

    def _ensure_logic_in_combo(self, path: str):
        if not path:
            return
        try:
            # If already present, select it. Otherwise add as custom entry.
            for i in range(self.logic_combo.count()):
                if self.logic_combo.itemData(i) == path:
                    self.logic_combo.setCurrentIndex(i)
                    return
            # Add custom path
            self.logic_combo.addItem(os.path.basename(path), path)
            self.logic_combo.setCurrentIndex(self.logic_combo.count() - 1)
        except Exception:
            pass

    def _browse_logic_file(self):
        start_dir = self.program_logic_dir if os.path.isdir(self.program_logic_dir) else os.path.expanduser('~')
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Select logic file', start_dir, 'Python Files (*.py)')
        if not fname:
            return
        # If file not in our logic dir, offer to copy
        try:
            target = os.path.join(self.program_logic_dir, os.path.basename(fname))
            if os.path.abspath(os.path.dirname(fname)) != os.path.abspath(self.program_logic_dir):
                resp = QtWidgets.QMessageBox.question(
                    self, 'Copy logic file?',
                    f'Copy selected file into:\n{self.program_logic_dir}\nfor easier reuse?',
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.Yes)
                if resp == QtWidgets.QMessageBox.Yes:
                    try:
                        import shutil
                        os.makedirs(self.program_logic_dir, exist_ok=True)
                        shutil.copy2(fname, target)
                        fname = target
                    except Exception as e:
                        QtWidgets.QMessageBox.warning(self, 'Copy failed', str(e))
            # Ensure in combo and select
            self._ensure_logic_in_combo(fname)
        except Exception:
            pass

    def _get_selected_logic_path(self) -> str:
        try:
            data = self.logic_combo.currentData()
            return data if isinstance(data, str) else ''
        except Exception:
            return ''

    def _auto_select_logic_for_config(self, config_path: str, instruments: list, content: dict) -> str:
        """Heuristically pick a programming logic .py in program_logic_dir based on
        the config filename and instrument names/types. Returns full path or ''."""
        try:
            logic_dir = getattr(self, 'program_logic_dir', os.path.dirname(__file__))
            try:
                files = [f for f in os.listdir(logic_dir) if f.lower().endswith('.py')]
            except Exception:
                files = []
            if not files:
                return ''

            def _norm(s: str) -> str:
                return ''.join(ch for ch in str(s).lower() if ch.isalnum())

            cfg_base = os.path.splitext(os.path.basename(config_path or ''))[0]
            cfg_tok = _norm(cfg_base)
            # 1) Prefer exact base-name match: product.json -> product.py (case-insensitive)
            if cfg_base:
                base_lower = cfg_base.lower()
                for f in files:
                    if os.path.splitext(f)[0].lower() == base_lower:
                        return os.path.join(logic_dir, f)
            tokens = set()
            if cfg_tok:
                tokens.add(cfg_tok)
            # Add instrument names and types as tokens
            try:
                for ent in instruments or []:
                    nm = ent.get('name') if isinstance(ent, dict) else None
                    tp = ent.get('type') if isinstance(ent, dict) else None
                    if nm:
                        t = _norm(nm)
                        if t:
                            tokens.add(t)
                    if tp:
                        t = _norm(tp)
                        if t:
                            tokens.add(t)
            except Exception:
                pass
            # Include any explicit part hint fields, if present
            try:
                for k in ('part', 'device', 'dut', 'design'):
                    v = content.get(k)
                    if v:
                        t = _norm(v)
                        if t:
                            tokens.add(t)
            except Exception:
                pass

            candidates = []  # (score, mtime, fullpath)
            for f in files:
                full = os.path.join(logic_dir, f)
                base = os.path.splitext(f)[0]
                bn = _norm(base)
                try:
                    mt = os.path.getmtime(full)
                except Exception:
                    mt = 0
                score = 0
                if bn and cfg_tok and bn == cfg_tok:
                    score = 3  # exact match to config name
                elif any(bn == t for t in tokens):
                    score = 3  # exact token match
                elif any((bn in t) or (t in bn) for t in tokens):
                    score = 2  # partial contains
                # else score remains 0
                candidates.append((score, mt, full))

            # Prefer highest score; break ties by newest mtime
            candidates.sort(key=lambda x: (x[0], x[1]))
            best = candidates[-1] if candidates else None
            if best and best[0] > 0:
                return best[2]
            # As a gentle fallback, pick newest overall, but only if there is a single file
            # or user likely expects auto-selection when only one logic exists.
            if len(files) == 1:
                return os.path.join(logic_dir, files[0])
            return ''
        except Exception:
            return ''

    def on_configure_part_clicked(self):
        logic_path = self._get_selected_logic_path()
        if not logic_path:
            QtWidgets.QMessageBox.information(self, 'No logic selected', 'Please select a logic .py file to run.')
            return
        if not os.path.exists(logic_path):
            QtWidgets.QMessageBox.warning(self, 'Missing file', f'Logic file not found:\n{logic_path}')
            return
        self.run_configure_part(logic_path)

    def run_configure_part(self, logic_path: str):
        # Run logic directly in a background thread (no temp file, no subprocess)
        if self._logic_thread is not None and self._logic_thread.is_alive():
            self._log('Configure Part is already running')
            return

        self._logic_abort = False
        self._log(f'Starting Configure Part with logic: {os.path.basename(logic_path)}')

        def post_log(msg: str):
            try:
                QtCore.QTimer.singleShot(0, lambda m=msg: self._log(m))
            except Exception:
                pass

        def worker():
            import sys as _sys
            import importlib.util as _importlib_util
            import threading as _th
            import time as _time
            import logging as _logging
            # Forward prints from logic to the GUI logs
            class _StreamForwarder:
                def __init__(self, cb):
                    self._cb = cb
                    self._buf = ''
                    self._lock = _th.Lock()
                def write(self, s):
                    if not s:
                        return
                    with self._lock:
                        self._buf += str(s)
                        while '\n' in self._buf:
                            line, self._buf = self._buf.split('\n', 1)
                            if line.strip():
                                try:
                                    self._cb(line)
                                except Exception:
                                    pass
                def flush(self):
                    # Flush any partial line as-is to keep UI fresh
                    with self._lock:
                        if self._buf.strip():
                            try:
                                self._cb(self._buf)
                            except Exception:
                                pass
                        self._buf = ''
            try:
                # Ensure ACE Client path is available
                ace_path = r'C:\\Program Files\\Analog Devices\\ACE\\Client'
                if ace_path not in _sys.path:
                    _sys.path.append(ace_path)
                import clr  # type: ignore
                clr.AddReference('AnalogDevices.Csa.Remoting.Clients')
                clr.AddReference('AnalogDevices.Csa.Remoting.Contracts')
                from AnalogDevices.Csa.Remoting.Clients import ClientManager  # type: ignore
                manager = ClientManager.Create()
                client = manager.CreateRequestClient('localhost:2357')
                # Redirect stdout/stderr so print() in logic shows in the GUI
                _old_out, _old_err = _sys.stdout, _sys.stderr
                _fw = _StreamForwarder(post_log)
                _sys.stdout = _fw
                _sys.stderr = _fw
                # Hook Python logging to the same stream so logging.info/etc show up
                _log_handler = _logging.StreamHandler(_fw)
                _log_handler.setLevel(_logging.DEBUG)
                _log_handler.setFormatter(_logging.Formatter('%(message)s'))
                _root_logger = _logging.getLogger()
                _root_logger.addHandler(_log_handler)
                # Periodic flusher to surface partial lines at ~0.5s cadence
                _stop_evt = _th.Event()
                def _periodic_flush():
                    while not _stop_evt.is_set():
                        try:
                            _fw.flush()
                        except Exception:
                            pass
                        _time.sleep(0.5)
                _flush_thread = _th.Thread(target=_periodic_flush, daemon=True)
                _flush_thread.start()
                # Use signal-based logger for thread-safe UI update
                try:
                    self._log('ACE connection established; running logic...')
                except Exception:
                    # Last-resort: attempt direct append if logs exist
                    try:
                        if hasattr(self, 'test_log'):
                            self.test_log.appendPlainText('ACE connection established; running logic...')
                        if hasattr(self, 'program_log'):
                            self.program_log.appendPlainText('ACE connection established; running logic...')
                    except Exception:
                        pass
                # Load logic module from given path
                spec = _importlib_util.spec_from_file_location('selected_logic', logic_path)
                mod = _importlib_util.module_from_spec(spec)
                assert spec and spec.loader
                spec.loader.exec_module(mod)  # type: ignore
                func = None
                if hasattr(mod, 'execute_macro'):
                    func = getattr(mod, 'execute_macro')
                elif hasattr(mod, 'logic'):
                    func = getattr(mod, 'logic')
                if func is None:
                    self._log('Selected logic file must define execute_macro(client) or logic(client)')
                    # Restore streams before returning
                    try:
                        _fw.flush()
                    except Exception:
                        pass
                    _sys.stdout, _sys.stderr = _old_out, _old_err
                    return
                try:
                    func(client)
                finally:
                    self._log('Configure Part finished')
                    # Flush any buffered text and restore streams
                    try:
                        _fw.flush()
                    except Exception:
                        pass
                    # Stop periodic flusher
                    try:
                        _stop_evt.set()
                        _flush_thread.join(timeout=1.0)
                    except Exception:
                        pass
                    # Detach logging handler
                    try:
                        _root_logger.removeHandler(_log_handler)
                    except Exception:
                        pass
                    _sys.stdout, _sys.stderr = _old_out, _old_err                    
            except Exception as e:
                post_log(f'Configure Part error: {e}')

        import threading as _threading
        self._logic_thread = _threading.Thread(target=worker, daemon=True)
        self._logic_thread.start()

        # Update UI while running
        try:
            if hasattr(self, 'abort_program_btn'):
                self.abort_program_btn.setEnabled(True)
            if hasattr(self, 'run_program_btn'):
                self.run_program_btn.setEnabled(False)
            if hasattr(self, 'program_now_btn'):
                self.program_now_btn.setEnabled(False)
            if hasattr(self, 'configure_part_btn'):
                self.configure_part_btn.setEnabled(False)
        except Exception:
            pass

        # Poll thread to re-enable UI on completion
        try:
            if self._logic_poll_timer is None:
                self._logic_poll_timer = QtCore.QTimer(self)
                self._logic_poll_timer.setInterval(300)
                self._logic_poll_timer.timeout.connect(self._on_logic_thread_check)
            if not self._logic_poll_timer.isActive():
                self._logic_poll_timer.start()
        except Exception:
            pass

    def _on_logic_thread_check(self):
        try:
            alive = self._logic_thread is not None and self._logic_thread.is_alive()
        except Exception:
            alive = False
        if not alive:
            try:
                if self._logic_poll_timer and self._logic_poll_timer.isActive():
                    self._logic_poll_timer.stop()
            except Exception:
                pass
            self._logic_thread = None
            # Reset UI buttons
            try:
                if hasattr(self, 'run_program_btn'):
                    self.run_program_btn.setEnabled(True)
                if hasattr(self, 'abort_program_btn'):
                    self.abort_program_btn.setEnabled(False)
                if hasattr(self, 'program_now_btn'):
                    self.program_now_btn.setEnabled(True)
                if hasattr(self, 'configure_part_btn'):
                    self.configure_part_btn.setEnabled(True)
            except Exception:
                pass

# Application entry point (restored)
if __name__ == '__main__':
    import sys
    from PyQt5 import QtWidgets
    try:
        app = QtWidgets.QApplication(sys.argv)
        win = MainWindow()
        win.show()
        sys.exit(app.exec_())
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f'Fatal startup error: {e}', file=sys.stderr)
        sys.exit(1)