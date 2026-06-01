import threading
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

class SupplyRecorder:
    def __init__(self, panels, excel_path, sheet_name="supply reads"):
        """
        panels: list of instrument panels (KeithleyPanel, KeysightE36233APanel, etc.)
        excel_path: path to excel file
        sheet_name: name of sheet to store readings
        """
        self.panels = panels
        self.get_readings_funcs = [p.get_all_readings for p in panels]
        self.panel_names = []
        for p in panels:
            name = getattr(p, 'name_edit', None)
            if name and name.text().strip():
                self.panel_names.append(name.text().strip())
            else:
                # fallback to serial/resource
                self.panel_names.append(getattr(p, 'resource', 'Unknown'))
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self._stop_event = threading.Event()
        self._thread = None
        self.buffer = []
        self.last_perf_time = None
        self.read_count = 0

    def start(self):
        if self._thread and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_event.set()
        if self._thread:
            self._thread.join()
        # Always flush and save buffer immediately
        self._flush_buffer()

    def _run(self):
        self.last_perf_time = time.time()
        self._recording_start_time = time.time()
        self.read_count = 0
        while not self._stop_event.is_set():
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            elapsed_s = round(time.time() - self._recording_start_time, 3)
            all_voltages = []
            all_currents = []
            for func in self.get_readings_funcs:
                voltages, currents = func()
                all_voltages.extend(voltages)
                all_currents.extend(currents)
            row = [now, elapsed_s] + all_voltages + all_currents
            self.buffer.append(row)
            self.read_count += 1
            # Performance reporting every 10 seconds
            if time.time() - self.last_perf_time >= 10:
                avg_rps = self.read_count / (time.time() - self.last_perf_time)
                print(f"[SupplyRecorder] Average reads/sec: {avg_rps:.2f} over last 10 seconds")
                self.last_perf_time = time.time()
                self.read_count = 0
                self._flush_buffer()
            time.sleep(0.01)  # ~100Hz, adjust as needed
        self._flush_buffer()

    def _flush_buffer(self):
        if not self.buffer:
            return
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
            if self.sheet_name in wb.sheetnames:
                ws = wb[self.sheet_name]
            else:
                ws = wb.create_sheet(self.sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = self.sheet_name
            # Write header matching row order: [timestamp, elapsed_s] + all voltages (all panels) + all currents (all panels)
            header = ["timestamp", "elapsed_s"]
            v_headers = []
            i_headers = []
            for idx, func in enumerate(self.get_readings_funcs):
                v, c = func()
                name = self.panel_names[idx]
                labels = getattr(self.panels[idx], 'channel_labels', None)
                for i in range(len(v)):
                    if labels and i < len(labels):
                        v_headers.append(f"{name} {labels[i]} V")
                    else:
                        v_headers.append(f"{name} V{i+1}")
                for i in range(len(c)):
                    if labels and i < len(labels):
                        i_headers.append(f"{name} {labels[i]} I")
                    else:
                        i_headers.append(f"{name} I{i+1}")
            ws.append(header + v_headers + i_headers)
        for row in self.buffer:
            ws.append(row)
        wb.save(self.excel_path)
        self.buffer.clear()
