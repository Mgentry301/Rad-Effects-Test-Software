"""
Excel workbook management mixin for MainWindow.

Provides shared workbook open/save/sheet helpers, save-status timer,
bench-info writing, and artifact cleanup.
"""
import os
import json
import datetime
import time

from PyQt5 import QtCore, QtWidgets

from Support_Scrips.Front_Panels.keithley_panel import KeithleyPanel
from Support_Scrips.Front_Panels.keysightEL_panel import KeysightELPanel
from Support_Scrips.Front_Panels.keysight_e36233a_panel import KeysightE36233APanel
from Support_Scrips.Front_Panels.hittite_siggen_panel import HittiteSigGenPanel
from Support_Scrips.Front_Panels.rhodeschwarz_sma_panel import RhodeSchwarzSMAPanel
from Support_Scrips.Front_Panels.fieldfox_sa_panel import FieldFoxSAPanel


class ExcelMixin:
    """Mixin that adds Excel workbook management to MainWindow."""

    # ---- workbook open / sheet / save (caller must hold _excel_lock) ----

    def _excel_open_locked(self, excel_path: str):
        """Open or return a cached workbook for this session. Caller must hold _excel_lock."""
        try:
            from openpyxl import Workbook, load_workbook
            try:
                target_path = getattr(self, '_excel_working_path', None) or excel_path
            except Exception:
                target_path = excel_path
            if getattr(self, '_excel_wb', None) is None or getattr(self, '_excel_wb_path', None) != target_path:
                try:
                    if os.path.exists(target_path):
                        self._excel_wb = load_workbook(target_path)
                    else:
                        self._excel_wb = Workbook()
                except Exception:
                    self._excel_wb = Workbook()
                self._excel_ws_cache = {}
                self._excel_wb_path = target_path
            return self._excel_wb
        except Exception:
            return None

    def _excel_get_sheet_locked(self, sheet_name: str):
        """Get or create a worksheet; cached by name. Caller must hold _excel_lock."""
        try:
            if sheet_name in self._excel_ws_cache:
                return self._excel_ws_cache[sheet_name]
            wb = getattr(self, '_excel_wb', None)
            if wb is None:
                return None
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                if getattr(wb, 'active', None) is not None and len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1 and not wb.active['A1'].value:
                    ws = wb.active
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(sheet_name)
            self._excel_ws_cache[sheet_name] = ws
            return ws
        except Exception:
            return None

    def _excel_save_locked(self, force: bool = False):
        """Save cached workbook to working file, and snapshot atomically to final.
        Caller holds _excel_lock.
        Policy: write working file frequently; snapshot final at least every 2.5 minutes
        to keep the final XLSX current and mitigate oversized working files.
        """
        try:
            import shutil
            wb = getattr(self, '_excel_wb', None)
            wpath = getattr(self, '_excel_wb_path', None)
            if not wb or not wpath:
                return
            now = time.time()
            last = getattr(self, '_excel_last_save', 0.0)
            save_interval = 10.0
            if force or (now - last) >= save_interval:
                self._excel_saving = True
                wb.save(wpath)
                self._excel_last_save = now
            fpath = getattr(self, '_excel_final_path', None)
            if fpath:
                last_snap = getattr(self, '_excel_last_snapshot', 0.0)
                snapshot_interval = 150.0
                if force or ((now - last_snap) >= snapshot_interval):
                    tmp_final = fpath + '.tmp'
                    try:
                        shutil.copyfile(wpath, tmp_final)
                        os.replace(tmp_final, fpath)
                        self._excel_last_snapshot = now
                    except Exception:
                        try:
                            if os.path.exists(tmp_final):
                                os.remove(tmp_final)
                        except Exception:
                            pass
            self._excel_saving = False
        except Exception:
            try:
                self._excel_saving = False
            except Exception:
                pass

    # ---- pending-writes counter ----

    def _get_total_pending_writes(self) -> int:
        """Return approximate total pending items across known writer queues."""
        total = 0
        try:
            if hasattr(self, 'global_recorder') and getattr(self.global_recorder, '_write_q', None):
                total += self.global_recorder._write_q.qsize()
        except Exception:
            pass
        try:
            if hasattr(self, '_spectrum_write_q') and getattr(self, '_spectrum_write_q', None):
                total += self._spectrum_write_q.qsize()
        except Exception:
            pass
        try:
            if hasattr(self, '_register_write_q') and getattr(self, '_register_write_q', None):
                total += self._register_write_q.qsize()
        except Exception:
            pass
        return total

    # ---- save-status timer callback ----

    def _update_save_status(self):
        """Update status bar hinting whether Excel is still being saved."""
        try:
            saving = bool(getattr(self, '_excel_saving', False))
            has_marker = False
            try:
                marker = getattr(self, '_excel_marker_path', '')
                has_marker = bool(marker) and os.path.exists(marker)
            except Exception:
                has_marker = False
            has_working = False
            try:
                wpath = getattr(self, '_excel_working_path', '')
                has_working = bool(wpath) and os.path.exists(wpath)
            except Exception:
                has_working = False
            pending = 0
            try:
                if hasattr(self, 'global_recorder') and getattr(self.global_recorder, '_write_q', None):
                    pending += self.global_recorder._write_q.qsize()
                if hasattr(self, '_spectrum_write_q') and getattr(self, '_spectrum_write_q', None):
                    pending += self._spectrum_write_q.qsize()
                if hasattr(self, '_register_write_q') and getattr(self, '_register_write_q', None):
                    pending += self._register_write_q.qsize()
            except Exception:
                pass
            if saving or has_marker or has_working or pending > 0:
                self.statusBar().showMessage('Saving Excel data… please do not open the file yet', 2000)
            else:
                self.statusBar().showMessage('Excel file ready', 1500)
        except Exception:
            pass

    # ---- artifact cleanup ----

    def _cleanup_excel_artifacts(self, finalize: bool, timeout_s: float = 3.0) -> bool:
        """Try to remove .saving marker and .working.xlsx file with small retries."""
        end = time.time() + max(0.5, float(timeout_s))
        ok = False
        while time.time() < end:
            ok = True
            try:
                marker = getattr(self, '_excel_marker_path', '')
                if marker and os.path.exists(marker):
                    try:
                        os.remove(marker)
                    except Exception:
                        ok = False
            except Exception:
                pass
            try:
                wpath = getattr(self, '_excel_working_path', '')
                if wpath and os.path.exists(wpath):
                    try:
                        os.remove(wpath)
                    except Exception:
                        ok = False
            except Exception:
                pass
            if ok:
                break
            time.sleep(0.2)
        if not finalize:
            try:
                self._excel_wb = None
                self._excel_ws_cache = {}
                self._excel_wb_path = None
            except Exception:
                pass
        return ok

    # ---- bench info sheet ----

    def _write_bench_info(self, excel_path: str):
        """Create or update a 'Bench Info' sheet with a summary of the bench equipment and settings."""
        from openpyxl import Workbook, load_workbook
        rows = []
        ts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        def parse_serial_from_idn(idn: str) -> str:
            try:
                parts = [p.strip() for p in idn.replace(";", ",").split(',') if p.strip()]
                if len(parts) >= 3:
                    return parts[2]
                for tok in parts:
                    if tok.upper().startswith('SN'):
                        return tok.split(':')[-1].strip()
            except Exception:
                pass
            return ''

        for i in range(self.tabs.count()):
            panel = self.tabs.widget(i)
            tab_name = self.tabs.tabText(i)
            if isinstance(panel, KeithleyPanel):
                inst = getattr(panel, 'inst', None)
                connected = bool(inst)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if inst:
                        idn = inst.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                settings = {
                    'ch1': {'V': getattr(panel.vol_edits.get(1), 'text', lambda: '0')()},
                    'ch2': {'V': getattr(panel.vol_edits.get(2), 'text', lambda: '0')()},
                    'ch3': {'V': getattr(panel.vol_edits.get(3), 'text', lambda: '0')()},
                }
                try:
                    settings['ch1']['I'] = panel.iam_edits[1].text()
                    settings['ch2']['I'] = panel.iam_edits[2].text()
                    settings['ch3']['I'] = panel.iam_edits[3].text()
                except Exception:
                    pass
                state_parts = []
                for ch in (1, 2, 3):
                    ch_on = None
                    if connected:
                        try:
                            ch_on = bool(inst.get_output_state(ch))
                        except Exception:
                            ch_on = None
                    state_parts.append(f'ch{ch}:{"on" if ch_on else ("off" if ch_on is not None else "?")}')
                state = ', '.join(state_parts)
                rows.append([ts, 'Keithley 2230', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            elif isinstance(panel, KeysightE36233APanel):
                supply = getattr(panel, 'supply', None)
                connected = bool(supply)
                resource = ''
                try:
                    resource = panel.name_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if supply and getattr(supply, 'inst', None):
                        idn = supply.inst.query("*IDN?")
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                settings = {'ch1': {}, 'ch2': {}}
                state_parts = []
                for ch in (1, 2):
                    try:
                        settings[f'ch{ch}']['V'] = str(supply.get_voltage_setpoint(ch)) if connected else ''
                    except Exception:
                        pass
                    try:
                        settings[f'ch{ch}']['I'] = str(supply.get_current_setpoint(ch)) if connected else ''
                    except Exception:
                        pass
                    on_val = None
                    try:
                        on_val = bool(supply.get_output_state(ch)) if connected else None
                    except Exception:
                        on_val = None
                    state_parts.append(f'ch{ch}:{"on" if on_val else ("off" if on_val is not None else "?")}')
                state = ', '.join(state_parts)
                rows.append([ts, 'Keysight E36233A', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            elif isinstance(panel, KeysightELPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                settings = {
                    'ch1': {
                        'mode': getattr(panel.mode_combo_ch1, 'currentText', lambda: '')(),
                        'value': getattr(panel.mode_value_ch1, 'text', lambda: '')(),
                    },
                    'ch2': {
                        'mode': getattr(panel.mode_combo_ch2, 'currentText', lambda: '')(),
                        'value': getattr(panel.mode_value_ch2, 'text', lambda: '')(),
                    },
                }
                state = f"ch1:{'on' if panel.input_toggle_ch1.isChecked() else 'off'}, ch2:{'on' if panel.input_toggle_ch2.isChecked() else 'off'}"
                rows.append([ts, 'Keysight EL34243A', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            elif isinstance(panel, HittiteSigGenPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.freq_unit_combo, 'currentText', lambda: '')()
                settings = {
                    'frequency': getattr(panel.freq_edit, 'text', lambda: '')(),
                    'freq_unit': unit,
                    'power_dB': getattr(panel.pow_edit, 'text', lambda: '')(),
                }
                state = 'on' if getattr(panel.output_btn, 'isChecked', lambda: False)() else 'off'
                rows.append([ts, 'Hittite Sig Gen', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            elif isinstance(panel, RhodeSchwarzSMAPanel):
                dev = getattr(panel, 'dev', None)
                connected = bool(dev)
                resource = ''
                try:
                    resource = panel.resource_edit.text().strip()
                except Exception:
                    pass
                serial = ''
                try:
                    if dev:
                        idn = dev.get_identification()
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.freq_unit_combo, 'currentText', lambda: '')()
                settings = {
                    'frequency': getattr(panel.freq_edit, 'text', lambda: '')(),
                    'freq_unit': unit,
                    'power_dBm': getattr(panel.pow_edit, 'text', lambda: '')(),
                }
                state = 'on' if getattr(panel.output_btn, 'isChecked', lambda: False)() else 'off'
                rows.append([ts, 'RhodeSchwarz SMA', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])
            elif isinstance(panel, FieldFoxSAPanel):
                sa = getattr(panel, 'sa', None)
                connected = bool(getattr(sa, 'inst', None))
                resource = ''
                try:
                    resource = getattr(sa, 'visa_address', '')
                except Exception:
                    pass
                serial = ''
                try:
                    if connected:
                        try:
                            idn = sa._safe_query("*IDN?") if hasattr(sa, '_safe_query') else sa.inst.query("*IDN?")
                        except Exception:
                            idn = ''
                        serial = parse_serial_from_idn(idn)
                except Exception:
                    serial = ''
                unit = getattr(panel.unit_combo, 'currentText', lambda: '')()
                settings = {
                    'center': getattr(panel.center_edit, 'text', lambda: '')(),
                    'span': getattr(panel.span_edit, 'text', lambda: '')(),
                    'start': getattr(panel.start_edit, 'text', lambda: '')(),
                    'stop': getattr(panel.stop_edit, 'text', lambda: '')(),
                    'unit': unit,
                }
                running = bool(getattr(panel, '_capture_thread', None))
                state = 'capture:running' if running else 'capture:stopped'
                rows.append([ts, 'Keysight FieldFox', tab_name, resource, serial, 'Yes' if connected else 'No', state, json.dumps(settings)])

        try:
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    wb = self._excel_open_locked(excel_path)
                    if wb is None:
                        return
                    if 'Bench Info' in wb.sheetnames:
                        ws = wb['Bench Info']
                        try:
                            for _ in range(ws.max_row, 0, -1):
                                ws.delete_rows(1)
                        except Exception:
                            pass
                    else:
                        ws = wb.create_sheet('Bench Info')
                    header = ['timestamp', 'instrument', 'name', 'resource', 'serial', 'connected', 'state', 'settings']
                    ws.append(header)
                    if rows:
                        for r in rows:
                            ws.append(r)
                    else:
                        ws.append([ts, '(none)', '(no instruments)', '', '', '', '', '{}'])
                        try:
                            QtCore.QTimer.singleShot(1000, lambda: self._write_bench_info(excel_path))
                        except Exception:
                            pass
                    try:
                        if 'Sheet' in wb.sheetnames and wb['Sheet'].max_row == 1 and wb['Sheet'].max_column == 1 and not wb['Sheet']['A1'].value:
                            del wb['Sheet']
                    except Exception:
                        pass
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception as e:
            try:
                self._log(f'Bench Info write error: {e}')
            except Exception:
                pass
