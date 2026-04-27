"""
Recording management mixin for MainWindow.

Provides record/prime/stop logic, supply recorder factory,
spectrum capture, and register recording.
"""
import os
import json
import datetime
import time
import threading
import queue
import re

from PyQt5 import QtWidgets, QtCore

from Support_Scrips.Front_Panels.fieldfox_sa_panel import FieldFoxSAPanel


class RecordingMixin:
    """Mixin that adds recording functionality to MainWindow."""

    def _resolve_register_read_array(self) -> list:
        """Return the active register_read_array, falling back to the loaded config file."""
        try:
            arr = getattr(self, 'register_read_array', None)
            if arr:
                return list(arr)
        except Exception:
            pass
        try:
            cfg_name = self.load_combo.currentText() if hasattr(self, 'load_combo') else ''
            if cfg_name:
                cfg_path = os.path.join(self.configs_dir, cfg_name)
                if os.path.exists(cfg_path):
                    with open(cfg_path, 'r') as f:
                        data = json.load(f)
                        return list(data.get('register_read_array', []) or [])
        except Exception:
            pass
        return []

    def open_register_monitor(self):
        """Open (or raise) the live register monitor dialog."""
        from register_monitor import RegisterMonitorDialog
        regs = self._resolve_register_read_array()
        if not regs:
            QtWidgets.QMessageBox.warning(
                self, 'No registers configured',
                'No register_read_array found in the current config.\n'
                'Load a config that defines it before opening the monitor.')
            return
        existing = getattr(self, '_register_monitor_dialog', None)
        if existing is not None:
            try:
                if existing.isVisible():
                    existing.raise_()
                    existing.activateWindow()
                    return
            except Exception:
                pass
        dlg = RegisterMonitorDialog(self, regs)
        self._register_monitor_dialog = dlg
        dlg.show()

    def on_record_clicked(self):
        """Unified record button: start/stop recording for selected metrics."""
        if self.record_btn.isChecked():
            if getattr(self, '_record_primed', False):
                try:
                    if getattr(self, '_record_start_event', None):
                        self._record_start_event.set()
                    self._record_sync_start = time.time()
                except Exception:
                    pass
                self.record_btn.setText('Stop Recording')
                self.statusBar().showMessage('Recording started', 3000)
                return
            excel_path = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Reads', '', 'Excel Files (*.xlsx)')[0]
            if excel_path and not excel_path.lower().endswith('.xlsx'):
                excel_path += '.xlsx'
            if not excel_path:
                return
            try:
                self._last_excel_path = excel_path
            except Exception:
                pass
            try:
                if not hasattr(self, '_excel_lock') or self._excel_lock is None:
                    self._excel_lock = threading.Lock()
            except Exception:
                self._excel_lock = None
            try:
                self._excel_wb = None
                self._excel_ws_cache = {}
                self._excel_final_path = excel_path
                self._excel_working_path = excel_path + '.working.xlsx'
                self._excel_marker_path = excel_path + '.saving'
                self._excel_wb_path = self._excel_working_path
                self._excel_last_save = 0.0
                self._excel_last_snapshot = 0.0
                self._excel_saving = False
            except Exception:
                pass
            try:
                with open(self._excel_marker_path, 'w') as f:
                    f.write('saving')
            except Exception:
                pass
            try:
                self._record_sync_start = time.time() + 1.5
            except Exception:
                self._record_sync_start = None
            try:
                self._write_bench_info(excel_path)
            except Exception as e:
                try:
                    self._log(f'Bench Info write failed: {e}')
                except Exception:
                    pass
            if self.supply_record_toggle.isChecked():
                panels = self.get_all_supply_panels()
                if not panels:
                    self.statusBar().showMessage('No supply panels to record', 4000)
                self.global_recorder = self._create_supply_recorder(panels, excel_path)
                self.global_recorder.start()
                self.statusBar().showMessage('Started recording supplies', 4000)
            if self.spectrum_record_toggle.isChecked():
                fieldfox_panel = None
                for i in range(self.tabs.count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel):
                        fieldfox_panel = w
                        break
                if fieldfox_panel is None:
                    self.statusBar().showMessage('No FieldFox panel found', 4000)
                    fieldfox_panel = None
                if fieldfox_panel is not None:
                    def update_spectrum_read_speed(sps: float):
                        self.spectrum_read_speed_label.setText(f'Spectrum Sample Rate: {sps:.2f} samples/sec')
                    self._start_spectrum_capture(fieldfox_panel, excel_path, update_spectrum_read_speed)
                    self._spectrum_recording = True
                    self._spectrum_thread_running = True
                    def running_flag():
                        return self._spectrum_recording and self._spectrum_thread_running
                    self.statusBar().showMessage('Started recording spectrum', 4000)
            if self.register_record_toggle.isChecked():
                registers = []
                try:
                    if hasattr(self, 'register_read_array') and self.register_read_array:
                        registers = self.register_read_array
                    else:
                        cfg_name = self.load_combo.currentText() if hasattr(self, 'load_combo') else ''
                        if cfg_name:
                            cfg_path = os.path.join(self.configs_dir, cfg_name)
                            if os.path.exists(cfg_path):
                                with open(cfg_path, 'r') as f:
                                    data = json.load(f)
                                    registers = data.get('register_read_array', [])
                except Exception:
                    registers = []
                if not registers:
                    QtWidgets.QMessageBox.warning(self, 'No registers configured', 'No register_read_array found in current config. Load a config that defines it.')
                else:
                    def update_register_read_speed(sps: float):
                        try:
                            self.register_read_speed_label.setText(f'Register Sample Rate: {sps:.2f} samples/sec')
                        except Exception:
                            pass
                    self._start_register_recording(excel_path, registers, update_register_read_speed)
                    self.statusBar().showMessage('Started recording registers', 4000)
            self.record_btn.setText('Stop Recording')
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
        else:
            self._stop_all_recordings()
            self.record_btn.setText('Record')
            self.statusBar().showMessage('Stopped recording', 4000)
            try:
                self._record_primed = False
                self.record_btn.setEnabled(False)
                for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                    cb.setEnabled(True)
                if hasattr(self, 'prime_btn'):
                    self.prime_btn.setText('Prime Recorders')
                    self.prime_btn.setStyleSheet('')
            except Exception:
                pass

    def _create_supply_recorder(self, panels, excel_path):
        """Factory for a PatchedSupplyRecorder with background writer and 25Hz throttle."""
        from supply_recorder import SupplyRecorder
        self_parent = self

        def update_supply_read_speed(sps: float):
            try:
                self_parent.supply_read_speed_label.setText(f'Supply Sample Rate: {sps:.2f} samples/sec')
            except Exception:
                pass

        class PatchedSupplyRecorder(SupplyRecorder):
            def __init__(self, panels, excel_path, sheet_name='supply reads'):
                super().__init__(panels, excel_path, sheet_name)
                self._write_q = queue.Queue(maxsize=100000)
                self._writer_stop = None
                self._writer_thread = None
                self._header = self._build_header(panels)

            def _build_header(self, panels):
                ts = ['timestamp', 'elapsed_s']
                v_headers = []
                i_headers = []
                for p in panels:
                    try:
                        tab_idx = self_parent.tabs.indexOf(p)
                        alias_name = self_parent.tabs.tabText(tab_idx) if tab_idx >= 0 else getattr(p, 'resource', 'Supply')
                    except Exception:
                        alias_name = getattr(p, 'resource', 'Supply')
                    if p.__class__.__name__.startswith('Keithley'):
                        ch_count = 3
                        try:
                            ch_names = [p.ch_name_edits[i].text() if hasattr(p, 'ch_name_edits') and p.ch_name_edits.get(i) else f'CH{i}' for i in (1, 2, 3)]
                        except Exception:
                            ch_names = [f'CH{i}' for i in (1, 2, 3)]
                    else:
                        ch_count = 2
                        ch_names = [f'CH{i}' for i in (1, 2)]
                    for i in range(1, ch_count + 1):
                        nm = ch_names[i - 1] if i - 1 < len(ch_names) else f'CH{i}'
                        v_headers.append(f'{alias_name}_{nm}_V')
                for p in panels:
                    try:
                        tab_idx = self_parent.tabs.indexOf(p)
                        alias_name = self_parent.tabs.tabText(tab_idx) if tab_idx >= 0 else getattr(p, 'resource', 'Supply')
                    except Exception:
                        alias_name = getattr(p, 'resource', 'Supply')
                    if p.__class__.__name__.startswith('Keithley'):
                        ch_count = 3
                        try:
                            ch_names = [p.ch_name_edits[i].text() if hasattr(p, 'ch_name_edits') and p.ch_name_edits.get(i) else f'CH{i}' for i in (1, 2, 3)]
                        except Exception:
                            ch_names = [f'CH{i}' for i in (1, 2, 3)]
                    else:
                        ch_count = 2
                        ch_names = [f'CH{i}' for i in (1, 2)]
                    for i in range(1, ch_count + 1):
                        nm = ch_names[i - 1] if i - 1 < len(ch_names) else f'CH{i}'
                        i_headers.append(f'{alias_name}_{nm}_I')
                return ts + v_headers + i_headers

            def start(self):
                self._writer_stop = threading.Event()

                def writer_loop():
                    lock = getattr(self_parent, '_excel_lock', None)
                    acquired = lock.acquire(timeout=5) if lock else True
                    if acquired:
                        try:
                            self_parent._excel_open_locked(self.excel_path)
                            ws = self_parent._excel_get_sheet_locked(self.sheet_name)
                            for ci, val in enumerate(self._header, start=1):
                                ws.cell(row=1, column=ci, value=val)
                            self_parent._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                    pending = []
                    last_save = time.time()
                    flush_rows = 1200
                    while True:
                        try:
                            item = self._write_q.get(timeout=0.3)
                            pending.append(item)
                            self._write_q.task_done()
                        except Exception:
                            pass
                        now = time.time()
                        should_save = pending and (len(pending) >= flush_rows or (now - last_save) >= 8.0)
                        should_stop = self._writer_stop.is_set() and self._write_q.empty()
                        if should_save or should_stop:
                            acquired = lock.acquire(timeout=10) if lock else True
                            if acquired:
                                try:
                                    self_parent._excel_open_locked(self.excel_path)
                                    ws = self_parent._excel_get_sheet_locked(self.sheet_name)
                                    for row in pending:
                                        ws.append(row)
                                    pending.clear()
                                    last_save = now
                                    self_parent._excel_save_locked()
                                finally:
                                    if lock and acquired:
                                        try:
                                            lock.release()
                                        except Exception:
                                            pass
                        if should_stop and not pending:
                            break

                self._writer_thread = threading.Thread(target=writer_loop, daemon=True)
                self._writer_thread.start()
                return super().start()

            def stop(self):
                try:
                    self._stop_event.set()
                    if self._thread:
                        self._thread.join()
                except Exception:
                    pass
                try:
                    if self._writer_stop:
                        self._writer_stop.set()
                    if self._writer_thread:
                        self._writer_thread.join(timeout=10)
                except Exception:
                    pass

            def _run(self):
                sample_count = 0
                try:
                    evt = getattr(self_parent, '_record_start_event', None)
                    if evt is not None:
                        while not evt.is_set() and not self._stop_event.is_set():
                            time.sleep(0.005)
                except Exception:
                    pass
                if getattr(self_parent, '_record_sync_start', None):
                    while time.time() < self_parent._record_sync_start and not self._stop_event.is_set():
                        time.sleep(0.005)
                start_time = time.time()
                last_report_time = start_time
                target_interval = getattr(self_parent, '_supply_target_interval', 0.04)
                next_tick = time.perf_counter()
                while not self._stop_event.is_set():
                    nowp = time.perf_counter()
                    if nowp < next_tick:
                        time.sleep(min(next_tick - nowp, 0.01))
                        continue
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    elapsed_s = round(time.time() - start_time, 3)
                    all_voltages = []
                    all_currents = []
                    for func in self.get_readings_funcs:
                        voltages, currents = func()
                        all_voltages.extend(voltages)
                        all_currents.extend(currents)
                    row = [now, elapsed_s] + all_voltages + all_currents
                    try:
                        self._write_q.put_nowait(row)
                    except Exception:
                        pass
                    sample_count += 1
                    now_ts = time.time()
                    if now_ts - last_report_time >= 2.0:
                        elapsed = max(now_ts - start_time, 1e-6)
                        sps = sample_count / elapsed
                        update_supply_read_speed(sps)
                        last_report_time = now_ts
                    next_tick += target_interval
                    if next_tick < nowp - target_interval:
                        next_tick = nowp + target_interval

        return PatchedSupplyRecorder(panels, excel_path, sheet_name='supply reads')

    def on_prime_clicked(self):
        """Prepare selected recorders: choose file, init Excel, and spin threads waiting for start."""
        try:
            if getattr(self, '_record_primed', False):
                try:
                    if hasattr(self, 'record_btn') and self.record_btn.isChecked():
                        self.statusBar().showMessage('Recording is running; stop it before unpriming', 4000)
                        return
                except Exception:
                    pass
                self._unprime_recorders()
                return
            excel_path = QtWidgets.QFileDialog.getSaveFileName(self, 'Save Reads', '', 'Excel Files (*.xlsx)')[0]
            if excel_path and not excel_path.lower().endswith('.xlsx'):
                excel_path += '.xlsx'
            if not excel_path:
                return
            if not hasattr(self, '_excel_lock') or self._excel_lock is None:
                self._excel_lock = threading.Lock()
            self._excel_wb = None
            self._excel_ws_cache = {}
            self._excel_final_path = excel_path
            self._excel_working_path = excel_path + '.working.xlsx'
            self._excel_marker_path = excel_path + '.saving'
            self._excel_wb_path = self._excel_working_path
            self._excel_last_save = 0.0
            self._excel_last_snapshot = 0.0
            self._excel_saving = False
            self._last_excel_path = excel_path
            try:
                with open(self._excel_marker_path, 'w') as f:
                    f.write('saving')
            except Exception:
                pass
            self._record_start_event = threading.Event()
            self._record_sync_start = time.time() + 1.0
            try:
                self._write_bench_info(excel_path)
            except Exception:
                pass
            if self.supply_record_toggle.isChecked():
                panels = self.get_all_supply_panels()
                if panels:
                    self.global_recorder = self._create_supply_recorder(panels, excel_path)
                    self.global_recorder.start()
            if self.spectrum_record_toggle.isChecked():
                fieldfox_panel = None
                for i in range(self.tabs.count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel):
                        fieldfox_panel = w
                        break
                if fieldfox_panel is not None:
                    def update_spectrum_read_speed(sps: float):
                        self.spectrum_read_speed_label.setText(f'Spectrum Sample Rate: {sps:.2f} samples/sec')
                    self._start_spectrum_capture(fieldfox_panel, excel_path, update_spectrum_read_speed)
            if self.register_record_toggle.isChecked():
                registers = []
                try:
                    if hasattr(self, 'register_read_array') and self.register_read_array:
                        registers = self.register_read_array
                    else:
                        cfg_name = self.load_combo.currentText() if hasattr(self, 'load_combo') else ''
                        if cfg_name:
                            cfg_path = os.path.join(self.configs_dir, cfg_name)
                            if os.path.exists(cfg_path):
                                with open(cfg_path, 'r') as f:
                                    data = json.load(f)
                                    registers = data.get('register_read_array', [])
                except Exception:
                    registers = []
                if registers:
                    def update_register_read_speed(sps: float):
                        try:
                            self.register_read_speed_label.setText(f'Register Sample Rate: {sps:.2f} samples/sec')
                        except Exception:
                            pass
                    self._start_register_recording(excel_path, registers, update_register_read_speed)
            self._record_primed = True
            self.record_btn.setEnabled(True)
            for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                cb.setEnabled(False)
            if hasattr(self, 'prime_btn'):
                self.prime_btn.setText('Recorders Primed (Click to Unprime)')
                self.prime_btn.setStyleSheet('background-color: #4CAF50; color: white;')
                try:
                    self.prime_btn.setToolTip('Click to unprime and tear down prepared recorders')
                except Exception:
                    pass
            self.statusBar().showMessage('Recorders primed. Press Record to start.', 4000)
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Prime failed', str(e))

    def _unprime_recorders(self):
        """Tear down any primed recorders and reset UI back to pre-primed state."""
        try:
            if hasattr(self, 'record_btn') and self.record_btn.isChecked():
                return
        except Exception:
            pass
        try:
            self._stop_all_recordings(finalize=False)
        except Exception:
            pass
        try:
            self._record_primed = False
            try:
                self._record_start_event = None
                self._record_sync_start = None
            except Exception:
                pass
            if hasattr(self, 'record_btn'):
                try:
                    self.record_btn.setChecked(False)
                except Exception:
                    pass
                self.record_btn.setText('Record')
                self.record_btn.setEnabled(False)
            for cb in (self.supply_record_toggle, self.spectrum_record_toggle, self.register_record_toggle):
                try:
                    cb.setEnabled(True)
                except Exception:
                    pass
            if hasattr(self, 'prime_btn'):
                try:
                    self.prime_btn.setText('Prime Recorders')
                    self.prime_btn.setStyleSheet('')
                    self.prime_btn.setToolTip('Prepare selected recorders and Excel file; enables Start')
                except Exception:
                    pass
            self.statusBar().showMessage('Recorders unprimed', 3000)
        except Exception:
            pass

    def _stop_all_recordings(self, finalize: bool = True):
        """Stop supply, spectrum, and register recordings and reset labels."""
        # Supply
        try:
            if hasattr(self, 'global_recorder') and self.global_recorder:
                try:
                    self.global_recorder.stop()
                except Exception:
                    pass
                if finalize:
                    try:
                        lock = getattr(self, '_excel_lock', None)
                        if lock and lock.acquire(timeout=5):
                            try:
                                if getattr(self.global_recorder, 'buffer', None):
                                    self.global_recorder._flush_buffer()
                            finally:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                    except Exception:
                        pass
                self.global_recorder = None
            if hasattr(self, 'supply_read_speed_label'):
                self.supply_read_speed_label.setText('Supply Sample Rate: --')
        except Exception:
            pass
        # Spectrum
        try:
            if hasattr(self, '_spectrum_thread') and self._spectrum_thread:
                self._spectrum_recording = False
                self._spectrum_thread_running = False
                try:
                    self._spectrum_thread.join(timeout=2)
                except Exception:
                    pass
                self._spectrum_thread = None
            try:
                for i in range(getattr(self, 'tabs', QtWidgets.QTabWidget()).count()):
                    w = self.tabs.widget(i)
                    if isinstance(w, FieldFoxSAPanel) and hasattr(w, 'toggle_streaming'):
                        if not w.streaming_enabled:
                            w.stream_btn.setChecked(True)
                            w.toggle_streaming(True)
            except Exception:
                pass
        except Exception:
            pass
        # Registers
        try:
            if hasattr(self, '_register_thread') and getattr(self, '_register_thread', None):
                try:
                    self._register_recording = False
                except Exception:
                    pass
                try:
                    self._register_thread_running = False
                except Exception:
                    pass
                try:
                    self._register_thread.join(timeout=2)
                except Exception:
                    pass
                self._register_thread = None
            if hasattr(self, 'register_read_speed_label'):
                self.register_read_speed_label.setText('Register Sample Rate: --')
        except Exception:
            pass
        # Finalize or discard workbook artifacts
        try:
            if finalize:
                try:
                    t0 = time.time()
                    while (time.time() - t0) < 2.0:
                        try:
                            pending = self._get_total_pending_writes()
                        except Exception:
                            pending = 0
                        if pending <= 0:
                            break
                        time.sleep(0.1)
                except Exception:
                    pass
            if finalize:
                lock = getattr(self, '_excel_lock', None)
                if lock and lock.acquire(timeout=5):
                    try:
                        # Highlight register transitions that follow a prolonged
                        # constant run before doing the final save.
                        try:
                            self._apply_register_change_highlights_locked()
                        except Exception as e:
                            try:
                                self._log(f'Register change highlighting failed: {e}')
                            except Exception:
                                pass
                        self._excel_save_locked(force=True)
                    finally:
                        try:
                            lock.release()
                        except Exception:
                            pass
            try:
                if not hasattr(self, '_save_status_timer'):
                    self._save_status_timer = QtCore.QTimer(self)
                    self._save_status_timer.setInterval(1000)
                    self._save_status_timer.timeout.connect(self._update_save_status)
                self._save_status_timer.start()
            except Exception:
                pass
            ok = self._cleanup_excel_artifacts(finalize=finalize, timeout_s=3.0)
            try:
                if hasattr(self, '_save_status_timer'):
                    self._save_status_timer.stop()
            except Exception:
                pass
            if not finalize:
                try:
                    self._excel_final_path = None
                    self._excel_working_path = None
                    self._excel_marker_path = None
                except Exception:
                    pass
        except Exception:
            pass

    def _apply_register_change_highlights_locked(self):
        """Highlight transitions in the 'register reads' sheet where a register
        value changes after holding the same value for at least
        ``register_change_min_run`` consecutive samples.

        Must be called with ``self._excel_lock`` already held. Uses the
        currently open workbook (``_excel_*_locked`` helpers).
        """
        from openpyxl.styles import PatternFill
        from openpyxl.comments import Comment

        sheet_name = 'register reads'
        try:
            wb = getattr(self, '_excel_wb', None)
            if wb is None or sheet_name not in wb.sheetnames:
                return
        except Exception:
            return

        try:
            min_run = int(getattr(self, 'register_change_min_run', 10))
        except Exception:
            min_run = 10
        if min_run < 2:
            min_run = 2

        try:
            ws = self._excel_get_sheet_locked(sheet_name)
        except Exception:
            return

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        # Header row 1, EXPECTED row 2, data starts row 3
        data_start = 3
        if max_row < data_start + 1 or max_col < 3:
            return

        # Use a yellow fill for "change after prolonged hold". The existing
        # mismatch-vs-EXPECTED highlighting (red) is applied via conditional
        # formatting and remains in place; this fill takes precedence on the
        # specific transition cells where it is applied.
        change_fill = PatternFill(start_color='FFEB3B', end_color='FFEB3B', fill_type='solid')

        highlight_count = 0
        for col in range(3, max_col + 1):
            prev_val = None
            run_len = 0
            for row in range(data_start, max_row + 1):
                cell = ws.cell(row=row, column=col)
                v = cell.value
                if v is None:
                    prev_val = None
                    run_len = 0
                    continue
                if prev_val is None:
                    prev_val = v
                    run_len = 1
                    continue
                if v == prev_val:
                    run_len += 1
                else:
                    if run_len >= min_run:
                        try:
                            cell.fill = change_fill
                            cell.comment = Comment(
                                f'Changed after holding previous value for '
                                f'{run_len} consecutive samples.',
                                'RegisterMonitor',
                            )
                            highlight_count += 1
                        except Exception:
                            pass
                    prev_val = v
                    run_len = 1
        try:
            self._log(
                f'Register change highlighting: marked {highlight_count} '
                f'transition(s) (min run = {min_run} samples).')
        except Exception:
            pass

    def _start_register_recording(self, excel_path: str, registers: list, rate_cb):
        """Start register recording in background with a non-blocking Excel writer."""
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import CellIsRule

        reg_list: list[int] = []
        for r in registers:
            try:
                if isinstance(r, str):
                    reg_list.append(int(r, 0))
                else:
                    reg_list.append(int(r))
            except Exception:
                continue
        if not reg_list:
            QtWidgets.QMessageBox.warning(self, 'No registers', 'No registers specified to record.')
            return

        # Publish address list + initialise live-monitor data structures so
        # the RegisterMonitorDialog can read them.
        try:
            self._register_addr_list = list(reg_list)
            self._register_latest_values = {}
            self._register_baseline = {}
        except Exception:
            pass

        sheet_name = 'register reads'
        start_time = time.time()
        last_report_time = start_time
        sample_count = 0

        # Row 2 is reserved for the EXPECTED baseline values; data starts at row 3.
        baseline_row = 2
        data_start_row = 3
        baseline_values: list = []  # filled on first successful read
        baseline_written = threading.Event()
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        try:
            header = ['timestamp', 'elapsed_s'] + [f'{addr:#x}' for addr in reg_list]
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    self._excel_open_locked(excel_path)
                    ws = self._excel_get_sheet_locked(sheet_name)
                    for ci, val in enumerate(header, start=1):
                        ws.cell(row=1, column=ci, value=val)
                    # Pre-label the baseline row
                    ws.cell(row=baseline_row, column=1, value='EXPECTED')
                    ws.cell(row=baseline_row, column=2, value=0)
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception:
            pass

        write_q: "queue.Queue[list]" = queue.Queue(maxsize=50000)
        try:
            self._register_write_q = write_q
        except Exception:
            pass
        writer_stop = threading.Event()
        flush_interval = 150
        header = ['timestamp', 'elapsed_s'] + [f'{addr:#x}' for addr in reg_list]

        def _write_baseline_and_formatting(ws, vals):
            """Write the EXPECTED row and apply conditional formatting once."""
            from openpyxl.utils import get_column_letter
            ws.cell(row=baseline_row, column=1, value='EXPECTED')
            ws.cell(row=baseline_row, column=2, value=0)
            for ci, v in enumerate(vals, start=3):
                ws.cell(row=baseline_row, column=ci, value=v)
            # Apply conditional formatting: highlight cells that differ from EXPECTED
            for ci in range(3, 3 + len(vals)):
                col_letter = get_column_letter(ci)
                ref_cell = f'${col_letter}${baseline_row}'
                cell_range = f'{col_letter}{data_start_row}:{col_letter}1048576'
                ws.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator='notEqual',
                        formula=[ref_cell],
                        fill=red_fill,
                    ),
                )

        def writer_loop():
            pending_rows: list[list] = []
            last_save = time.time()
            while True:
                try:
                    item = write_q.get(timeout=0.3)
                    pending_rows.append(item)
                    write_q.task_done()
                except queue.Empty:
                    pass
                now = time.time()
                should_save = pending_rows and (len(pending_rows) >= flush_interval or (now - last_save) >= 8.0)
                should_stop = writer_stop.is_set() and write_q.empty()
                if should_save or should_stop:
                    lock = getattr(self, '_excel_lock', None)
                    acquired = lock.acquire(timeout=10) if lock else True
                    if acquired:
                        try:
                            self._excel_open_locked(excel_path)
                            ws = self._excel_get_sheet_locked(sheet_name)
                            try:
                                for ci, val in enumerate(header, start=1):
                                    ws.cell(row=1, column=ci, value=val)
                            except Exception:
                                pass
                            # Write baseline + formatting on first flush
                            if baseline_written.is_set() and not getattr(ws, '_baseline_done', False):
                                try:
                                    _write_baseline_and_formatting(ws, baseline_values)
                                    ws._baseline_done = True
                                except Exception:
                                    pass
                            for row in pending_rows:
                                ws.append(row)
                            pending_rows.clear()
                            last_save = now
                            self._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                if should_stop and not pending_rows:
                    break

        writer_thread = threading.Thread(target=writer_loop, daemon=True)
        writer_thread.start()

        def running_flag():
            return getattr(self, '_register_recording', False) and getattr(self, '_register_thread_running', False)

        def worker():
            nonlocal sample_count, last_report_time, start_time
            try:
                evt = getattr(self, '_record_start_event', None)
                if evt is not None:
                    while not evt.is_set() and running_flag():
                        time.sleep(0.01)
                if getattr(self, '_record_sync_start', None):
                    while time.time() < self._record_sync_start and running_flag():
                        time.sleep(0.01)
            except Exception:
                pass
            recording_start_time = time.time()
            try:
                import clr  # type: ignore
                clr.AddReference('AnalogDevices.Csa.Remoting.Clients')
                clr.AddReference('AnalogDevices.Csa.Remoting.Contracts')
                from AnalogDevices.Csa.Remoting.Clients import ClientManager  # type: ignore
                manager = ClientManager.Create()
                client = manager.CreateRequestClient('localhost:2357')
            except Exception as e:
                try:
                    self._log(f'ACE connection failed: {e}')
                except Exception:
                    pass
                return
            def _parse_register_value(raw_val):
                """Convert a raw ACE register response to an integer."""
                if hasattr(raw_val, 'strip'):
                    raw_val = raw_val.strip('\r\n')
                if isinstance(raw_val, str):
                    sval = raw_val.strip()
                    if re.fullmatch(r'0[xX][0-9A-Fa-f]+', sval):
                        return int(sval, 16)
                    elif re.fullmatch(r'[0-9A-Fa-f]+', sval):
                        return int(sval, 16)
                    else:
                        return int(sval)
                return int(raw_val)

            while running_flag():
                try:
                    nowts = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    elapsed_s = round(time.time() - recording_start_time, 3)
                    values = []
                    for addr in reg_list:
                        try:
                            val = _parse_register_value(client.ReadRegister(str(int(addr))))
                        except Exception as err:
                            val = f'ERR:{err}'
                        values.append(val)
                    # Capture baseline from the first successful read
                    if not baseline_written.is_set():
                        if all(isinstance(v, int) for v in values):
                            baseline_values.extend(values)
                            baseline_written.set()
                            try:
                                self._register_baseline = {
                                    a: v for a, v in zip(reg_list, values)
                                }
                            except Exception:
                                pass
                    # Publish latest values for the live monitor (best effort)
                    try:
                        latest = getattr(self, '_register_latest_values', None)
                        if latest is None:
                            latest = {}
                            self._register_latest_values = latest
                        for a, v in zip(reg_list, values):
                            latest[a] = (v, nowts, elapsed_s)
                    except Exception:
                        pass
                    try:
                        write_q.put_nowait([nowts, elapsed_s] + values)
                    except queue.Full:
                        pass
                    sample_count += 1
                    now_time = time.time()
                    if now_time - last_report_time >= 2.0:
                        elapsed = max(now_time - start_time, 1e-6)
                        try:
                            rate_cb(sample_count / elapsed)
                        except Exception:
                            pass
                        last_report_time = now_time
                    time.sleep(0.05)
                except Exception as e:
                    try:
                        self._log(f'Register read loop error: {e}')
                    except Exception:
                        pass
                    time.sleep(0.1)
            try:
                writer_stop.set()
            except Exception:
                pass
            try:
                writer_thread.join(timeout=10)
            except Exception:
                pass
            try:
                self._register_thread_running = False
            except Exception:
                pass

        self._register_recording = True
        self._register_thread_running = True
        self._register_thread = threading.Thread(target=worker, daemon=True)
        self._register_thread.start()

    def _start_spectrum_capture(self, panel: FieldFoxSAPanel, excel_path: str, rate_cb):
        """Start spectrum capture in background with a non-blocking Excel writer."""
        from openpyxl import Workbook, load_workbook

        try:
            if not getattr(self, 'spectrum_record_toggle', None) or not self.spectrum_record_toggle.isChecked():
                return
        except Exception:
            return

        self._spectrum_recording = True
        self._spectrum_thread_running = True

        try:
            if getattr(panel.sa, 'inst', None) is None:
                panel.on_connect()
        except Exception:
            pass

        try:
            freq = panel.sa.get_freq_axis(panel.unit_combo.currentText())
        except Exception:
            freq = []

        try:
            lock = getattr(self, '_excel_lock', None)
            acquired = lock.acquire(timeout=5) if lock else True
            if acquired:
                try:
                    self._excel_open_locked(excel_path)
                    ws = self._excel_get_sheet_locked('capture data')
                    try:
                        n = len(freq)
                    except Exception:
                        n = 0
                    header = ['timestamp', 'elapsed_s'] + ([f"{f:.6f}" for f in freq] if n > 0 else [])
                    for ci, val in enumerate(header, start=1):
                        ws.cell(row=1, column=ci, value=val)
                    self._excel_save_locked()
                finally:
                    if lock and acquired:
                        try:
                            lock.release()
                        except Exception:
                            pass
        except Exception:
            pass

        try:
            if hasattr(panel, 'stop_capture_thread'):
                panel.stop_capture_thread()
        except Exception:
            pass
        try:
            if hasattr(panel.sa, 'sync'):
                panel.sa.sync()
        except Exception:
            pass

        try:
            n0 = len(freq)
        except Exception:
            n0 = 0
        desired_header = ['timestamp', 'elapsed_s'] + ([f"{f:.6f}" for f in freq] if n0 > 0 else [])
        write_q: "queue.Queue[list]" = queue.Queue(maxsize=20000)
        try:
            self._spectrum_write_q = write_q
        except Exception:
            pass
        writer_stop = threading.Event()
        flush_interval = 200

        def spec_writer_loop():
            pending_rows: list[list] = []
            last_save = time.time()
            while True:
                try:
                    item = write_q.get(timeout=0.3)
                    pending_rows.append(item)
                    write_q.task_done()
                except queue.Empty:
                    pass
                now = time.time()
                should_save = pending_rows and (len(pending_rows) >= flush_interval or (now - last_save) >= 6.0)
                should_stop = writer_stop.is_set() and write_q.empty()
                if should_save or should_stop:
                    lock = getattr(self, '_excel_lock', None)
                    acquired = lock.acquire(timeout=10) if lock else True
                    if acquired:
                        try:
                            self._excel_open_locked(excel_path)
                            ws = self._excel_get_sheet_locked('capture data')
                            try:
                                for ci, val in enumerate(desired_header, start=1):
                                    ws.cell(row=1, column=ci, value=val)
                            except Exception:
                                pass
                            for row in pending_rows:
                                ws.append(row)
                            pending_rows.clear()
                            last_save = now
                            self._excel_save_locked()
                        finally:
                            if lock and acquired:
                                try:
                                    lock.release()
                                except Exception:
                                    pass
                if should_stop and not pending_rows:
                    break

        spec_writer_thread = threading.Thread(target=spec_writer_loop, daemon=True)
        spec_writer_thread.start()

        start_time = time.time()
        last_report_time = start_time
        sample_count = 0

        def running_flag():
            return getattr(self, '_spectrum_recording', False) and getattr(self, '_spectrum_thread_running', False)

        def worker():
            nonlocal sample_count, last_report_time, freq, desired_header
            import pyvisa
            try:
                evt = getattr(self, '_record_start_event', None)
                if evt is not None:
                    while not evt.is_set() and running_flag():
                        time.sleep(0.01)
                if getattr(self, '_record_sync_start', None):
                    while time.time() < self._record_sync_start and running_flag():
                        time.sleep(0.01)
            except Exception:
                pass
            recording_start_time = time.time()
            while running_flag():
                try:
                    if getattr(panel.sa, 'inst', None) is None:
                        try:
                            panel.on_connect()
                        except Exception:
                            pass
                        time.sleep(0.3)
                        continue
                    amplitudes = panel.sa.capture_spectrum()
                    try:
                        n_freq = len(freq)
                    except Exception:
                        n_freq = 0
                    if not n_freq or (hasattr(amplitudes, '__len__') and len(amplitudes) != n_freq):
                        try:
                            freq = panel.sa.get_freq_axis(panel.unit_combo.currentText())
                            try:
                                n1 = len(freq)
                            except Exception:
                                n1 = 0
                            desired_header = ['timestamp', 'elapsed_s'] + ([f"{f:.6f}" for f in freq] if n1 > 0 else [])
                        except Exception:
                            pass
                    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')[:-3]
                    elapsed_s = round(time.time() - recording_start_time, 3)
                    row = [now, elapsed_s] + [float(a) for a in amplitudes]
                    try:
                        write_q.put_nowait(row)
                    except queue.Full:
                        pass
                    sample_count += 1
                except pyvisa.errors.InvalidSession:
                    break
                except Exception as e:
                    msg = ''
                    try:
                        msg = str(e.args[0]) if getattr(e, 'args', None) else str(e)
                    except Exception:
                        msg = str(e)
                    if any(tok in msg for tok in ['-410', 'Query Interrupted', '-420', 'Query UNTERMINATED', 'Query Unterminated']):
                        try:
                            if hasattr(panel.sa, 'sync'):
                                panel.sa.sync()
                        except Exception:
                            pass
                        time.sleep(0.05)
                        continue
                now_time = time.time()
                if now_time - last_report_time >= 2.0:
                    elapsed = now_time - start_time
                    rate_cb(sample_count / elapsed if elapsed > 0 else 0.0)
                    last_report_time = now_time
                time.sleep(0.1)
            try:
                writer_stop.set()
            except Exception:
                pass
            try:
                spec_writer_thread.join(timeout=10)
            except Exception:
                pass

        self._spectrum_thread = threading.Thread(target=worker, daemon=True)
        self._spectrum_thread.start()
